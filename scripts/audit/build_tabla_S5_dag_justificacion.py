"""Tabla S5 — Justificación bibliográfica por arista del DAG.

Construye Post_Auditoria/Suplementario/Tabla_S5_justificacion_DAG.xlsx con:

- Una fila por arista del DAG (visual + especificación .R formal)
- Claim causal en lenguaje natural
- Referencia principal (autor, año, revista, PMID/DOI)
- Fuerza de evidencia (Alta / Media / Baja)
- Notas sobre limitaciones o falta de evidencia específica peruana

Sigue la recomendación de Ferguson et al. 2020 (IJE) y Tennant et al. 2021 (IJE)
de documentar la base teórica/empírica de cada arista en un DAG publicable.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data" / "output" / "Post_Auditoria" / "Suplementario" / "Tabla_S5_justificacion_DAG.xlsx"


FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
FILL_TYPE_BG = {
    "Confusor → Exposición": PatternFill("solid", fgColor="D4EDDA"),
    "Confusor → Desenlace":  PatternFill("solid", fgColor="D4EDDA"),
    "Confusor ↔ Confusor":   PatternFill("solid", fgColor="EAF6E4"),
    "Confusor → Mediador":   PatternFill("solid", fgColor="DFF0D8"),
    "Exposición → Desenlace": PatternFill("solid", fgColor="DCE6F1"),
    "Exposición → Mediador": PatternFill("solid", fgColor="FFE0B2"),
    "Mediador → Desenlace":  PatternFill("solid", fgColor="FFE0B2"),
    "Mediador → Mediador":   PatternFill("solid", fgColor="FFEFD5"),
    "No medido → Exposición": PatternFill("solid", fgColor="F3E5F5"),
    "No medido → Desenlace": PatternFill("solid", fgColor="F3E5F5"),
}
FILL_STRENGTH = {
    "Alta":  PatternFill("solid", fgColor="C8E6C9"),
    "Media": PatternFill("solid", fgColor="FFF9C4"),
    "Baja":  PatternFill("solid", fgColor="FFCCBC"),
}


# (id, tipo, origen, destino, claim, referencia_principal, identificador, fuerza, notas)
EDGES = [
    # ------------------------------------------------------------------------
    # CONFUSORES → EXPOSICIÓN (PHQ-9)
    # ------------------------------------------------------------------------
    ("A1", "Confusor → Exposición", "Edad", "PHQ-9",
     "La edad modifica la prevalencia de síntomas depresivos; en LMIC suele haber mayor depresión en adultos mayores con curva no lineal.",
     "Patel V, Kleinman A. Bull World Health Organ. 2003;81(8):609-15.",
     "PMID: 14576893",
     "Alta",
     "Patrón consistente en revisiones; magnitud específica varía por país."),

    ("A2", "Confusor → Exposición", "Sexo", "PHQ-9",
     "Las mujeres presentan aproximadamente el doble de prevalencia de depresión que los hombres en estudios poblacionales.",
     "Kessler RC, et al. Arch Gen Psychiatry. 1994;51(1):8-19.",
     "PMID: 8279933",
     "Alta",
     "Hallazgo replicado en cientos de estudios; mecanismo multifactorial."),

    ("A3", "Confusor → Exposición", "Educación", "PHQ-9",
     "Menor educación se asocia con mayor prevalencia e incidencia de depresión.",
     "Lorant V, et al. Am J Epidemiol. 2003;157(2):98-112.",
     "PMID: 12522017",
     "Alta",
     "Meta-análisis de cohortes y transversales; gradiente educativo robusto."),

    ("A4", "Confusor → Exposición", "Área urbano/rural", "PHQ-9",
     "Las áreas urbanas suelen mostrar mayor prevalencia de depresión, aunque la dirección puede invertirse en LMIC con privación rural.",
     "Peen J, et al. Acta Psychiatr Scand. 2010;121(2):84-93.",
     "PMID: 19624573",
     "Media",
     "Patrón menos consistente en LMIC; en Perú PMC9021682 (PLOS One 2019) sugiere asociación con altitud."),

    ("A5", "Confusor → Exposición", "Quintil de riqueza", "PHQ-9",
     "La pobreza/menor riqueza incrementa el riesgo de depresión y trastornos mentales comunes.",
     "Lund C, et al. Soc Sci Med. 2010;71(3):517-28.",
     "PMID: 20621748",
     "Alta",
     "Revisión sistemática en LMIC; gradiente socioeconómico documentado."),

    ("A6", "Confusor → Exposición", "Violencia de pareja", "PHQ-9",
     "La violencia de pareja incrementa significativamente el riesgo de depresión incidente y persistente.",
     "Devries KM, et al. PLoS Med. 2013;10(5):e1001439.",
     "PMID: 23671407",
     "Alta",
     "Revisión sistemática de estudios longitudinales; relación causal sustentada."),

    ("A7", "Confusor → Exposición", "Año del estudio", "PHQ-9",
     "Existen tendencias temporales en prevalencia de depresión; el año 2020 mostró aumento global por COVID-19.",
     "Santomauro DF, et al. Lancet. 2021;398(10312):1700-12.",
     "PMID: 34634250",
     "Alta",
     "Impacto COVID-19 cuantificado a nivel global; aplica al sub-muestreo 2020 de ENDES."),

    ("A8", "Confusor → Exposición", "Altitud", "PHQ-9",
     "En Perú, mayor altitud de residencia se asocia con MÁS síntomas depresivos (cada 100 m de ascenso ↑ el puntaje de depresión); mecanismo propuesto: hipoxia hipobárica y su efecto sobre la función cerebral.",
     "Hernández-Vásquez A, et al. J Affect Disord. 2022;299:536-44.",
     "PMID: 34942223",
     "Alta",
     "Confusor geográfico incorporado al conjunto de ajuste durante el análisis (enmienda 2026-06-01). Junto con la arista B8 (altitud → MENOR HTA), las direcciones opuestas inducen la asociación inversa espuria PHQ-9 ↔ PAE que se atenúa al ajustar por altitud."),

    # ------------------------------------------------------------------------
    # CONFUSORES → DESENLACE (PAE)
    # ------------------------------------------------------------------------
    ("B1", "Confusor → Desenlace", "Edad", "PA elevada",
     "La presión arterial aumenta con la edad; la prevalencia de hipertensión crece de forma monotónica desde los 30 años.",
     "Mills KT, et al. Nat Rev Nephrol. 2020;16(4):223-37.",
     "PMID: 32024986",
     "Alta",
     "Hallazgo universal; uno de los efectos cardiovasculares mejor establecidos."),

    ("B2", "Confusor → Desenlace", "Sexo", "PA elevada",
     "Los hombres presentan mayor prevalencia de HTA hasta la sexta década; las mujeres post-menopáusicas pueden igualar o superar.",
     "Mills KT, et al. Nat Rev Nephrol. 2020;16(4):223-37.",
     "PMID: 32024986",
     "Alta",
     "Curva por edad y sexo bien caracterizada."),

    ("B3", "Confusor → Desenlace", "Educación", "PA elevada",
     "Menor educación se asocia con mayor prevalencia e incidencia de hipertensión en países desarrollados y en LMIC.",
     "Leng B, et al. J Hypertens. 2015;33(2):221-9.",
     "PMID: 25479029",
     "Alta",
     "Meta-análisis con 51 estudios; asociación inversa robusta."),

    ("B4", "Confusor → Desenlace", "Área urbano/rural", "PA elevada",
     "En LMIC en transición epidemiológica, las áreas urbanas suelen mostrar mayor prevalencia de HTA por estilos de vida.",
     "Hernández-Hernández R, et al. J Hypertens. 2010;28(1):24-34. (CARMELA study)",
     "PMID: 19809362; DOI: 10.1097/HJH.0b013e328332c353",
     "Media",
     "Patrón menos consistente en cohortes recientes; en Perú la transición rural→urbana puede atenuar el gradiente."),

    ("B5", "Confusor → Desenlace", "Quintil de riqueza", "PA elevada",
     "Relación compleja en LMIC: tradicionalmente mayor SES → más HTA por transición epidemiológica; ahora más mixto.",
     "Mills KT, et al. Nat Rev Nephrol. 2020;16(4):223-37.",
     "PMID: 32024986",
     "Media",
     "Dirección puede variar por contexto; en ENDES queda como confusor a controlar sin asumir signo."),

    ("B6", "Confusor → Desenlace", "Violencia de pareja", "PA elevada",
     "La violencia de pareja se asocia con mayor riesgo de hipertensión incidente en mujeres, posiblemente vía estrés crónico.",
     "Mason SM, et al. Ann Epidemiol. 2012;22(8):562-7.",
     "PMID: 22717307; DOI: 10.1016/j.annepidem.2012.05.003",
     "Media",
     "Asociación documentada pero menos estudiada en LMIC; mecanismo neuroendocrino plausible."),

    ("B7", "Confusor → Desenlace", "Año del estudio", "PA elevada",
     "Tendencias temporales en HTA y, en ENDES 2020, sub-muestreo COVID que pudo alterar la prevalencia medible.",
     "NCD Risk Factor Collaboration. Lancet. 2021;398(10304):957-80.",
     "PMID: 34450083",
     "Alta",
     "Aplica especialmente al año 2020 con restricciones operativas COVID-19 en ENDES."),

    ("B8", "Confusor → Desenlace", "Altitud", "PA elevada",
     "En Perú, mayor altitud se asocia con MENOR prevalencia de hipertensión (relación inversa, marcada ≥ 2 500 m); mecanismos: adaptación crónica a la hipoxia y menor resistencia vascular periférica.",
     "Mendoza-Quispe D, et al. J Hypertens. 2023;41(7):1142-51.",
     "PMID: 37071440",
     "Alta",
     "Confusor geográfico (enmienda 2026-06-01). Las direcciones opuestas de A8 (altitud → +depresión) y B8 (altitud → −HTA) son la base causal de la asociación inversa espuria; al entrar la altitud al ajuste, el efecto de PHQ-9 se atenúa hacia la nulidad."),

    # ------------------------------------------------------------------------
    # CONFUSOR ↔ CONFUSOR (correlaciones graficadas)
    # ------------------------------------------------------------------------
    ("C1", "Confusor ↔ Confusor", "Educación", "Quintil de riqueza",
     "Educación y riqueza están correlacionadas; ambas son dimensiones del estatus socioeconómico.",
     "Adler NE, Newman K. Health Aff (Millwood). 2002;21(2):60-76.",
     "PMID: 11900187",
     "Alta",
     "Correlación estándar en epidemiología social."),

    ("C2", "Confusor ↔ Confusor", "Edad", "Año del estudio",
     "Estructura de cohorte: la distribución etaria puede variar año a año por efectos de muestreo y mortalidad diferencial.",
     "INEI. Manual de la Entrevistadora, ENDES 2024.",
     "https://proyectos.inei.gob.pe/iinei/srienaho/Descarga/DocumentosMetodologicos/2024-5/ManualEntrevistadora.pdf",
     "Media",
     "Documento metodológico oficial INEI; aplicar la misma referencia para años 2019-2024 (manuales por año disponibles en la misma plataforma)."),

    # ------------------------------------------------------------------------
    # EXPOSICIÓN → DESENLACE (efecto total a estimar)
    # ------------------------------------------------------------------------
    ("D1", "Exposición → Desenlace", "PHQ-9", "PA elevada",
     "Hipótesis principal del estudio: los síntomas depresivos se asocian con presión arterial elevada (dirección esperada incierta a priori; literatura previa mixta).",
     "Zhang H, Xu Y, Xu Y. Front Psychiatry. 2024;15:1433990.",
     "PMID: 39355374; DOI: 10.3389/fpsyt.2024.1433990",
     "Media",
     "Paper guía local (Trabajos Guía/fpsyt-15-1433990.pdf). Evidencia previa mixta; este estudio aporta evidencia poblacional peruana 2019-2024."),

    # ------------------------------------------------------------------------
    # EXPOSICIÓN → MEDIADOR
    # ------------------------------------------------------------------------
    ("E1", "Exposición → Mediador", "PHQ-9", "IMC",
     "Asociación bidireccional bien establecida: depresión incrementa el riesgo de obesidad incidente.",
     "Luppino FS, et al. Arch Gen Psychiatry. 2010;67(3):220-9.",
     "PMID: 20194822",
     "Alta",
     "Meta-análisis de 15 estudios longitudinales, n=58 745. RR ~1,58."),

    ("E2", "Exposición → Mediador", "PHQ-9", "Circunferencia abdominal",
     "Síntomas depresivos se asocian con mayor adiposidad abdominal (visceral), independiente del IMC.",
     "Hach I, et al. J Affect Disord. 2006;92(2-3):305-8.",
     "PMID: 16503357; DOI: 10.1016/j.jad.2006.01.023",
     "Media",
     "Menos estudios longitudinales que para IMC; mecanismo plausible vía cortisol y conducta."),

    ("E3", "Exposición → Mediador", "PHQ-9", "Consumo de tabaco (últimos 30 días)",
     "Los síntomas depresivos se asocian con mayor consumo de tabaco (auto-medicación / regulación afectiva); la relación es bidireccional.",
     "Wootton RE, et al. Psychol Med. 2020;50(14):2435-43.",
     "PMID: 31689377; DOI: 10.1017/S0033291719002678",
     "Media",
     "Etiqueta corregida 2026-06-01 (QS201 = tabaco, no alcohol). MR bidireccional: evidencia fuerte tabaco→depresión y más débil depresión→tabaco; mediador plausible."),

    ("E4", "Exposición → Mediador", "PHQ-9", "Consumo problemático de alcohol",
     "Depresión se asocia con mayor riesgo de abuso/dependencia alcohólica (no solo consumo).",
     "Boden JM, Fergusson DM. Addiction. 2011;106(5):906-14.",
     "PMID: 21382111",
     "Alta",
     "Mismo paper que E3; AUDIT-positividad como marcador."),

    ("E5", "Exposición → Mediador", "PHQ-9", "Calidad de dieta",
     "La depresión se asocia con menor calidad de dieta (más alimentos procesados, menos vegetales).",
     "Quirk SE, et al. BMC Psychiatry. 2013;13:175.",
     "PMID: 23802679",
     "Alta",
     "Revisión sistemática; relación bidireccional con énfasis en patrones occidentales."),

    ("E6", "Exposición → Mediador", "PHQ-9", "Diagnóstico de diabetes",
     "La depresión se asocia con mayor riesgo de diabetes tipo 2 incidente (relación bidireccional).",
     "Mezuk B, et al. Diabetes Care. 2008;31(12):2383-90.",
     "PMID: 19033418; DOI: 10.2337/dc08-0985",
     "Alta",
     "Etiqueta corregida 2026-06-01 (QS109 = diagnóstico de diabetes, no medicación antihipertensiva). Meta-análisis: depresión → ~60% mayor riesgo de DM2 incidente."),

    # ------------------------------------------------------------------------
    # MEDIADOR → DESENLACE
    # ------------------------------------------------------------------------
    ("F1", "Mediador → Desenlace", "IMC", "PA elevada",
     "El IMC elevado es uno de los factores de riesgo más potentes para HTA; relación dosis-respuesta establecida.",
     "Mills KT, et al. Nat Rev Nephrol. 2020;16(4):223-37.",
     "PMID: 32024986",
     "Alta",
     "Universalmente aceptado; ~50% del riesgo poblacional de HTA atribuible a sobrepeso/obesidad."),

    ("F2", "Mediador → Desenlace", "Circunferencia abdominal", "PA elevada",
     "Adiposidad abdominal independiente del IMC predice HTA mejor que el IMC en algunas poblaciones.",
     "Janssen I, et al. Arch Intern Med. 2002;162(18):2074-9.",
     "PMID: 12374515",
     "Alta",
     "Magnitud relativa varía por etnia; adipocitokinas y resistencia a insulina como mecanismo."),

    ("F3", "Mediador → Desenlace", "Consumo de tabaco (últimos 30 días)", "PA elevada",
     "El tabaquismo eleva la presión arterial de forma aguda y se asocia con mayor riesgo de hipertensión incidente (efecto más fuerte en fumadores intensos).",
     "Bowman TS, Gaziano JM, Buring JE, Sesso HD. J Am Coll Cardiol. 2007;50(21):2085-92.",
     "PMID: 18021879; DOI: 10.1016/j.jacc.2007.08.017",
     "Media",
     "Etiqueta corregida 2026-06-01 (QS201 = tabaco). Cohorte prospectiva: asociación modesta, más fuerte con ≥15 cig/día; relación tabaco–PA compleja."),

    ("F4", "Mediador → Desenlace", "Consumo problemático de alcohol", "PA elevada",
     "Consumo problemático (binge, dependencia) genera HTA más severa y resistente al tratamiento.",
     "Roerecke M, et al. Lancet Public Health. 2017;2(2):e108-20.",
     "PMID: 29253389; DOI: 10.1016/S2468-2667(17)30003-8",
     "Alta",
     "Mismo meta-análisis; efecto incremental sobre HTA."),

    ("F5", "Mediador → Desenlace", "Calidad de dieta", "PA elevada",
     "Dieta DASH y similares reducen significativamente la presión arterial; sodio y procesados la elevan.",
     "Sacks FM, et al. N Engl J Med. 2001;344(1):3-10.",
     "PMID: 11136953",
     "Alta",
     "Ensayo clínico randomizado; relación causal directa."),

    ("F6", "Mediador → Desenlace", "Diagnóstico de diabetes", "PA elevada",
     "La diabetes mellitus es un predictor independiente de hipertensión incidente y coexiste con PA elevada (resistencia a insulina, rigidez arterial, nefropatía).",
     "Tsimihodimos V, et al. Hypertension. 2018;71(3):422-28.",
     "PMID: 29335249; DOI: 10.1161/HYPERTENSIONAHA.117.10546",
     "Alta",
     "Etiqueta corregida 2026-06-01 (QS109 = diabetes). DM basal predice HTA incidente (OR ~3,1) independiente de edad/IMC/sexo."),

    # ------------------------------------------------------------------------
    # CONFUSOR → MEDIADOR (en visual)
    # ------------------------------------------------------------------------
    ("G1", "Confusor → Mediador", "Edad", "IMC",
     "El IMC se incrementa con la edad hasta los 60-69 años y luego comienza a declinar después de los 70 años (trayectoria invertida en U a lo largo de la vida).",
     "Yang YC, et al. Proc Natl Acad Sci USA. 2021;118(17):e2020167118. (Life-course trajectories of BMI from adolescence to old age)",
     "PMID: 33875595; DOI: 10.1073/pnas.2020167118",
     "Alta",
     "Reemplazo verificado tras la corrección bibliográfica (Heo 2003 invalidado). Cohortes longitudinales con seguimiento desde adolescencia hasta vejez; el paper también discute disparidades por raza y educación útiles para la Discusión."),

    ("G2", "Confusor → Mediador", "Sexo", "Circunferencia abdominal",
     "Dimorfismo sexual claro: los hombres acumulan grasa visceral, las mujeres subcutánea; CC y umbrales de riesgo difieren por sexo.",
     "Lemieux S, et al. Am J Clin Nutr. 1993;58(4):463-7.",
     "PMID: 8379501",
     "Alta",
     "Hallazgo fisiológico bien establecido; informa los puntos de corte ALAD/IDF por sexo."),

    ("G3", "Confusor → Mediador", "Quintil de riqueza", "Calidad de dieta",
     "Mayor pobreza se asocia con dieta de mayor densidad energética y menor calidad nutricional.",
     "Drewnowski A, Specter SE. Am J Clin Nutr. 2004;79(1):6-16.",
     "PMID: 14684391",
     "Alta",
     "Mecanismo: precio relativo de alimentos densos en energía. Confirmado en LMIC."),

    ("G4", "Confusor → Mediador", "Educación", "Calidad de dieta",
     "Mayor educación se asocia con mejor calidad de dieta vía alfabetización nutricional y autoeficacia.",
     "Lallukka T, et al. Eur J Clin Nutr. 2007;61(6):701-10.",
     "PMID: 17180154; DOI: 10.1038/sj.ejcn.1602583",
     "Alta",
     "Efecto independiente del nivel de ingresos."),

    ("G5", "Confusor → Mediador", "Edad", "Diagnóstico de diabetes",
     "La prevalencia de diabetes tipo 2 aumenta marcadamente con la edad, con pico entre los 65-79 años.",
     "Saeedi P, et al. Diabetes Res Clin Pract. 2019;157:107843. (IDF Diabetes Atlas, 9.ª ed.)",
     "PMID: 31518657; DOI: 10.1016/j.diabres.2019.107843",
     "Alta",
     "Etiqueta corregida 2026-06-01 (QS109 = diabetes). Gradiente etario de prevalencia de DM bien establecido a nivel global."),

    # ------------------------------------------------------------------------
    # CONFUSOR → MEDIADOR (en .R formal, omitidas del visual por legibilidad)
    # ------------------------------------------------------------------------
    ("G6", "Confusor → Mediador", "Sexo", "IMC",
     "Diferencias por sexo en el IMC medio y en la prevalencia de obesidad por categorías.",
     "NCD Risk Factor Collaboration. Lancet. 2016;387(10026):1377-96.",
     "PMID: 27115820",
     "Alta",
     "Omitida del visual por legibilidad; presente en 02_DAG_modelos.R."),

    ("G7", "Confusor → Mediador", "Área urbano/rural", "Calidad de dieta",
     "Áreas urbanas tienen mayor acceso a alimentos procesados; rurales pueden consumir más alimentos tradicionales.",
     "Popkin BM. Nutr Rev. 2017;75(2):73-82. (Relationship between shifts in food system dynamics and acceleration of the global nutrition transition)",
     "PMID: 28395033; DOI: 10.1093/nutrit/nuw064",
     "Media",
     "Transición nutricional en LMIC; dirección puede variar por país. Omitida del visual."),

    ("G8", "Confusor → Mediador", "Quintil de riqueza", "Consumo de tabaco (últimos 30 días)",
     "El tabaquismo sigue un gradiente socioeconómico: en general, menor nivel socioeconómico se asocia con mayor prevalencia de consumo de tabaco.",
     "Hiscock R, et al. Ann N Y Acad Sci. 2012;1248:107-23. (Socioeconomic status and smoking: a review)",
     "PMID: 22092035; DOI: 10.1111/j.1749-6632.2011.06202.x",
     "Alta",
     "Etiqueta corregida 2026-06-01 (QS201 = tabaco, no alcohol). Gradiente socioeconómico del tabaquismo robusto; en LMIC la dirección puede variar por contexto. Omitida del visual."),

    # ------------------------------------------------------------------------
    # MEDIADOR → MEDIADOR (en .R formal, omitidas del visual)
    # ------------------------------------------------------------------------
    ("H1", "Mediador → Mediador", "Consumo de tabaco (últimos 30 días)", "Consumo problemático de alcohol",
     "El consumo de tabaco y el consumo problemático de alcohol co-ocurren con fuerza (uso concurrente de sustancias / responsabilidad compartida); se grafican como mediadores correlacionados.",
     "Falk DE, Yi HY, Hiller-Sturmhöfel S. Alcohol Res Health. 2006;29(3):162-71.",
     "PMID: 17373404",
     "Alta",
     "Arista reconceptuada 2026-06-01 tras la corrección de etiqueta (QS201 = tabaco): de 'frecuencia→problemático de alcohol' a co-ocurrencia tabaco–alcohol. Correlación, no causalidad direccional. Omitida del visual."),

    ("H2", "Mediador → Mediador", "Calidad de dieta", "IMC",
     "Una dieta de mayor calidad reduce el balance energético positivo y previene aumento de IMC.",
     "Mozaffarian D, et al. N Engl J Med. 2011;364(25):2392-404.",
     "PMID: 21696306",
     "Alta",
     "Cohorte longitudinal; efecto de componentes específicos de dieta sobre cambios de peso. Omitida del visual."),

    ("H3", "Mediador → Mediador", "IMC", "Circunferencia abdominal",
     "IMC y CC están altamente correlacionados; el IMC predice parcialmente la CC pero no completamente.",
     "Janssen I, et al. Am J Clin Nutr. 2004;79(3):379-84.",
     "PMID: 14985210",
     "Alta",
     "Correlación r > 0,7 en la mayoría de poblaciones. Omitida del visual."),

    # ------------------------------------------------------------------------
    # NO MEDIDO (U) → EXPOSICIÓN y DESENLACE
    # ------------------------------------------------------------------------
    ("U1", "No medido → Exposición", "Estrés crónico", "PHQ-9",
     "El estrés crónico es el predictor proximal más fuerte de síntomas depresivos incidentes.",
     "Hammen C. Annu Rev Clin Psychol. 2005;1:293-319.",
     "PMID: 17716090",
     "Alta",
     "Variable proxy posible vía VPAR y SES; confusión residual probable."),

    ("U2", "No medido → Desenlace", "Estrés crónico", "PA elevada",
     "El estrés crónico eleva la PA vía activación simpática y eje hipotálamo-pituitario-adrenal.",
     "Cohen S, et al. JAMA. 2007;298(14):1685-7.",
     "PMID: 17925521",
     "Alta",
     "Mecanismo neuroendocrino. Confusión residual con depresión vía estrés común."),

    ("U3", "No medido → Exposición", "Calidad de sueño", "PHQ-9",
     "Insomnio y mala calidad del sueño predicen depresión incidente.",
     "Baglioni C, et al. J Affect Disord. 2011;135(1-3):10-9.",
     "PMID: 21300408",
     "Alta",
     "Meta-análisis longitudinal; OR ~2 para depresión incidente."),

    ("U4", "No medido → Desenlace", "Apnea del sueño / sueño corto", "PA elevada",
     "Sueño corto y apnea obstructiva incrementan la PA y el riesgo de HTA.",
     "Calhoun DA, Harding SM. Chest. 2010;138(2):434-43.",
     "PMID: 20682533",
     "Alta",
     "Mecanismo bien establecido para apnea; sueño corto con efecto menor pero significativo."),

    ("U5", "No medido → Desenlace", "Comorbilidad CV no medida (dislipidemia, ERC, enf. coronaria)", "PA elevada",
     "La dislipidemia, la enfermedad renal crónica y la enfermedad coronaria coexisten con HTA y modifican su prevalencia y control.",
     "Whelton PK, et al. Hypertension. 2018;71(6):e13-e115. (Guía ACC/AHA HTA)",
     "PMID: 29133356; DOI: 10.1161/HYP.0000000000000065",
     "Alta",
     "Actualizado 2026-06-01: la diabetes ya está medida (QS109) y entra como mediador (E6/F6/G5); dislipidemia, ERC y enf. coronaria siguen no medidas → confusión residual probable."),

    ("U6", "No medido → Exposición y Desenlace", "Sesgo de detección/cuidado", "PHQ-9 y PA elevada",
     "Las personas diagnosticadas con HTA pueden reportar más síntomas depresivos (estrés del manejo) y tener PA controlada por tratamiento — explica asociación inversa observada en datos transversales.",
     "Ferguson KD, et al. Int J Epidemiol. 2020;49(1):322-9. (ESC-DAGs: evidence synthesis for DAGs)",
     "PMID: 31325312; DOI: 10.1093/ije/dyz150",
     "Media",
     "Hipótesis explicativa del hallazgo del estudio; debería explicitarse en Discusión."),
]


def main() -> None:
    wb = Workbook(); wb.remove(wb.active)

    # ---------------- Portada ----------------------------------------------
    ws = wb.create_sheet("Portada")
    ws["A1"] = "Tabla S5 — Justificación bibliográfica de las aristas del DAG"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = ("Fecha original: 2026-05-29 · Correcciones bibliográficas: 2026-05-30 · "
                "Corrección de etiquetas QS201 (tabaco) / QS109 (diabetes) y re-fundamentación de aristas E3, E6, F3, F6, G5, G8, H1, U5: 2026-06-01")
    ws["A4"] = "Propósito"
    ws["A4"].font = Font(bold=True)
    propósito = [
        "  Documentar la base teórica o empírica de cada arista del DAG (Figura S2).",
        "  Sigue las recomendaciones de Tennant et al. 2021 (IJE) y Ferguson et al. 2020 (IJE)",
        "  para DAGs publicables: cada arista debe tener justificación citable.",
    ]
    for i, line in enumerate(propósito, start=5):
        ws.cell(row=i, column=1, value=line)

    ws["A9"] = "Cómo se construyó esta tabla"
    ws["A9"].font = Font(bold=True)
    cons = [
        "  - Inventario de aristas tomado del archivo 02_DAG_modelos.R (especificación dagitty formal),",
        "    incluyendo aristas omitidas del visual por legibilidad (marcadas en columna 'Notas').",
        "  - Para cada arista se identificó: claim causal, referencia principal (PMID/DOI), fuerza de evidencia.",
        "  - Las correlaciones entre confusores y las aristas con U (no medidos) se incluyen explícitamente.",
        "  - Fuerza de evidencia:",
        "      Alta:  Evidencia de meta-análisis, ECA o múltiples cohortes longitudinales.",
        "      Media: Evidencia de algunas cohortes o transversales; mecanismo plausible.",
        "      Baja:  Solo razonamiento teórico o evidencia indirecta.",
    ]
    for i, line in enumerate(cons, start=10):
        ws.cell(row=i, column=1, value=line)

    ws["A19"] = "Limitaciones de esta tabla"
    ws["A19"].font = Font(bold=True)
    limits = [
        "  - La búsqueda bibliográfica NO es exhaustiva; representa una selección de literatura central.",
        "  - Para evidencia poblacional específica de Perú, varias citas son extrapolaciones de literatura internacional.",
        "  - Idealmente, la siguiente iteración co-construye el DAG con expertos clínicos peruanos",
        "    (Ferguson et al. 2020 IJE), documentando consenso en una sesión formal.",
        "  - Las citas marcadas con 'Verificar:' requieren consulta directa al paper original antes de citar en el manuscrito.",
    ]
    for i, line in enumerate(limits, start=20):
        ws.cell(row=i, column=1, value=line)

    ws["A26"] = "Resumen por tipo de arista"
    ws["A26"].font = Font(bold=True)
    # Contar por tipo
    from collections import Counter
    counter = Counter(e[1] for e in EDGES)
    row = 27
    for tipo, n in counter.most_common():
        ws.cell(row=row, column=1, value=f"  {tipo}: {n} aristas")
        row += 1

    ws.column_dimensions["A"].width = 110

    # ---------------- Hoja principal: Justificacion_aristas ----------------
    ws = wb.create_sheet("Justificacion_aristas")

    headers = ["ID", "Tipo", "Origen", "Destino", "Claim causal",
               "Referencia principal", "PMID / DOI", "Fuerza", "Notas"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = FILL_HEADER
    ws.row_dimensions[1].height = 32

    for ridx, (eid, tipo, ori, dst, claim, ref, ident, strength, notas) in enumerate(EDGES, start=2):
        cells = [
            (1, eid, "center"),
            (2, tipo, "left"),
            (3, ori, "left"),
            (4, dst, "left"),
            (5, claim, "left"),
            (6, ref, "left"),
            (7, ident, "left"),
            (8, strength, "center"),
            (9, notas, "left"),
        ]
        for col, val, align in cells:
            c = ws.cell(row=ridx, column=col, value=val)
            c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
        # Color por tipo (col Tipo)
        if tipo in FILL_TYPE_BG:
            ws.cell(row=ridx, column=2).fill = FILL_TYPE_BG[tipo]
        # Color por fuerza
        if strength in FILL_STRENGTH:
            ws.cell(row=ridx, column=8).fill = FILL_STRENGTH[strength]
        ws.row_dimensions[ridx].height = 56

    widths = {1: 6, 2: 22, 3: 22, 4: 22, 5: 45, 6: 38, 7: 22, 8: 9, 9: 38}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # ---------------- Hoja de referencias consolidadas ---------------------
    ws = wb.create_sheet("Referencias")
    ws["A1"] = "Referencias consolidadas (orden alfabético por primer autor)"
    ws["A1"].font = Font(bold=True, size=12)

    refs = sorted(set((e[5], e[6]) for e in EDGES), key=lambda x: x[0].split(",")[0])
    headers = ["Referencia", "PMID / DOI"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True); c.fill = FILL_HEADER
    for ridx, (ref, ident) in enumerate(refs, start=4):
        ws.cell(row=ridx, column=1, value=ref).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=ridx, column=2, value=ident).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 26

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"  -> {OUT.relative_to(ROOT)}")
    print(f"  {len(EDGES)} aristas documentadas")
    print(f"  {len(refs)} referencias únicas consolidadas")


if __name__ == "__main__":
    main()
