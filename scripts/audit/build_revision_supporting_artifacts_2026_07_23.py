"""Construye los artefactos suplementarios exigidos por la guía del 23-07-2026.

Fuentes de verdad:
  - base observada e imputaciones congeladas de ``data/output_2025``;
  - salidas corregidas del panel D1 en ``analysis/models``;
  - validación independiente con ``mitml``.

El script conserva copias ``.pre_guia_2026-07-23.xlsx`` de S2, S3 y S7 antes
de modificarlos.
"""

from __future__ import annotations

import math
import shutil
from copy import copy
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
ANALYSIS = ROOT / "data" / "output_2025" / "analysis"
MODELS = ANALYSIS / "models"
TABLES = ANALYSIS / "tables"
IMPUTED = ANALYSIS / "imputed"
OBSERVED = ROOT / "data" / "output_2025" / "endes_hta_depresion_2019_2025.parquet"
SUBMISSION = ROOT / "ENVIO_PLOS_ONE_2019-2025"

S2 = SUBMISSION / "S2_Table.xlsx"
S3 = SUBMISSION / "S3_Table.xlsx"
S7 = SUBMISSION / "S7_Table.xlsx"

D1_VALIDATION = MODELS / "effect_modification_d1_mitml_validation.csv"
D1_PANEL = MODELS / "effect_modification_panel.csv"
SEX_SLOPES = MODELS / "effect_modification_sex_stratified.csv"
MODEL_DIAGNOSTICS = MODELS / "model_diagnostics_summary.csv"
DISPERSION = (
    ROOT
    / "data"
    / "output_2025"
    / "comparacion_2019_2025"
    / "dispersion_model2_H7.csv"
)
MI_DIAGNOSTICS = ANALYSIS / "mice_observed_imputed_diagnostics.csv"


VARIABLES = {
    "QS23": {"label": "Age", "kind": "continuous", "allowed": None},
    "QSSEXO": {"label": "Sex", "kind": "categorical", "allowed": [1, 2]},
    "HV025": {"label": "Area of residence", "kind": "categorical", "allowed": [1, 2]},
    "HV270": {"label": "Wealth quintile", "kind": "categorical", "allowed": [1, 2, 3, 4, 5]},
    "VIOLENCIA_PAREJA": {
        "label": "Intimate partner violence",
        "kind": "categorical",
        "allowed": [0, 1, 2],
    },
    "IMC": {"label": "Body mass index", "kind": "continuous", "allowed": None},
    "QS907": {"label": "Waist circumference", "kind": "continuous", "allowed": None},
    "CALIDAD_DIETA": {"label": "Diet quality", "kind": "categorical", "allowed": [0, 1]},
    "QS109": {"label": "Diabetes diagnosis", "kind": "categorical", "allowed": [1, 2]},
}

LABELS = {
    ("VIOLENCIA_PAREJA", 0): "no violence",
    ("VIOLENCIA_PAREJA", 1): "physical violence",
    ("VIOLENCIA_PAREJA", 2): "no partner",
    ("CALIDAD_DIETA", 0): "inadequate",
    ("CALIDAD_DIETA", 1): "adequate",
    ("QS109", 1): "yes",
    ("QS109", 2): "no",
    ("QSSEXO", 1): "men",
    ("QSSEXO", 2): "women",
    ("HV025", 1): "urban",
    ("HV025", 2): "rural",
}

MODIFIER_LABELS = {
    "sex": "Sex",
    "year": "Year",
    "area": "Area",
    "wealth": "Wealth quintile",
    "dxhta": "Prior HTN diagnosis",
    "altitude": "Altitude",
}

MODEL_LABELS = {
    "interaction_altitud": "Interaction PHQ-9 × Altitude",
    "interaction_area": "Interaction PHQ-9 × Area",
    "interaction_dxhta": "Interaction PHQ-9 × Prior HTN dx",
    "interaction_riqueza": "Interaction PHQ-9 × Wealth",
    "interaction_sex": "Interaction PHQ-9 × Sex",
    "interaction_year": "Interaction PHQ-9 × Year",
    "model_1": "Model 1 — Crude",
    "model_2": "Model 2 — Structural (main)",
    "model_3": "Model 3 — Exploratory",
    "sensitivity_no_2020": "Sensitivity: without 2020",
    "sensitivity_outcome_130_80": "Sensitivity: outcome ≥130/80",
    "sensitivity_second_bp_measure": "Sensitivity: second BP measure",
    "submodel_adherence": "Cascade: non-adherence",
    "submodel_domain_bp": "Cascade: uncontrolled BP",
}


def _require(paths: list[Path]) -> None:
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError("Faltan artefactos requeridos:\n" + "\n".join(missing))


def _normalize_effect_panel_roles() -> None:
    """Migra la etiqueta de rol del proceso ya iniciado antes del cambio terminológico."""
    panel = pd.read_csv(D1_PANEL)
    panel["modificador"] = panel["modificador"].replace(
        {"riqueza": "wealth", "altitud": "altitude"}
    )
    expected = {"sex", "year", "area", "wealth", "dxhta", "altitude"}
    if set(panel["modificador"]) != expected:
        raise RuntimeError("El panel D1 no contiene exactamente los seis modificadores esperados.")
    panel.loc[panel["modificador"].isin(["sex", "year"]), "tipo"] = "informado_por_teoria"
    panel.loc[
        panel["modificador"].isin(["area", "wealth", "dxhta", "altitude"]),
        "tipo",
    ] = "exploratorio"
    panel.to_csv(D1_PANEL, index=False, encoding="utf-8-sig")


def _backup(path: Path) -> Path:
    backup = path.with_name(f"{path.stem}.pre_guia_2026-07-23{path.suffix}")
    if not backup.exists():
        shutil.copy2(path, backup)
    return backup


def _safe_values(series: pd.Series) -> np.ndarray:
    return pd.to_numeric(series, errors="coerce").dropna().to_numpy(dtype=float)


def _build_mi_diagnostics() -> pd.DataFrame:
    observed = pd.read_parquet(OBSERVED, columns=list(VARIABLES))
    missing_masks = {var: observed[var].isna().to_numpy() for var in VARIABLES}
    per_imputation: dict[tuple[str, str], list[float]] = {}
    pooled_imputed: dict[str, list[np.ndarray]] = {var: [] for var in VARIABLES}

    paths = sorted(IMPUTED.glob("imputation_*.parquet"))
    if len(paths) != 20:
        raise RuntimeError(f"Se esperaban 20 imputaciones y se encontraron {len(paths)}.")

    for path in paths:
        frame = pd.read_parquet(path, columns=list(VARIABLES))
        for var, spec in VARIABLES.items():
            mask = missing_masks[var]
            values = _safe_values(frame.loc[mask, var])
            if not mask.any():
                if values.size:
                    raise AssertionError(f"{var}: aparecieron valores imputados sin faltantes originales.")
                continue
            if values.size != int(mask.sum()):
                raise AssertionError(f"{var}: quedaron faltantes en {path.name}.")
            pooled_imputed[var].append(values)
            if spec["kind"] == "continuous":
                per_imputation.setdefault((var, "mean"), []).append(float(np.mean(values)))
                per_imputation.setdefault((var, "sd"), []).append(float(np.std(values, ddof=1)))
            else:
                allowed = np.asarray(spec["allowed"], dtype=float)
                if not np.isin(values, allowed).all():
                    bad = np.unique(values[~np.isin(values, allowed)])
                    raise AssertionError(f"{var}: categorías imputadas inválidas: {bad.tolist()}")
                for level in allowed:
                    per_imputation.setdefault((var, f"proportion_{level:g}"), []).append(
                        float(np.mean(values == level))
                    )

    rows: list[dict[str, object]] = []
    for var, spec in VARIABLES.items():
        obs_values = _safe_values(observed[var])
        missing_n = int(missing_masks[var].sum())
        if missing_n == 0:
            rows.append(
                {
                    "variable": var,
                    "label": spec["label"],
                    "metric": "no_missing_values",
                    "level": "",
                    "n_observed": int(obs_values.size),
                    "n_missing_before": 0,
                    "observed_value": np.nan,
                    "imputed_pooled_value": np.nan,
                    "between_imputation_min": np.nan,
                    "between_imputation_max": np.nan,
                    "restriction_check": "not imputed; target was complete",
                }
            )
            continue

        imputed_values = np.concatenate(pooled_imputed[var])
        if spec["kind"] == "continuous":
            observed_min = float(np.min(obs_values))
            observed_max = float(np.max(obs_values))
            if np.min(imputed_values) < observed_min or np.max(imputed_values) > observed_max:
                raise AssertionError(f"{var}: una imputación quedó fuera del rango observado.")
            for metric, obs_metric, imp_metric in [
                ("mean", float(np.mean(obs_values)), float(np.mean(imputed_values))),
                ("sd", float(np.std(obs_values, ddof=1)), float(np.std(imputed_values, ddof=1))),
            ]:
                stability = per_imputation[(var, metric)]
                rows.append(
                    {
                        "variable": var,
                        "label": spec["label"],
                        "metric": metric,
                        "level": "",
                        "n_observed": int(obs_values.size),
                        "n_missing_before": missing_n,
                        "observed_value": obs_metric,
                        "imputed_pooled_value": imp_metric,
                        "between_imputation_min": min(stability),
                        "between_imputation_max": max(stability),
                        "restriction_check": (
                            f"within observed range [{observed_min:.3f}, {observed_max:.3f}]"
                        ),
                    }
                )
        else:
            for level in spec["allowed"]:
                level_float = float(level)
                metric = f"proportion_{level_float:g}"
                stability = per_imputation[(var, metric)]
                rows.append(
                    {
                        "variable": var,
                        "label": spec["label"],
                        "metric": "proportion",
                        "level": LABELS.get((var, level), str(level)),
                        "n_observed": int(obs_values.size),
                        "n_missing_before": missing_n,
                        "observed_value": float(np.mean(obs_values == level_float)),
                        "imputed_pooled_value": float(np.mean(imputed_values == level_float)),
                        "between_imputation_min": min(stability),
                        "between_imputation_max": max(stability),
                        "restriction_check": "valid observed category",
                    }
                )

    result = pd.DataFrame(rows)
    result.to_csv(MI_DIAGNOSTICS, index=False, encoding="utf-8-sig")
    return result


def _copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)


def _style_section(ws, row: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = Font(bold=True)


def _style_header(ws, row: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor="EDEDED")
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _clear_merged_ranges_intersecting(ws, min_row: int, max_row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if not (merged.max_row < min_row or merged.min_row > max_row):
            ws.unmerge_cells(str(merged))


def _format_p(value: float) -> str:
    if not math.isfinite(value):
        return "—"
    return "< 0.001" if value < 0.001 else f"{value:.3f}"


def _update_s2(mi: pd.DataFrame) -> None:
    source = _backup(S2)
    wb = load_workbook(source)
    ws = wb["S2 Table"]
    ws.cell(
        1,
        1,
        (
            "S2 Table. Design effects (DEFF) and model diagnostics "
            "(ENDES 2019-2025, 20 parametric chained-equations imputations)."
        ),
    )
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str):
                value = value.replace(
                    "across the 20 MICE imputations",
                    "across the 20 parametric chained-equations imputations",
                )
                value = value.replace(
                    "first MICE imputation",
                    "first completed dataset",
                )
                ws.cell(row, col, value)

    diagnostics = pd.read_csv(MODEL_DIAGNOSTICS)
    diag_by_model = diagnostics.set_index("model")
    row_by_label = {
        str(ws.cell(row, 1).value): row
        for row in range(10, 24)
        if ws.cell(row, 1).value
    }
    for model, label in MODEL_LABELS.items():
        if model not in diag_by_model.index or label not in row_by_label:
            continue
        row = row_by_label[label]
        values = diag_by_model.loc[model]
        ws.cell(row, 2, round(float(values["mean_df_resid"])))
        ws.cell(row, 3, round(float(values["median_scale_deviance"]), 2))
        ws.cell(row, 4, str(int(values["imputations"])))

    dispersion = pd.read_csv(DISPERSION)
    phi = float(
        dispersion.loc[
            dispersion["cohorte"].astype(str).str.contains("2019-2025"),
            "dispersion_svyglm_design",
        ].iloc[0]
    )
    ws.cell(24, 1, "Design-based quasi-Poisson dispersion (φ), Model 2")
    ws.cell(24, 2, round(phi, 4))
    ws.cell(24, 3, "Mild underdispersion")
    ws.cell(24, 4, "20")
    for col in range(1, 5):
        _copy_cell_style(ws.cell(23, col), ws.cell(24, col))

    base_row = ws.max_row + 2
    ws.cell(base_row, 1, "D. Multiple-imputation diagnostics")
    _style_section(ws, base_row, 4)
    ws.merge_cells(start_row=base_row, start_column=1, end_row=base_row, end_column=4)
    base_row += 1
    headers = [
        "Target / metric",
        "Observed",
        "Imputed-only pooled",
        "Between-imputation range",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(base_row, col, header)
    _style_header(ws, base_row, 4)
    base_row += 1

    report = mi.loc[
        (mi["n_missing_before"] > 0)
        & (
            ((mi["metric"] == "mean") & mi["variable"].isin(["IMC", "QS907"]))
            | (
                (mi["metric"] == "proportion")
                & mi["variable"].isin(["VIOLENCIA_PAREJA", "CALIDAD_DIETA", "QS109"])
            )
        )
    ]
    for _, item in report.iterrows():
        level = f" — {item['level']}" if str(item["level"]) else ""
        ws.cell(base_row, 1, f"{item['label']}{level} ({item['metric']})")
        percentage = item["metric"] == "proportion"
        if percentage:
            ws.cell(base_row, 2, f"{100 * item['observed_value']:.1f}%")
            ws.cell(base_row, 3, f"{100 * item['imputed_pooled_value']:.1f}%")
            ws.cell(
                base_row,
                4,
                (
                    f"{100 * item['between_imputation_min']:.1f}%–"
                    f"{100 * item['between_imputation_max']:.1f}%"
                ),
            )
        else:
            ws.cell(base_row, 2, f"{item['observed_value']:.2f}")
            ws.cell(base_row, 3, f"{item['imputed_pooled_value']:.2f}")
            ws.cell(
                base_row,
                4,
                (
                    f"{item['between_imputation_min']:.2f}–"
                    f"{item['between_imputation_max']:.2f}"
                ),
            )
        base_row += 1

    notes = [
        (
            "D. Diagnostics compare observed values with the values imputed only at originally "
            "missing positions. All continuous draws remained inside the observed range; categorical "
            "draws were reassigned to valid observed categories."
        ),
        (
            "The between-imputation range summarizes stability across 20 posterior-sampling draws "
            "with distinct seeds. IterativeImputer returns completed draws and does not retain a "
            "chain-iteration convergence trace; no such trace was available for retrospective review."
        ),
        (
            "m = 20 was chosen to reduce Monte Carlo error and is conservative relative to the "
            "largest observed target-specific missing proportion (diet quality, 4.4%). "
            f"Exact machine-readable diagnostics: {MI_DIAGNOSTICS.relative_to(ROOT).as_posix()}."
        ),
        (
            f"B. The design-based dispersion parameter for Model 2 was φ = {phi:.4f}, indicating "
            "mild underdispersion. Median scaled deviance is a separate descriptive statistic."
        ),
    ]
    for note in notes:
        ws.cell(base_row, 1, note)
        ws.cell(base_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=base_row, start_column=1, end_row=base_row, end_column=4)
        base_row += 1

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 55)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 26
    wb.save(S2)


def _update_s3() -> None:
    source = _backup(S3)
    panel = pd.read_csv(D1_PANEL)
    validation = pd.read_csv(D1_VALIDATION)
    slopes = pd.read_csv(SEX_SLOPES)

    if "match_1e_10" not in validation.columns:
        raise RuntimeError("La validación mitml no contiene la bandera match_1e_10.")
    matches = validation["match_1e_10"].map(
        lambda value: str(value).strip().lower() == "true"
    )
    if not matches.all():
        raise RuntimeError("La validación D1 con mitml no coincide con la implementación Python.")

    wb = load_workbook(source)
    ws = wb["S3 Table"]
    _clear_merged_ranges_intersecting(ws, 11, ws.max_row + 30)
    for row in range(11, ws.max_row + 1):
        for col in range(1, 7):
            ws.cell(row, col).value = None

    ws.cell(11, 1, "B. Effect modification (PHQ-9 × modifier; pooled D1 test)")
    ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=6)
    _style_section(ws, 11, 6)
    headers = ["Modifier", "Analysis role", "df1", "df2", "D1 F", "p / Holm p"]
    for col, header in enumerate(headers, 1):
        ws.cell(12, col, header)
    _style_header(ws, 12, 6)

    for row, (_, item) in enumerate(panel.iterrows(), start=13):
        modifier = str(item["modificador"])
        ws.cell(row, 1, MODIFIER_LABELS.get(modifier, modifier))
        ws.cell(
            row,
            2,
            "theory-informed" if modifier in {"sex", "year"} else "exploratory",
        )
        ws.cell(row, 3, int(item["df_num"]))
        ws.cell(row, 4, f"{float(item['df_den']):.1f}")
        ws.cell(row, 5, f"{float(item['statistic']):.2f}")
        p_text = _format_p(float(item["p_value"]))
        if str(item["tipo"]) == "exploratorio":
            holm = float(item["p_holm_exploratorios"])
            p_text += f" / Holm {_format_p(holm)}"
        ws.cell(row, 6, p_text)

    section_c = 13 + len(panel) + 1
    ws.cell(section_c, 1, "C. Sex-specific PHQ-9 slopes from the interaction model")
    ws.merge_cells(start_row=section_c, start_column=1, end_row=section_c, end_column=6)
    _style_section(ws, section_c, 6)
    headers_c = ["Sex", "PR per PHQ-9 point", "95% CI", "p", "Imputations", "Model n"]
    for col, header in enumerate(headers_c, 1):
        ws.cell(section_c + 1, col, header)
    _style_header(ws, section_c + 1, 6)
    for row, (_, item) in enumerate(slopes.iterrows(), start=section_c + 2):
        ws.cell(row, 1, str(item["sex"]).capitalize())
        ws.cell(row, 2, f"{float(item['pr']):.3f}")
        ws.cell(row, 3, f"{float(item['ci_low']):.3f} to {float(item['ci_high']):.3f}")
        ws.cell(row, 4, _format_p(float(item["p_value"])))
        ws.cell(row, 5, int(item["imputations_used"]))
        ws.cell(row, 6, f"{float(item['mean_n_obs']):,.0f}")

    notes_row = section_c + 2 + len(slopes) + 1
    notes = [
        (
            "Notes: PR = prevalence ratio (quasi-Poisson, log link). Scalar estimates and "
            "sex-specific linear combinations were pooled with Rubin's rules over 20 imputations."
        ),
        (
            "Joint interaction tests used the D1 Li–Rubin F procedure with finite denominator "
            "degrees of freedom. The Python implementation was reproduced from the exact coefficient "
            "and covariance inputs with mitml 0.4-5 (R 4.5.3); all F, df2, relative-increase-in-"
            "variance and p values agreed within the stated numerical tolerances."
        ),
        (
            "Holm correction was applied only to the four exploratory modifiers (area, wealth, "
            "prior hypertension diagnosis and altitude). Records with missing prior-diagnosis status "
            "were excluded from that binary one-df interaction."
        ),
        (
            "Spline: four knots at PHQ-9 = 0, 4, 9 and 14; the non-linearity p value (0.098) "
            "was combined using D2, not a median p value."
        ),
    ]
    for note in notes:
        ws.cell(notes_row, 1, note)
        ws.cell(notes_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=6)
        notes_row += 1

    for col, width in enumerate([28, 22, 14, 16, 14, 22], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    wb.save(S3)


def _update_s7() -> None:
    source = _backup(S7)
    panel = pd.read_csv(D1_PANEL)
    altitude = panel.loc[panel["modificador"] == "altitude"].iloc[0]
    wb = load_workbook(source)
    ws = wb["S7 Table"]
    _clear_merged_ranges_intersecting(ws, 15, ws.max_row + 5)

    ws.cell(15, 1, "C. Effect modification by altitude (pooled D1 test)")
    ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=4)
    _style_section(ws, 15, 4)
    for col, value in enumerate(["df1 / df2", "D1 F", "p", "Holm p"], start=1):
        ws.cell(16, col, value)
    _style_header(ws, 16, 4)
    ws.cell(17, 1, f"{int(altitude['df_num'])} / {float(altitude['df_den']):.1f}")
    ws.cell(17, 2, f"{float(altitude['statistic']):.2f}")
    ws.cell(17, 3, _format_p(float(altitude["p_value"])))
    ws.cell(17, 4, _format_p(float(altitude["p_holm_exploratorios"])))
    note = (
        "Notes: PR = prevalence ratio (quasi-Poisson, log link); scalar estimates were pooled "
        "over 20 imputations with Rubin's rules and the joint altitude interaction used D1. "
        "The change after adding altitude is descriptive attenuation and does not by itself prove "
        "causal confounding. Within each altitude stratum, Model 2 omits the altitude term. "
        "Altitude = HV040 (complete; not imputed). Source: ENDES 2019-2025 (INEI)."
    )
    ws.cell(19, 1, note)
    ws.cell(19, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=19, start_column=1, end_row=19, end_column=4)
    wb.save(S7)


def main() -> None:
    _require(
        [
            OBSERVED,
            D1_PANEL,
            D1_VALIDATION,
            SEX_SLOPES,
            MODEL_DIAGNOSTICS,
            DISPERSION,
            S2,
            S3,
            S7,
        ]
    )
    _normalize_effect_panel_roles()
    mi = _build_mi_diagnostics()
    _update_s2(mi)
    _update_s3()
    _update_s7()
    print(f"Diagnósticos MI: {MI_DIAGNOSTICS.relative_to(ROOT)}")
    print(f"Actualizados: {S2.name}, {S3.name}, {S7.name}")


if __name__ == "__main__":
    main()
