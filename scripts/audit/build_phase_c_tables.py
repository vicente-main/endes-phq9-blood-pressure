"""Fase C — Producción de tablas v2 autorizadas por la matriz B.

Cubre:
- C1: Tabla_1_v2.xlsx        (B04 regenerar_v2)
- C2: Tabla_3_v2.xlsx        (B06 regenerar_v2)
- C3: Tabla_4_v2.xlsx        (B07 regenerar_v2)
- C4: Tabla_S2_v2.xlsx       (B10 regenerar_v2)
- C5: Tabla_S3_v2.xlsx       (B11 regenerar_v2)
- C6: Tabla_5_v2.xlsx        (B08 renombrar — None → '—')
- C7: Tabla_S4_v2.xlsx       (B12 renombrar — encabezado 'variable')

Estrategia: cargar los xlsx fuente desde Para Publicar Redisenado/{Principal,Suplementario}/,
aplicar reemplazos celda por celda con LABEL_MAP, y guardar en Auditoria_Integral/.

Cada salida lleva un encabezado de metadato en la primera hoja (Portada) que cita la
fila B0X que la autoriza.
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


ROOT = Path(__file__).resolve().parents[2]
SRC_PRINCIPAL = ROOT / "data" / "output" / "Para Enviar" / "Para Publicar Redisenado" / "Principal"
SRC_SUPL = ROOT / "data" / "output" / "Para Enviar" / "Para Publicar Redisenado" / "Suplementario"
OUT = ROOT / "data" / "output" / "Auditoria_Integral"


# ---------------------------------------------------------------------------
# Mapeo canónico de etiquetas (definido en 04_tablas_auditoria.xlsx hoja
# Nomenclatura, materializado aquí para reproducibilidad).
# ---------------------------------------------------------------------------

EXACT_REPLACEMENTS: dict[str, str] = {
    # Año (factor(ANIO))
    "factor(ANIO)2020": "Año 2020 (vs 2019)",
    "factor(ANIO)2021": "Año 2021 (vs 2019)",
    "factor(ANIO)2022": "Año 2022 (vs 2019)",
    "factor(ANIO)2023": "Año 2023 (vs 2019)",
    "factor(ANIO)2024": "Año 2024 (vs 2019)",
    # Sexo
    "factor(QSSEXO)2": "Mujer (vs Hombre)",
    "factor(QSSEXO)1": "Hombre (referencia)",
    # Área
    "factor(HV025)2": "Rural (vs Urbano)",
    "factor(HV025)1": "Urbano (referencia)",
    # Riqueza
    "factor(HV270)1": "Quintil 1 — Más pobre (referencia)",
    "factor(HV270)2": "Quintil 2 — Pobre",
    "factor(HV270)3": "Quintil 3 — Medio",
    "factor(HV270)4": "Quintil 4 — Rico",
    "factor(HV270)5": "Quintil 5 — Más rico",
    # Educación (QS25N) — diccionario INEI real (auditoría 2026-07-16):
    # 0 Sin nivel/inicial (incluye QS24=2 nunca asistió), 1 Primaria, 2 Secundaria,
    # 3 Superior no universitaria, 4 Superior universitaria, 5 Postgrado.
    "factor(QS25N)0": "Sin nivel / inicial (referencia)",
    "factor(QS25N)1": "Primaria",
    "factor(QS25N)2": "Secundaria",
    "factor(QS25N)3": "Superior no universitaria",
    "factor(QS25N)4": "Superior universitaria",
    "factor(QS25N)5": "Postgrado",
    # Violencia de pareja
    "factor(VIOLENCIA_PAREJA)0": "Sin violencia (referencia)",
    "factor(VIOLENCIA_PAREJA)1": "Violencia leve / única",
    "factor(VIOLENCIA_PAREJA)2": "Violencia múltiple / severa",
    # Alcohol problemático
    "factor(ALCOHOL_PROBLEMATICO)0": "Sin consumo problemático (referencia)",
    "factor(ALCOHOL_PROBLEMATICO)1": "Consumo problemático",
    "factor(ALCOHOL_PROBLEMATICO)NaN": "No aplica (no consume; salto del cuestionario)",
    # Consumo de tabaco en los últimos 30 días (QS201; ENDES Sí=1 / No=2)
    "factor(QS201)1": "Fumó en últimos 30 días (referencia)",
    "factor(QS201)2": "No fumó en últimos 30 días",
    "factor(QS201)NaN": "No aplica (no fumó en últimos 12 meses; salto del cuestionario)",
    # Diagnóstico de diabetes (QS109; ENDES Sí=1 / No=2)
    "factor(QS109)1": "Diagnóstico de diabetes (referencia)",
    "factor(QS109)2": "Sin diagnóstico de diabetes",
    # Calidad de dieta
    "factor(CALIDAD_DIETA)0": "Dieta no adecuada (referencia)",
    "factor(CALIDAD_DIETA)1": "Dieta adecuada",
    # Continuas - cuando aparecen como Categoría en Tabla 1
    "PHQ9_TOTAL": "Puntaje PHQ-9 (continuo, 0–27)",
    "QS23": "Edad (años, continuo)",
    "QS907": "Circunferencia abdominal (cm, continuo)",
    "IMC": "Índice de masa corporal (kg/m²)",
    # Términos especiales
    "(Intercept)": "Intercepto",
    "PHQ9_TOTAL:factor(QSSEXO)2": "Interacción PHQ-9 × Mujer",
    "PHQ9_TOTAL:factor(ANIO)2020": "Interacción PHQ-9 × Año 2020",
    "PHQ9_TOTAL:factor(ANIO)2021": "Interacción PHQ-9 × Año 2021",
    "PHQ9_TOTAL:factor(ANIO)2022": "Interacción PHQ-9 × Año 2022",
    "PHQ9_TOTAL:factor(ANIO)2023": "Interacción PHQ-9 × Año 2023",
    "PHQ9_TOTAL:factor(ANIO)2024": "Interacción PHQ-9 × Año 2024",
    # Grupos
    "overall": "General",
    # Métodos internos
    "rubin_d1_chisq_approx": "Combinación tipo Rubin (D1 chi-cuadrado)",
    # Familias estadísticas
    "quasipoisson": "Cuasi-Poisson",
    "quasibinomial": "Cuasi-binomial",
    # Tabla 5 — celdas vacías
    None: None,  # placeholder for clarity; manejo aparte
}


# Patrones regex para .tmp_predN — se reemplaza por descripción genérica del nivel
_TMP_PATTERN = re.compile(r"\.tmp_pred([A-Za-z0-9_]+)")


def _humanize_tmp_pred(match: re.Match[str]) -> str:
    raw = match.group(1)
    if raw == "NaN":
        return "[No aplica]"
    # Niveles numéricos: .tmp_pred1, .tmp_pred2, ...
    if raw.isdigit():
        return f"[Nivel {raw}]"
    # Niveles categóricos textuales: .tmp_predLeve, .tmp_predSevera...
    return f"[{raw}]"


def transform_cell(value: object) -> object:
    """Aplica los reemplazos de etiqueta a una celda. Conserva tipos no-string."""
    if value is None:
        return value
    if not isinstance(value, str):
        return value
    s = value
    # Reemplazo exacto si toda la celda coincide
    if s in EXACT_REPLACEMENTS:
        repl = EXACT_REPLACEMENTS[s]
        if repl is not None:
            return repl
    # Reemplazo de tokens internos (.tmp_predXxx, listas de varios separados por coma)
    if ".tmp_pred" in s:
        # Caso: lista separada por comas
        parts = [p.strip() for p in s.split(",")]
        humanized_parts = []
        for part in parts:
            humanized_parts.append(_TMP_PATTERN.sub(_humanize_tmp_pred, part))
        return ", ".join(humanized_parts)
    return s


def _add_portada(wb, *, tabla_v2: str, autorizado_por: str, fuente: str, transformaciones: list[str]) -> None:
    """Inserta una hoja Portada al inicio del workbook con metadato de Fase C."""
    ws = wb.create_sheet("Portada", 0)
    ws["A1"] = f"{tabla_v2} — versión v2 (Auditoría Integral Fase C)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Fecha: 2026-05-27"
    ws["A4"] = "Autorización"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = f"  Matriz B fila: {autorizado_por}"
    ws["A6"] = f"  Fuente original: {fuente}"
    ws["A7"] = "  Reglas aplicadas:"
    for i, t in enumerate(transformaciones, start=8):
        ws.cell(row=i, column=1, value=f"   - {t}")
    ws["A20"] = "Test plan asociado: 06_decision_principal_suplementario.xlsx hoja 'Test_Plan'."
    ws.column_dimensions["A"].width = 100


def regenerate_table(
    src_path: Path,
    out_path: Path,
    *,
    tabla_v2: str,
    autorizado_por: str,
    transformaciones: list[str],
    nan_replacement: str | None = None,
) -> None:
    """Copia el xlsx fuente, aplica reemplazos en TODAS las celdas, inserta Portada y guarda."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(src_path, out_path)
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                # Reemplazo principal por mapeo
                new_value = transform_cell(cell.value)
                # Reemplazar None con '—' si lo pide la opción
                if nan_replacement is not None and new_value is None:
                    new_value = nan_replacement
                if new_value != cell.value:
                    cell.value = new_value
    _add_portada(
        wb,
        tabla_v2=tabla_v2,
        autorizado_por=autorizado_por,
        fuente=str(src_path.relative_to(ROOT)),
        transformaciones=transformaciones,
    )
    wb.save(out_path)
    print(f"  -> {out_path.name}")


def regenerate_tabla_s4(src_path: Path, out_path: Path) -> None:
    """Tabla S4: renombrar columna 'variable' y traducir 'overall' a 'General'.

    Mantiene los códigos ENDES (PHQ9_TOTAL etc.) en la columna 'Variable (código ENDES)'
    porque son la identificación canónica de la variable graficada y la columna
    explícitamente lo declara.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(src_path, out_path)
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        # Renombrar el encabezado 'variable'
        for cell in ws[1]:
            if cell.value == "variable":
                cell.value = "Variable (código ENDES)"
            elif cell.value == "group":
                cell.value = "Grupo (estratificación)"
        # Reemplazar 'overall' por 'General' en TODAS las celdas (es etiqueta de grupo)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value == "overall":
                    cell.value = "General"
    _add_portada(
        wb,
        tabla_v2="Tabla S4 (datos figuras)",
        autorizado_por="B12 renombrar",
        fuente=str(src_path.relative_to(ROOT)),
        transformaciones=[
            "Encabezado de columna 'variable' -> 'Variable (código ENDES)' para clarificar que es código fuente.",
            "Encabezado de columna 'group' -> 'Grupo (estratificación)'.",
            "Valor 'overall' (col Grupo) -> 'General' (consistencia idiomática).",
            "El cuerpo de la columna Variable preserva los códigos ENDES por ser archivo de datos fuente.",
        ],
    )
    wb.save(out_path)
    print(f"  -> {out_path.name}")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    print("Fase C — generación de tablas v2:")
    print()

    # C1 — Tabla 1
    regenerate_table(
        SRC_PRINCIPAL / "Tabla_1.xlsx",
        OUT / "Tabla_1_v2.xlsx",
        tabla_v2="Tabla 1 (basal weighted)",
        autorizado_por="B04 regenerar_v2",
        transformaciones=[
            "'overall' → 'General' (consistencia idiomática)",
            "Códigos ENDES en col 'Categoría' (PHQ9_TOTAL, QS23, QS907) → etiquetas legibles",
            "Sin tokens factor(...), .tmp_pred*, rubin_d1, model_N",
        ],
    )

    # C2 — Tabla 3
    regenerate_table(
        SRC_PRINCIPAL / "Tabla_3.xlsx",
        OUT / "Tabla_3_v2.xlsx",
        tabla_v2="Tabla 3 (modelos principales)",
        autorizado_por="B06 regenerar_v2",
        transformaciones=[
            "factor(ANIO)20XX → 'Año 20XX (vs 2019)'",
            "factor(QSSEXO)/HV025/HV270/QS25N/VIOLENCIA_PAREJA → etiquetas categóricas legibles",
            "factor(ALCOHOL_PROBLEMATICO/QS201/QS109/CALIDAD_DIETA) → etiquetas con referencia explícita",
            "Términos de interacción PHQ9_TOTAL:factor(...) → 'PHQ-9 × <subgrupo>'",
            "Códigos QS23/IMC/QS907 → etiquetas (Edad, IMC, Circunferencia abdominal)",
        ],
    )

    # C3 — Tabla 4
    regenerate_table(
        SRC_PRINCIPAL / "Tabla_4.xlsx",
        OUT / "Tabla_4_v2.xlsx",
        tabla_v2="Tabla 4 (cascada de cuidado)",
        autorizado_por="B07 regenerar_v2",
        transformaciones=[
            "Mismo mapeo que Tabla 3 v2",
        ],
    )

    # C4 — Tabla S2
    regenerate_table(
        SRC_SUPL / "Tabla_S2_soporte_tablas.xlsx",
        OUT / "Tabla_S2_v2.xlsx",
        tabla_v2="Tabla S2 (soporte tablas)",
        autorizado_por="B10 regenerar_v2",
        transformaciones=[
            "Mismo mapeo categórico que Tabla 3 v2",
            "rubin_d1_chisq_approx → 'Combinación tipo Rubin (D1 chi-cuadrado)'",
            ".tmp_predN / .tmp_predNaN → '[Nivel N]' / '[No aplica]'",
            "Familia 'quasipoisson' → 'Cuasi-Poisson'; 'quasibinomial' → 'Cuasi-binomial'",
        ],
    )

    # C5 — Tabla S3
    regenerate_table(
        SRC_SUPL / "Tabla_S3_sensibilidad_interacciones.xlsx",
        OUT / "Tabla_S3_v2.xlsx",
        tabla_v2="Tabla S3 (sensibilidad e interacciones)",
        autorizado_por="B11 regenerar_v2",
        transformaciones=[
            "Mismo mapeo categórico que Tabla 3 v2",
            "Términos de interacción PHQ-9 × Año o × Sexo con etiquetas legibles",
        ],
    )

    # C6 — Tabla 5 (renombrar)
    regenerate_table(
        SRC_PRINCIPAL / "Tabla_5.xlsx",
        OUT / "Tabla_5_v2.xlsx",
        tabla_v2="Tabla 5 (sensibilidad principal)",
        autorizado_por="B08 renombrar",
        transformaciones=[
            "Celdas vacías (None) → '—' para evitar lectura como dato faltante real",
            "Mapeo categórico aplicado si aparece factor(...) (no debería)",
        ],
        nan_replacement="—",
    )

    # C7 — Tabla S4 (renombrar suave)
    regenerate_tabla_s4(
        SRC_SUPL / "Tabla_S4_datos_figuras.xlsx",
        OUT / "Tabla_S4_v2.xlsx",
    )

    print()
    print("Listo. Verifica tests con scripts/audit/test_phase_c_tables.py.")


if __name__ == "__main__":
    main()
