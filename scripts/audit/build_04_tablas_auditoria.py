"""Construye 04_tablas_auditoria.xlsx con 5 hojas: problemas, nomenclatura actual,
nomenclatura recomendada, prioridad y decisión propuesta.

Insumos: tablas xlsx en data/output/Para Enviar/Para Publicar Redisenado/{Principal,Suplementario}.
Salida: data/output/Auditoria_Integral/04_tablas_auditoria.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data" / "output" / "Auditoria_Integral" / "04_tablas_auditoria.xlsx"


PROBLEMAS = [
    # (id, tabla, ubicacion, problema, prioridad, decision_propuesta)
    ("P01", "Tabla_3 (Principal)", "col 'Término' filas 5-9, 29-33", "Aparece literalmente 'factor(ANIO)2020'…'factor(ANIO)2024' en columna publicable", "alta", "regenerar_v2"),
    ("P02", "Tabla_4 (Principal)", "col 'Término' filas 4-8, 24-28", "Aparece literalmente 'factor(ANIO)20XX' en columna publicable", "alta", "regenerar_v2"),
    ("P03", "Tabla_S3 Sensibilidad_interacciones", "col 'Término' múltiples filas", "Aparece literalmente 'factor(ANIO)20XX'", "alta", "regenerar_v2"),
    ("P04", "Tabla_S3 Sensibilidad_logistica", "col 'Término' filas 4-8", "Aparece literalmente 'factor(ANIO)20XX'", "alta", "regenerar_v2"),
    ("P05", "Tabla_1 (Resumen_general, Por_desenlace)", "col 'Categoría' filas finales (39-41, 76-81)", "Aparecen códigos crudos 'PHQ9_TOTAL', 'QS23', 'QS907' como categoría de variable continua", "alta", "regenerar_v2"),
    ("P06", "Tabla_S2 Tabla_1_por_sexo, Por_severidad", "col 'Categoría' filas finales", "Aparecen códigos crudos 'PHQ9_TOTAL', 'QS23', 'QS907'", "alta", "regenerar_v2"),
    ("P07", "Tabla_S2 Tabla_2_detalle", "col 'Método de pooling' filas 2-11", "Método interno 'rubin_d1_chisq_approx' visible en tabla publicable", "media", "regenerar_v2"),
    ("P08", "Tabla_S2 Tabla_2_detalle", "col 'Términos' filas 2-11", "Nombres internos de R '.tmp_pred1', '.tmp_predNaN', '.tmp_predMinima' etc. en tabla publicable", "alta", "regenerar_v2"),
    ("P09", "Tabla_S4 Figura_2_spline", "col 'variable' filas 2-29", "Código crudo 'PHQ9_TOTAL' como nombre de variable; aceptable en datos fuente pero conviene etiquetarlo en encabezado", "baja", "renombrar"),
    ("P10", "Tabla_1 (todas)", "col 'Grupo' valores", "Valor 'overall' (inglés) usado como etiqueta del grupo general; inconsistente con resto en español", "media", "renombrar"),
    ("P11", "Tabla_3 (Principal)", "col 'Término' filas con HV270", "Niveles de riqueza ('factor(HV270)2..5') deberían mostrarse como 'Q2 — Pobre' … 'Q5 — Más rico' (quintiles ordenados)", "alta", "regenerar_v2"),
    ("P12", "Tabla_3 (Principal)", "col 'Término' filas con QS25N", "Niveles de educación ('factor(QS25N)1..5') deberían mostrarse con etiquetas: 'Sin educación', 'Primaria incompleta', 'Primaria completa', 'Secundaria', 'Superior'", "alta", "regenerar_v2"),
    ("P13", "Tabla_3, S3", "filas de interacción", "Términos de interacción `PHQ9_TOTAL:factor(QSSEXO)2` o similares deben mostrarse como 'PHQ-9 × Sexo (mujer)'", "media", "regenerar_v2"),
    ("P14", "Todas las tablas", "encabezados", "Encabezado 'n medio' es ambiguo: ¿n promedio entre imputaciones, ponderado o no? Aclarar como 'n medio entre imputaciones (no ponderado)'", "baja", "renombrar"),
    ("P15", "Tabla_1 Resumen_general y derivadas", "col 'DEFF promedio'", "DEFF mostrado solo en Resumen_general y no en Por_desenlace/Por_sexo/Por_severidad — verificar consistencia o anotar deliberadamente", "baja", "mantener"),
    ("P16", "Tabla_5 (Principal)", "fila 'Spline corregida'", "Estimación/IC/p con valor None: aceptable porque es resumen de no-linealidad, pero conviene poner '—' explícito en lugar de celda vacía", "baja", "renombrar"),
    ("P17", "Tabla_S2 Tabla_2_detalle", "col 'gl'", "gl como string vacío en algunas filas; uniformar como entero o '—'", "baja", "renombrar"),
    ("P18", "Tabla_S2 Diagnosticos_modelos", "col 'Familia'", "'quasipoisson' / 'quasibinomial' en minúsculas: convención R, no de revista. Reemplazar por 'Cuasi-Poisson' / 'Cuasi-binomial'", "baja", "renombrar"),
    ("P19", "Tabla_4 (Principal)", "fila 'No adherencia terapéutica' valores p de año", "p = 0,008 para factor(ANIO)2024 en submodelo no adherencia: el dato existe; pero el lector no técnico puede confundirlo con efecto sustantivo. Conviene un pie de tabla explicando: el efecto de año captura cambios temporales no atribuibles a PHQ-9", "baja", "mantener"),
    ("P20", "Tabla_S4 Figura_1_STROBE", "todas", "Datos de la figura STROBE están duplicados con Tabla_S1 — verificar que ambos conjuntos coincidan tras cualquier v2 de la figura STROBE", "media", "mantener"),
]


NOMENCLATURA = [
    # (id, tabla, nombre_actual, nombre_recomendado, justificacion)
    ("N01", "Tabla_3 / Tabla_4 / S3", "factor(ANIO)2020", "Año 2020 (vs 2019)", "Año como referencia explícita; quita sintaxis interna de R"),
    ("N02", "Tabla_3 / Tabla_4 / S3", "factor(ANIO)2021", "Año 2021 (vs 2019)", ""),
    ("N03", "Tabla_3 / Tabla_4 / S3", "factor(ANIO)2022", "Año 2022 (vs 2019)", ""),
    ("N04", "Tabla_3 / Tabla_4 / S3", "factor(ANIO)2023", "Año 2023 (vs 2019)", ""),
    ("N05", "Tabla_3 / Tabla_4 / S3", "factor(ANIO)2024", "Año 2024 (vs 2019)", ""),
    ("N06", "Tabla_3 / Tabla_4 / S3", "factor(QSSEXO)2", "Mujer (vs Hombre)", "Etiqueta clara del nivel categórico"),
    ("N07", "Tabla_3 / Tabla_4 / S3", "factor(HV025)2", "Rural (vs Urbano)", ""),
    ("N08", "Tabla_3 / Tabla_4 / S3", "factor(HV270)2", "Quintil 2 — Pobre (vs Q1 más pobre)", "HV270 es índice de riqueza por quintiles"),
    ("N09", "Tabla_3 / Tabla_4 / S3", "factor(HV270)3", "Quintil 3 — Medio (vs Q1)", ""),
    ("N10", "Tabla_3 / Tabla_4 / S3", "factor(HV270)4", "Quintil 4 — Rico (vs Q1)", ""),
    ("N11", "Tabla_3 / Tabla_4 / S3", "factor(HV270)5", "Quintil 5 — Más rico (vs Q1)", ""),
    ("N12", "Tabla_3 / Tabla_4 / S3", "factor(QS25N)1", "Sin educación / inicial (vs ref.)", "Recategorización de QS25N según diccionario ENDES"),
    ("N13", "Tabla_3 / Tabla_4 / S3", "factor(QS25N)2", "Primaria incompleta", ""),
    ("N14", "Tabla_3 / Tabla_4 / S3", "factor(QS25N)3", "Primaria completa", ""),
    ("N15", "Tabla_3 / Tabla_4 / S3", "factor(QS25N)4", "Secundaria", ""),
    ("N16", "Tabla_3 / Tabla_4 / S3", "factor(QS25N)5", "Superior", ""),
    ("N17", "Tabla_3 / Tabla_4 / S3", "factor(VIOLENCIA_PAREJA)1", "Violencia leve / única (vs sin violencia)", "Verificar codificación exacta de VPAR en pipeline"),
    ("N18", "Tabla_3 / Tabla_4 / S3", "factor(VIOLENCIA_PAREJA)2", "Violencia múltiple/severa (vs sin violencia)", ""),
    ("N19", "Tabla_3 (Modelo 3) / S3", "factor(ALCOHOL_PROBLEMATICO)1", "Consumo problemático (vs no)", ""),
    ("N20", "Tabla_3 (Modelo 3) / S3", "factor(ALCOHOL_PROBLEMATICO)NaN", "No aplica (no consume; salto del cuestionario)", "Nivel explícito por salto estructural, no faltante MAR"),
    ("N21", "Tabla_3 (Modelo 3) / S3", "factor(QS201)2", "Consume rara vez (vs frecuente)", "Verificar dirección con diccionario ENDES"),
    ("N22", "Tabla_3 (Modelo 3) / S3", "factor(QS201)NaN", "No aplica (no consume)", "Salto estructural del cuestionario"),
    ("N23", "Tabla_3 (Modelo 3) / S3", "factor(QS109)2", "Sin medicación antihipertensiva (vs con)", "Verificar dirección con diccionario ENDES"),
    ("N24", "Tabla_3 (Modelo 3) / S3", "factor(CALIDAD_DIETA)1", "Dieta adecuada (vs no adecuada)", ""),
    ("N25", "Tabla_3 (Modelo 3) / S3", "IMC", "Índice de masa corporal (kg/m²)", "Continuo"),
    ("N26", "Tabla_3 (Modelo 3) / S3", "QS907", "Circunferencia abdominal (cm)", "Continuo"),
    ("N27", "Tabla_1 / S2", "Categoría: PHQ9_TOTAL", "Puntaje PHQ-9 (continuo, 0-27)", "Variable continua: la celda contiene la métrica, no la categoría"),
    ("N28", "Tabla_1 / S2", "Categoría: QS23", "Edad (años, continuo)", ""),
    ("N29", "Tabla_1 / S2", "Categoría: QS907", "Circunferencia abdominal (cm, continuo)", ""),
    ("N30", "Tabla_1 / S2", "Grupo: overall", "General (toda la cohorte)", "Reemplazar 'overall' por su equivalente español"),
    ("N31", "Tabla_S2 Tabla_2_detalle", "Método: rubin_d1_chisq_approx", "Combinación Rubin (D1 aproximado, chi-cuadrado)", "Etiqueta legible"),
    ("N32", "Tabla_S2 Tabla_2_detalle", "Términos: .tmp_pred1, .tmp_predNaN, etc.", "Niveles probados: <descripción legible por variable>", "Eliminar o expandir; los .tmp_predN no son interpretables"),
    ("N33", "Tabla_S2 Diagnosticos_modelos", "Familia: quasipoisson", "Cuasi-Poisson", ""),
    ("N34", "Tabla_S2 Diagnosticos_modelos", "Familia: quasibinomial", "Cuasi-binomial", ""),
    ("N35", "Todas", "Columna: 'n medio'", "n medio entre imputaciones (no ponderado)", "Aclara la métrica"),
    ("N36", "Tabla_3 / S3", "Término: PHQ9_TOTAL:factor(QSSEXO)2", "Interacción PHQ-9 × Mujer", "Notación legible de la interacción"),
    ("N37", "Tabla_3 / S3", "Término: PHQ9_TOTAL:factor(ANIO)2020", "Interacción PHQ-9 × Año 2020", ""),
]


DECISIONES = [
    # (tabla, decision, justificacion_breve)
    ("Tabla_1 (Principal)", "regenerar_v2", "Códigos crudos en col Categoría (PHQ9_TOTAL/QS23/QS907) y 'overall' → reemplazar antes de publicar"),
    ("Tabla_2 (Principal)", "mantener", "Encabezados y categorías ya legibles; sin tokens raw"),
    ("Tabla_3 (Principal)", "regenerar_v2", "factor(ANIO)/factor(QSSEXO)/factor(HV270)/factor(QS25N)/factor(QS201) etc. requieren etiquetas legibles"),
    ("Tabla_4 (Principal)", "regenerar_v2", "Mismos factor() problemas que Tabla_3"),
    ("Tabla_5 (Principal)", "renombrar", "Solo aclarar celdas None como '—' y reformular fila spline"),
    ("Tabla_S1_flujo_STROBE", "mantener", "Datos numéricos correctos; etiquetas ya legibles"),
    ("Tabla_S2 soporte_tablas", "regenerar_v2", "rubin_d1_chisq_approx, .tmp_predN y códigos QS23/QS907 → no son publicables tal cual"),
    ("Tabla_S3 sensibilidad_interacciones", "regenerar_v2", "Tokens factor() en términos de modelos"),
    ("Tabla_S4 datos_figuras", "renombrar", "Datos fuente correctos; mejorar etiqueta de la columna 'variable' en encabezado"),
]


PRIORIDAD_NOTAS = [
    ("alta", "tokens raw R en celdas publicables (factor(...), .tmp_pred*)", "bloquea publicación si la tabla va al manuscrito"),
    ("alta", "códigos crudos ENDES como categoría (PHQ9_TOTAL, QS23, QS907)", "lector no técnico no los entiende"),
    ("media", "etiquetas mixtas español/inglés ('overall')", "consistencia editorial; trivial de arreglar"),
    ("media", "método interno 'rubin_d1_chisq_approx' en tabla", "explicación didáctica deficiente"),
    ("baja", "encabezados ambiguos ('n medio'), familia minúscula ('quasipoisson')", "no bloquea pero mejora legibilidad"),
]


def write_xlsx(path: Path) -> None:
    wb = Workbook()

    # --- Hoja 0: portada -----------------------------------------------------
    ws = wb.active
    ws.title = "Portada"
    ws["A1"] = "Auditoría Integral — Tarea A4: Tablas"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Fecha: 2026-05-27"
    ws["A4"] = "Hojas:"
    ws["A5"] = "  1. Problemas — issues identificados por tabla con prioridad y decisión propuesta"
    ws["A6"] = "  2. Nomenclatura — tokens actuales y reemplazos sugeridos"
    ws["A7"] = "  3. Decisiones — qué hacer con cada tabla"
    ws["A8"] = "  4. Prioridad — criterios de priorización"
    ws["A10"] = "Reglas implícitas (heredadas de la audit general):"
    ws["A11"] = "  - p con 3 decimales, estimadores con 2, porcentajes con 1."
    ws["A12"] = "  - Coma decimal (RPMESP)."
    ws["A13"] = "  - Ningún factor(...), model_N, ni código crudo ENDES en tablas publicables."
    ws["A14"] = "  - Etiquetas en español; niveles categóricos con descripción explícita."
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 90

    # --- Hoja 1: Problemas ---------------------------------------------------
    ws = wb.create_sheet("Problemas")
    headers = ["ID", "Tabla", "Ubicación", "Problema", "Prioridad", "Decisión propuesta"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for row in PROBLEMAS:
        ws.append(row)
    for col_idx, width in enumerate([8, 30, 35, 90, 12, 22], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # --- Hoja 2: Nomenclatura ------------------------------------------------
    ws = wb.create_sheet("Nomenclatura")
    headers = ["ID", "Tabla afectada", "Nombre actual", "Nombre recomendado", "Justificación / nota"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for row in NOMENCLATURA:
        ws.append(row)
    for col_idx, width in enumerate([8, 32, 50, 50, 60], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # --- Hoja 3: Decisiones --------------------------------------------------
    ws = wb.create_sheet("Decisiones")
    headers = ["Tabla", "Decisión propuesta", "Justificación"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for row in DECISIONES:
        ws.append(row)
        # color por decisión
        last_row = ws.max_row
        color = {
            "mantener": "E2EFDA",
            "renombrar": "FFF2CC",
            "regenerar_v2": "FCE4D6",
            "descartar": "F8CBAD",
        }.get(row[1], None)
        if color:
            ws.cell(row=last_row, column=2).fill = PatternFill("solid", fgColor=color)
    for col_idx, width in enumerate([42, 22, 90], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # --- Hoja 4: Prioridad ---------------------------------------------------
    ws = wb.create_sheet("Prioridad")
    headers = ["Nivel", "Criterio", "Impacto editorial"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for row in PRIORIDAD_NOTAS:
        ws.append(row)
    for col_idx, width in enumerate([10, 60, 60], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Escrito: {path}")
    # Resumen
    cnt_alta = sum(1 for p in PROBLEMAS if p[4] == "alta")
    cnt_media = sum(1 for p in PROBLEMAS if p[4] == "media")
    cnt_baja = sum(1 for p in PROBLEMAS if p[4] == "baja")
    print(f"  Problemas alta:  {cnt_alta}")
    print(f"  Problemas media: {cnt_media}")
    print(f"  Problemas baja:  {cnt_baja}")
    print(f"  Tablas con decisión 'regenerar_v2': {sum(1 for d in DECISIONES if d[1] == 'regenerar_v2')}")


if __name__ == "__main__":
    write_xlsx(OUT)
