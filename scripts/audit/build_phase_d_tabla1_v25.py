"""Tabla_1_v2_5.xlsx — Prototipo Tabla 1 estilo NHANES alineado al núcleo del proyecto.

Decisiones (acordadas con el autor el 2026-05-29):
- 5 niveles de SEVERIDAD_DEPRESIVA preservados (Mínima/Leve/Moderada/Mod_Severa/Severa),
  los mismos que usa el pipeline. NO se colapsan a 3.
- Continuas (edad, IMC, cintura) en una fila con media (DE) ponderada — formato
  consistente con la Tabla 1 v2 actual.
- p-value: Rao-Scott (svychisq para categóricas, regTermTest sobre svyglm para continuas).
  Mismo motor que rao_scott_summary.csv del pipeline.
- Variantes: solo "journal Latam" (n no ponderado, % ponderado). Una sola hoja.

Datos fuente: data/output/analysis/imputed/imputation_01.parquet (prototipo con la
primera imputación; el pooling sobre 20 puede activarse si la tabla pasa a definitiva).

Salida: data/output/Auditoria_Integral/Tabla_1_v2_5.xlsx
"""
from __future__ import annotations

import json
import subprocess
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from etiquetas_educacion import EDU_LABELS_ES, EDU_ORDER_ES
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
IMP_PATH = ROOT / "data" / "output" / "analysis" / "imputed" / "imputation_01.parquet"
CONFIG_PATH = ROOT / "config" / "pipeline_config.json"
OUT_PATH = ROOT / "data" / "output" / "Auditoria_Integral" / "Tabla_1_v2_5.xlsx"
RSCRIPT_EXE = Path("C:/Program Files/R/R-4.5.3/bin/Rscript.exe")
R_LIB_USER = Path("C:/Users/Trabajo/Documents/R/win-library/4.5")


# Severidad en el orden del pipeline (Mínima → Severa) con rangos PHQ-9 oficiales
SEV_LEVELS = ["Minima", "Leve", "Moderada", "Mod_Severa", "Severa"]
SEV_HEADERS = {
    "Minima": ("Mínima", "0–4"),
    "Leve": ("Leve", "5–9"),
    "Moderada": ("Moderada", "10–14"),
    "Mod_Severa": ("Mod. severa", "15–19"),
    "Severa": ("Severa", "20–27"),
}
SEX_ORDER = ["Hombres", "Mujeres"]


# Categorías "lisas" (mapeo desde códigos ENDES a etiquetas legibles), preservando
# la misma codificación que usa el pipeline (analysis.py / pipeline.py).
CATEGORICAL_VARS = [
    # (col_id, label_for_table, mapper_dict_or_None_for_NaN, ordered_levels)
    ("HV025", "Área de residencia", {1: "Urbano", 2: "Rural"}, ["Urbano", "Rural"]),
    (
        "HV270", "Quintil de riqueza",
        {1: "Q1 — Más pobre", 2: "Q2", 3: "Q3", 4: "Q4", 5: "Q5 — Más rico"},
        ["Q1 — Más pobre", "Q2", "Q3", "Q4", "Q5 — Más rico"],
    ),
    (
        # Etiquetas corregidas (auditoria 2026-07-16): codificacion real 0..5 del
        # diccionario INEI; fuente unica en etiquetas_educacion.py.
        "QS25N", "Nivel educativo",
        dict(EDU_LABELS_ES),
        list(EDU_ORDER_ES),
    ),
    (
        # Codificación real del pipeline (auditoría 2026-07-16): 0 = sin violencia;
        # 1 = con violencia FÍSICA (QS710/QS711 son actos físicos); 2 = sin pareja.
        "VIOLENCIA_PAREJA", "Violencia de pareja (último año)",
        {0: "Sin violencia", 1: "Con violencia (física)", 2: "Sin pareja"},
        ["Sin violencia", "Con violencia (física)", "Sin pareja"],
    ),
    (
        "ALCOHOL_PROBLEMATICO", "Consumo problemático de alcohol",
        {0: "No", 1: "Sí", "NA": "No aplica (no consume)"},
        ["No", "Sí", "No aplica (no consume)"],
    ),
    (
        "QS201", "Consumo de tabaco (últimos 30 días)",
        {1: "Sí (fumó últimos 30 días)", 2: "No fumó", "NA": "No aplica (no fumó últimos 12 meses)"},
        ["Sí (fumó últimos 30 días)", "No fumó", "No aplica (no fumó últimos 12 meses)"],
    ),
    (
        "QS109", "Diagnóstico de diabetes",
        {1: "Sí (diabetes)", 2: "No"},
        ["Sí (diabetes)", "No"],
    ),
    (
        "CALIDAD_DIETA", "Calidad de la dieta",
        {0: "No adecuada", 1: "Adecuada"},
        ["Adecuada", "No adecuada"],
    ),
    (
        "DX_HTA_PREVIO", "Diagnóstico previo de HTA",
        {0: "No", 1: "Sí"},
        ["No", "Sí"],
    ),
    (
        "PRESION_ARTERIAL_ELEVADA", "Presión arterial elevada (medición actual)",
        {0: "No", 1: "Sí"},
        ["No", "Sí"],
    ),
]

# Continuas: (col, label, unidades)
CONTINUOUS_VARS = [
    ("QS23", "Edad", "años"),
    ("IMC", "Índice de masa corporal", "kg/m²"),
    ("QS907", "Circunferencia abdominal", "cm"),
]


# ---------------------------------------------------------------------------
# Preparación de datos
# ---------------------------------------------------------------------------

def load_and_prepare() -> pd.DataFrame:
    df = pd.read_parquet(IMP_PATH)
    df["SEXO_LBL"] = df["QSSEXO"].astype(int).map({1: "Hombres", 2: "Mujeres"})

    # Materializar etiquetas legibles para cada categórica (preservando NA explícito
    # como "No aplica" en las variables con salto estructural).
    for col, _label, mapper, _levels in CATEGORICAL_VARS:
        out_col = f"{col}_LBL"
        series = df[col]
        # Tratar NaN/NA explícitamente si la variable tiene "NA" en su mapper
        if "NA" in mapper:
            na_label = mapper["NA"]
            mapped = pd.Series(np.nan, index=df.index, dtype=object)
            mapped[series.isna()] = na_label
            for code, lbl in mapper.items():
                if code == "NA":
                    continue
                mapped[series == code] = lbl
        else:
            mapped = series.map(mapper)
        df[out_col] = mapped

    return df


def weighted_mean_sd(x: np.ndarray, w: np.ndarray) -> tuple[float, float]:
    x = np.asarray(x, dtype=float)
    w = np.asarray(w, dtype=float)
    valid = (~np.isnan(x)) & (w > 0) & np.isfinite(w)
    x = x[valid]
    w = w[valid]
    if x.size == 0:
        return float("nan"), float("nan")
    mean = float(np.sum(x * w) / np.sum(w))
    var = float(np.sum(w * (x - mean) ** 2) / np.sum(w))
    return mean, float(np.sqrt(var))


# ---------------------------------------------------------------------------
# Rao-Scott vía rpy2 (mismo bridge que el pipeline)
# ---------------------------------------------------------------------------

_RSCRIPT_TEMPLATE = r"""
.libPaths(c('{r_lib_user}', .libPaths()))
suppressPackageStartupMessages({{
  library(survey)
}})
options(survey.lonely.psu="adjust")

args <- commandArgs(trailingOnly = TRUE)
csv_in   <- args[1]
csv_jobs <- args[2]
csv_out  <- args[3]

df <- read.csv(csv_in, stringsAsFactors = FALSE, fileEncoding = "UTF-8")
jobs <- read.csv(csv_jobs, stringsAsFactors = FALSE, fileEncoding = "UTF-8")

build_design_sub <- function(df, sex_filter) {{
  sub <- df[df$SEXO_LBL == sex_filter, , drop = FALSE]
  survey::svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, data=sub, nest=TRUE)
}}

rao_scott_categorical <- function(df, var, sex_filter) {{
  d <- build_design_sub(df, sex_filter)
  f <- as.formula(paste0("~", var, " + SEVERIDAD_DEPRESIVA"))
  res <- tryCatch(
    suppressWarnings(survey::svychisq(f, design=d, statistic="F")),
    error = function(e) NULL
  )
  if (is.null(res)) return(NA_real_)
  as.numeric(res$p.value)
}}

rao_scott_continuous <- function(df, var, sex_filter) {{
  d <- build_design_sub(df, sex_filter)
  f <- as.formula(paste0(var, " ~ factor(SEVERIDAD_DEPRESIVA)"))
  fit <- tryCatch(
    suppressWarnings(survey::svyglm(f, design=d)),
    error = function(e) NULL
  )
  if (is.null(fit)) return(NA_real_)
  test <- tryCatch(
    suppressWarnings(survey::regTermTest(fit, "factor(SEVERIDAD_DEPRESIVA)")),
    error = function(e) NULL
  )
  if (is.null(test)) return(NA_real_)
  as.numeric(test$p)
}}

results <- data.frame(
  variable = character(),
  sex      = character(),
  type     = character(),
  p_value  = numeric(),
  stringsAsFactors = FALSE
)

for (i in seq_len(nrow(jobs))) {{
  var  <- jobs$variable[i]
  sex  <- jobs$sex[i]
  type <- jobs$type[i]
  cat(sprintf("[%d/%d] %s | %s (%s)\n", i, nrow(jobs), var, sex, type))
  if (type == "categorical") {{
    p <- rao_scott_categorical(df, var, sex)
  }} else {{
    p <- rao_scott_continuous(df, var, sex)
  }}
  results <- rbind(results, data.frame(
    variable = var, sex = sex, type = type, p_value = p,
    stringsAsFactors = FALSE
  ))
}}

write.csv(results, csv_out, row.names = FALSE, fileEncoding = "UTF-8")
cat("OK\n")
"""


def compute_rao_scott_pvalues(df: pd.DataFrame) -> dict:
    """Computa Rao-Scott p-values vía Rscript subprocess (más robusto que rpy2 en Windows)."""
    needed_cols = (
        ["HV001", "HV022", "PESO_FINAL", "SEVERIDAD_DEPRESIVA", "SEXO_LBL"]
        + [c for c, *_ in CATEGORICAL_VARS]
        + [c for c, *_ in CONTINUOUS_VARS]
        + [f"{c}_LBL" for c, *_ in CATEGORICAL_VARS]
    )
    df_for_r = df[needed_cols].copy()
    for col in df_for_r.columns:
        if pd.api.types.is_integer_dtype(df_for_r[col].dtype):
            df_for_r[col] = df_for_r[col].astype("float64")

    # Jobs: una fila por (variable, sex)
    jobs = []
    for col, *_ in CATEGORICAL_VARS:
        for sex in SEX_ORDER:
            jobs.append({"variable": f"{col}_LBL", "sex": sex, "type": "categorical", "var_original": col})
    for col, *_ in CONTINUOUS_VARS:
        for sex in SEX_ORDER:
            jobs.append({"variable": col, "sex": sex, "type": "continuous", "var_original": col})
    jobs_df = pd.DataFrame(jobs)

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        csv_in = tmp_dir / "data.csv"
        csv_jobs = tmp_dir / "jobs.csv"
        csv_out = tmp_dir / "results.csv"
        r_script = tmp_dir / "compute_rao_scott.R"

        df_for_r.to_csv(csv_in, index=False, encoding="utf-8")
        jobs_df[["variable", "sex", "type"]].to_csv(csv_jobs, index=False, encoding="utf-8")
        r_script.write_text(
            _RSCRIPT_TEMPLATE.format(r_lib_user=str(R_LIB_USER).replace("\\", "/")),
            encoding="utf-8",
        )

        cmd = [str(RSCRIPT_EXE), "--vanilla", str(r_script), str(csv_in), str(csv_jobs), str(csv_out)]
        proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
        if proc.returncode != 0:
            print("STDOUT:", proc.stdout)
            print("STDERR:", proc.stderr)
            raise RuntimeError("Rscript falló al computar Rao-Scott")
        # Imprimir resumen R
        last_lines = [line for line in proc.stdout.splitlines() if line.strip()][-5:]
        for line in last_lines:
            print(f"    R| {line}")

        results = pd.read_csv(csv_out)

    # Map a dict (variable_original, sex) -> p
    var_map = {row["variable"]: row["var_original"] for _, row in jobs_df.iterrows()}
    pvalues: dict[tuple[str, str], float] = {}
    for _, row in results.iterrows():
        original = var_map[row["variable"]]
        pvalues[(original, row["sex"])] = float(row["p_value"]) if pd.notna(row["p_value"]) else float("nan")
    return pvalues


# ---------------------------------------------------------------------------
# Construcción del workbook
# ---------------------------------------------------------------------------

def _fmt_n_pct(n_unw: int, pct_w: float) -> str:
    n_str = f"{int(n_unw):,}".replace(",", " ")
    pct_str = f"{pct_w:.1f}".replace(".", ",")
    return f"{n_str} ({pct_str})"


def _fmt_mean_sd(mean: float, sd: float) -> str:
    if not np.isfinite(mean):
        return "—"
    m_str = f"{mean:.1f}".replace(".", ",")
    s_str = f"{sd:.1f}".replace(".", ",")
    return f"{m_str} ({s_str})"


def _fmt_p(p: float) -> str:
    if p is None or not np.isfinite(p):
        return "—"
    if p < 0.001:
        return "< 0,001"
    return f"{p:.3f}".replace(".", ",")


def build_workbook(df: pd.DataFrame, pvalues: dict) -> Workbook:
    # Totales por (sexo, severidad)
    cell_totals_w: dict[tuple[str, str], float] = {}
    cell_totals_unw: dict[tuple[str, str], int] = {}
    for sex in SEX_ORDER:
        for sev in SEV_LEVELS:
            mask = (df["SEXO_LBL"] == sex) & (df["SEVERIDAD_DEPRESIVA"] == sev)
            cell_totals_w[(sex, sev)] = float(df.loc[mask, "PESO_FINAL"].sum())
            cell_totals_unw[(sex, sev)] = int(mask.sum())

    wb = Workbook()

    # --- Portada ------------------------------------------------------------
    ws = wb.active
    ws.title = "Portada"
    ws["A1"] = "Tabla 1 v2.5 — Características basales por sexo y severidad PHQ-9 (5 niveles)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Fecha de generación: 2026-05-29"
    ws["A4"] = "Decisiones acordadas"
    ws["A4"].font = Font(bold=True)
    notas_decisiones = [
        "  - 5 niveles de SEVERIDAD_DEPRESIVA preservados (Mínima/Leve/Moderada/Mod. severa/Severa),",
        "    los mismos del pipeline (SEVERIDAD_DEPRESIVA en analysis.py).",
        "  - Continuas (Edad, IMC, Cintura) en una fila con media (DE) ponderada.",
        "  - Categóricas con n no ponderado (% ponderado) por celda.",
        "  - p-value: Rao-Scott vía survey::svychisq (categóricas) y regTermTest sobre",
        "    svyglm (continuas). Mismo motor que rao_scott_summary.csv del pipeline.",
        "  - Una sola variante (journal Latam). Sin réplica estilo NHANES estricto.",
    ]
    for i, line in enumerate(notas_decisiones, start=5):
        ws.cell(row=i, column=1, value=line)

    ws.cell(row=13, column=1, value="Datos fuente").font = Font(bold=True)
    ws.cell(row=14, column=1, value="  - imputation_01.parquet (primera imputación de 20). Pooling completo bajo solicitud.")
    ws.cell(row=15, column=1, value="  - Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE).")
    ws.cell(row=16, column=1, value="  - options(survey.lonely.psu='adjust').")

    ws.cell(row=18, column=1, value="Convenciones de la tabla").font = Font(bold=True)
    ws.cell(row=19, column=1, value="  - Categóricas: 'n no ponderado (% ponderado)'.")
    ws.cell(row=20, column=1, value="  - Continuas: 'media (DE)' ambas ponderadas.")
    ws.cell(row=21, column=1, value="  - p < 0,001 mostrado como '< 0,001'; resto con 3 decimales y coma decimal.")
    ws.cell(row=22, column=1, value="  - QS201 y ALCOHOL_PROBLEMATICO mantienen categoría 'No aplica' (salto del cuestionario).")

    ws.cell(row=24, column=1, value="Coexiste con (no reemplaza)").font = Font(bold=True)
    ws.cell(row=25, column=1, value="  - Tabla_1_v2.xlsx (Auditoría Fase C, RPMESP detallado).")
    ws.cell(row=26, column=1, value="  - Tabla_1.xlsx en Para Publicar Redisenado/Principal/.")

    ws.column_dimensions["A"].width = 110

    # --- Tabla principal ----------------------------------------------------
    ws = wb.create_sheet("Tabla_1_v2_5")

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    panel_men_fill = PatternFill("solid", fgColor="E7F0F8")
    panel_women_fill = PatternFill("solid", fgColor="FCE7E7")
    block_fill = PatternFill("solid", fgColor="F2F2F2")
    cont_fill = PatternFill("solid", fgColor="FFF6E5")
    bold = Font(bold=True)
    italic9 = Font(italic=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    indent_left = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center")

    # Geometría: col 1 = característica; cols 2..6 = severidad H; col 7 = p_H;
    # cols 8..12 = severidad M; col 13 = p_M. Total = 13 columnas.
    n_total_unw = sum(cell_totals_unw.values())

    # Fila 1: título
    ws.cell(row=1, column=1, value=(
        f"Tabla 1 v2.5. Características basales según sexo y severidad de síntomas depresivos (PHQ-9), "
        f"ENDES 2019-2024. n total no ponderado = {n_total_unw:,} (1.ª imputación MICE)."
    ).replace(",", " "))
    ws.cell(row=1, column=1).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 34

    # Fila 2: panel headers
    ws.cell(row=2, column=1, value="").fill = header_fill
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    c = ws.cell(row=2, column=2, value="Hombres")
    c.font = bold; c.alignment = center; c.fill = panel_men_fill
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)
    c = ws.cell(row=2, column=8, value="Mujeres")
    c.font = bold; c.alignment = center; c.fill = panel_women_fill

    # Fila 3: severidad + p
    c = ws.cell(row=3, column=1, value="Característica")
    c.font = bold; c.alignment = left; c.fill = header_fill
    col_idx = 2
    for sex in SEX_ORDER:
        panel_fill = panel_men_fill if sex == "Hombres" else panel_women_fill
        for sev in SEV_LEVELS:
            label_main, phq_range = SEV_HEADERS[sev]
            n_unw = cell_totals_unw[(sex, sev)]
            header_text = f"{label_main}\n(PHQ-9 {phq_range})\nn = {n_unw:,}".replace(",", " ")
            c = ws.cell(row=3, column=col_idx, value=header_text)
            c.font = bold; c.alignment = center; c.fill = panel_fill
            col_idx += 1
        c = ws.cell(row=3, column=col_idx, value="p")
        c.font = bold; c.alignment = center; c.fill = panel_fill
        col_idx += 1
    ws.row_dimensions[3].height = 48

    # Cuerpo: continuas primero, luego categóricas
    row_cursor = 4

    # --- Continuas: una fila por variable ----------------------------------
    for col, label, unidades in CONTINUOUS_VARS:
        c = ws.cell(row=row_cursor, column=1, value=f"{label} ({unidades}), media (DE)")
        c.font = bold; c.alignment = left; c.fill = cont_fill
        col_idx = 2
        for sex in SEX_ORDER:
            for sev in SEV_LEVELS:
                mask = (df["SEXO_LBL"] == sex) & (df["SEVERIDAD_DEPRESIVA"] == sev)
                m, s = weighted_mean_sd(df.loc[mask, col].to_numpy(), df.loc[mask, "PESO_FINAL"].to_numpy())
                cc = ws.cell(row=row_cursor, column=col_idx, value=_fmt_mean_sd(m, s))
                cc.alignment = right; cc.fill = cont_fill
                col_idx += 1
            p = pvalues.get((col, sex), float("nan"))
            cp = ws.cell(row=row_cursor, column=col_idx, value=_fmt_p(p))
            cp.alignment = center; cp.font = bold; cp.fill = cont_fill
            col_idx += 1
        row_cursor += 1

    # --- Categóricas: bloque (header + sub-filas) --------------------------
    for col, label, mapper, levels in CATEGORICAL_VARS:
        # Header de variable + p por sexo
        c = ws.cell(row=row_cursor, column=1, value=label)
        c.font = bold; c.fill = block_fill; c.alignment = left
        col_idx = 2
        for sex in SEX_ORDER:
            for _ in SEV_LEVELS:
                ws.cell(row=row_cursor, column=col_idx).fill = block_fill
                col_idx += 1
            p = pvalues.get((col, sex), float("nan"))
            cp = ws.cell(row=row_cursor, column=col_idx, value=_fmt_p(p))
            cp.alignment = center; cp.font = bold; cp.fill = block_fill
            col_idx += 1
        row_cursor += 1

        # Pre-agregar: n_unweighted y % weighted por (sexo, severidad, nivel)
        lbl_col = f"{col}_LBL"
        agg_w = df.groupby(["SEXO_LBL", "SEVERIDAD_DEPRESIVA", lbl_col], observed=True)["PESO_FINAL"].sum().reset_index()
        agg_unw = df.groupby(["SEXO_LBL", "SEVERIDAD_DEPRESIVA", lbl_col], observed=True).size().reset_index(name="n_unw")

        for level in levels:
            c = ws.cell(row=row_cursor, column=1, value=f"   {level}")
            c.alignment = indent_left
            col_idx = 2
            for sex in SEX_ORDER:
                for sev in SEV_LEVELS:
                    mask_w = (agg_w["SEXO_LBL"] == sex) & (agg_w["SEVERIDAD_DEPRESIVA"] == sev) & (agg_w[lbl_col] == level)
                    mask_unw = (agg_unw["SEXO_LBL"] == sex) & (agg_unw["SEVERIDAD_DEPRESIVA"] == sev) & (agg_unw[lbl_col] == level)
                    n_w = float(agg_w.loc[mask_w, "PESO_FINAL"].iloc[0]) if mask_w.any() else 0.0
                    n_unw = int(agg_unw.loc[mask_unw, "n_unw"].iloc[0]) if mask_unw.any() else 0
                    total_w = cell_totals_w[(sex, sev)]
                    pct = (n_w / total_w * 100) if total_w > 0 else 0.0
                    cc = ws.cell(row=row_cursor, column=col_idx, value=_fmt_n_pct(n_unw, pct))
                    cc.alignment = right
                    col_idx += 1
                col_idx += 1  # saltar columna p
            row_cursor += 1

    # Notas al pie
    row_cursor += 1
    c = ws.cell(row=row_cursor, column=1, value="Notas:")
    c.font = bold; c.alignment = Alignment(horizontal="left", vertical="top")
    row_cursor += 1
    notes = [
        "Continuas: media (DE) ambas ponderadas por PESO_FINAL. Categóricas: n no ponderado (porcentaje ponderado).",
        "Encabezados de columna: n no ponderado de cada estrato (sexo × severidad).",
        "Severidad PHQ-9: Mínima 0–4, Leve 5–9, Moderada 10–14, Mod. severa 15–19, Severa 20–27.",
        "PHQ-9: Patient Health Questionnaire-9.",
        "p: Rao-Scott (svychisq estadístico F) para categóricas; svyglm + regTermTest para continuas. Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE) con survey.lonely.psu='adjust'.",
        "Pesos: PESO_FINAL multianual ENDES 2019-2024 (divisor=6). Datos: primera imputación MICE (m=1 de 20).",
        "Niveles 'No aplica' en consumo de alcohol corresponden a saltos estructurales del cuestionario (no imputables).",
    ]
    for note in notes:
        c = ws.cell(row=row_cursor, column=1, value=f"  {note}")
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=13)
        row_cursor += 1

    # Anchos
    ws.column_dimensions["A"].width = 46
    for cc in range(2, 7):
        ws.column_dimensions[get_column_letter(cc)].width = 14
    ws.column_dimensions[get_column_letter(7)].width = 9
    for cc in range(8, 13):
        ws.column_dimensions[get_column_letter(cc)].width = 14
    ws.column_dimensions[get_column_letter(13)].width = 9
    ws.freeze_panes = "B4"

    return wb


def main() -> None:
    print("Tabla 1 v2.5 (5 niveles, mean(SD), Rao-Scott, journal Latam):")
    print(f"  cargando {IMP_PATH.relative_to(ROOT)}")
    df = load_and_prepare()
    print(f"  n filas = {len(df):,}")

    print("  totales (n no ponderado) por (sexo, severidad):")
    for sex in SEX_ORDER:
        line = "    " + sex + ": "
        parts = []
        for sev in SEV_LEVELS:
            mask = (df["SEXO_LBL"] == sex) & (df["SEVERIDAD_DEPRESIVA"] == sev)
            parts.append(f"{sev}={int(mask.sum()):,}".replace(",", " "))
        print(line + ", ".join(parts))

    print("  computando Rao-Scott p-values (esto puede tomar 1-2 minutos)...")
    pvalues = compute_rao_scott_pvalues(df)
    print(f"    {len(pvalues)} p-values calculados")

    wb = build_workbook(df, pvalues)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"  -> {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
