"""Tabla_1_v3.xlsx — Layout Sexo × PAE (Sin/Con), reusando Rao-Scott pre-computado.

Decisiones acordadas (2026-05-29):
- Layout: 2 paneles (Hombres, Mujeres), cada uno dividido en Sin PAE / Con PAE.
  Una sola columna 'p' a la derecha (no estratificada por sexo).
- p-value para categóricas: reusa data/output/analysis/tables/rao_scott_summary.csv
  (mediana de p Rao-Scott pooled sobre 20 imputaciones). Coherente con table2_bivariate_pooled.
- p-value para 3 continuas + DX_HTA_PREVIO: calculados frescos vía svyglm + regTermTest
  (no estaban pre-computados; 1.ª imputación).
- Continuas en una fila con media (DE) ponderada.
- Categóricas con n no ponderado (% ponderado).

Coexiste con Tabla_1_v2.xlsx (Fase C) y Tabla_1_v2_5.xlsx (prototipo Sexo × Severidad).
NO reemplaza ninguna.
"""
from __future__ import annotations

import subprocess
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from etiquetas_educacion import EDU_LABELS_ES, EDU_ORDER_ES
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
IMP_PATH = ROOT / "data" / "output" / "analysis" / "imputed" / "imputation_01.parquet"
RAO_SCOTT_POOLED = ROOT / "data" / "output" / "analysis" / "tables" / "rao_scott_summary.csv"
OUT_PATH = ROOT / "data" / "output" / "Auditoria_Integral" / "Tabla_1_v3.xlsx"
RSCRIPT_EXE = Path("C:/Program Files/R/R-4.5.3/bin/Rscript.exe")
R_LIB_USER = Path("C:/Users/Trabajo/Documents/R/win-library/4.5")


SEX_ORDER = ["Hombres", "Mujeres"]
PAE_ORDER = ["Sin PAE", "Con PAE"]


# Variables categóricas y sus mapeos a etiquetas legibles
CATEGORICAL_VARS = [
    (
        "SEVERIDAD_DEPRESIVA", "Severidad PHQ-9 (exposición principal)",
        {"Minima": "Mínima (0–4)", "Leve": "Leve (5–9)",
         "Moderada": "Moderada (10–14)", "Mod_Severa": "Mod. severa (15–19)",
         "Severa": "Severa (20–27)"},
        ["Mínima (0–4)", "Leve (5–9)", "Moderada (10–14)", "Mod. severa (15–19)", "Severa (20–27)"],
    ),
    ("HV025", "Área de residencia", {1: "Urbano", 2: "Rural"}, ["Urbano", "Rural"]),
    (
        "HV270", "Quintil de riqueza",
        {1: "Q1 — Más pobre", 2: "Q2", 3: "Q3", 4: "Q4", 5: "Q5 — Más rico"},
        ["Q1 — Más pobre", "Q2", "Q3", "Q4", "Q5 — Más rico"],
    ),
    (
        # Etiquetas corregidas (auditoria 2026-07-16): codificacion real 0..5 del
        # diccionario INEI; fuente unica en etiquetas_educacion.py.
        "QS25N", "Nivel educativo",
        dict(EDU_LABELS_ES),
        list(EDU_ORDER_ES),
    ),
    (
        # Codificación real del pipeline (auditoría 2026-07-16): 0 = sin violencia;
        # 1 = con violencia FÍSICA (QS710/QS711 son actos físicos); 2 = sin pareja.
        "VIOLENCIA_PAREJA", "Violencia de pareja (último año)",
        {0: "Sin violencia", 1: "Con violencia (física)", 2: "Sin pareja"},
        ["Sin violencia", "Con violencia (física)", "Sin pareja"],
    ),
    (
        "ALCOHOL_PROBLEMATICO", "Consumo problemático de alcohol",
        {0: "No", 1: "Sí", "NA": "No aplica (no consume)"},
        ["No", "Sí", "No aplica (no consume)"],
    ),
    (
        "QS201", "Consumo de tabaco (últimos 30 días)",
        {1: "Sí (fumó últimos 30 días)", 2: "No fumó", "NA": "No aplica (no fumó últimos 12 meses)"},
        ["Sí (fumó últimos 30 días)", "No fumó", "No aplica (no fumó últimos 12 meses)"],
    ),
    (
        "QS109", "Diagnóstico de diabetes",
        {1: "Sí (diabetes)", 2: "No"},
        ["Sí (diabetes)", "No"],
    ),
    (
        "CALIDAD_DIETA", "Calidad de la dieta",
        {0: "No adecuada", 1: "Adecuada"},
        ["Adecuada", "No adecuada"],
    ),
    (
        "DX_HTA_PREVIO", "Diagnóstico previo de HTA",
        {0: "No", 1: "Sí"},
        ["No", "Sí"],
    ),
]

# Variables continuas
CONTINUOUS_VARS = [
    ("QS23", "Edad", "años"),
    ("IMC", "Índice de masa corporal", "kg/m²"),
    ("QS907", "Circunferencia abdominal", "cm"),
]

# Cuáles tienen p en rao_scott_summary.csv (pooled sobre 20 imp).
# El resto se calcula fresh.
P_FROM_POOLED = {
    "SEVERIDAD_DEPRESIVA", "HV025", "HV270", "QS25N",
    "VIOLENCIA_PAREJA", "ALCOHOL_PROBLEMATICO", "QS201",
    "QS109", "CALIDAD_DIETA",
}
# QSSEXO también está en el pool pero NO la incluimos como fila (es el estratificador)
P_FRESH = {"DX_HTA_PREVIO", "QS23", "IMC", "QS907"}


# ---------------------------------------------------------------------------
# Carga y preparación
# ---------------------------------------------------------------------------

def load_and_prepare() -> pd.DataFrame:
    df = pd.read_parquet(IMP_PATH)
    df["SEXO_LBL"] = df["QSSEXO"].astype(int).map({1: "Hombres", 2: "Mujeres"})
    df["PAE_LBL"] = df["PRESION_ARTERIAL_ELEVADA"].astype(int).map({0: "Sin PAE", 1: "Con PAE"})

    for col, _label, mapper, _levels in CATEGORICAL_VARS:
        out_col = f"{col}_LBL"
        series = df[col]
        if "NA" in mapper:
            na_label = mapper["NA"]
            mapped = pd.Series(np.nan, index=df.index, dtype=object)
            mapped[series.isna()] = na_label
            for code, lbl in mapper.items():
                if code == "NA":
                    continue
                mapped[series == code] = lbl
        else:
            mapped = series.map(mapper)
        df[out_col] = mapped
    return df


def weighted_mean_sd(x: np.ndarray, w: np.ndarray) -> tuple[float, float]:
    x = np.asarray(x, dtype=float)
    w = np.asarray(w, dtype=float)
    valid = (~np.isnan(x)) & (w > 0) & np.isfinite(w)
    x = x[valid]; w = w[valid]
    if x.size == 0:
        return float("nan"), float("nan")
    mean = float(np.sum(x * w) / np.sum(w))
    var = float(np.sum(w * (x - mean) ** 2) / np.sum(w))
    return mean, float(np.sqrt(var))


# ---------------------------------------------------------------------------
# p-values: pooled + fresh
# ---------------------------------------------------------------------------

def load_pooled_pvalues() -> dict:
    rs = pd.read_csv(RAO_SCOTT_POOLED)
    return {row["variable"]: float(row["median_p_value"]) for _, row in rs.iterrows()}


_RSCRIPT_FRESH = r"""
.libPaths(c('{r_lib_user}', .libPaths()))
suppressPackageStartupMessages({{ library(survey) }})
options(survey.lonely.psu="adjust")

args <- commandArgs(trailingOnly = TRUE)
csv_in <- args[1]; csv_jobs <- args[2]; csv_out <- args[3]
df   <- read.csv(csv_in,   stringsAsFactors = FALSE, fileEncoding = "UTF-8")
jobs <- read.csv(csv_jobs, stringsAsFactors = FALSE, fileEncoding = "UTF-8")

build_design <- function(df) survey::svydesign(
  id=~HV001, strata=~HV022, weights=~PESO_FINAL, data=df, nest=TRUE)

rao_scott_cat <- function(df, var) {{
  d <- build_design(df)
  f <- as.formula(paste0("~", var, " + PRESION_ARTERIAL_ELEVADA"))
  res <- tryCatch(suppressWarnings(survey::svychisq(f, design=d, statistic="F")), error=function(e) NULL)
  if (is.null(res)) return(NA_real_)
  as.numeric(res$p.value)
}}

rao_scott_cont <- function(df, var) {{
  d <- build_design(df)
  f <- as.formula(paste0(var, " ~ factor(PRESION_ARTERIAL_ELEVADA)"))
  fit <- tryCatch(suppressWarnings(survey::svyglm(f, design=d)), error=function(e) NULL)
  if (is.null(fit)) return(NA_real_)
  test <- tryCatch(suppressWarnings(survey::regTermTest(fit, "factor(PRESION_ARTERIAL_ELEVADA)")), error=function(e) NULL)
  if (is.null(test)) return(NA_real_)
  as.numeric(test$p)
}}

results <- data.frame(variable=character(), type=character(), p_value=numeric(), stringsAsFactors=FALSE)
for (i in seq_len(nrow(jobs))) {{
  var <- jobs$variable[i]; type <- jobs$type[i]
  cat(sprintf("[%d/%d] %s (%s)\n", i, nrow(jobs), var, type))
  p <- if (type == "categorical") rao_scott_cat(df, var) else rao_scott_cont(df, var)
  results <- rbind(results, data.frame(variable=var, type=type, p_value=p, stringsAsFactors=FALSE))
}}
write.csv(results, csv_out, row.names=FALSE, fileEncoding="UTF-8")
cat("OK\n")
"""


def compute_fresh_pvalues(df: pd.DataFrame) -> dict:
    """Computa los p-values que NO están pre-computados (DX_HTA_PREVIO + 3 continuas)."""
    jobs_list = [
        {"variable": "DX_HTA_PREVIO_LBL", "type": "categorical", "var_original": "DX_HTA_PREVIO"},
        {"variable": "QS23",   "type": "continuous", "var_original": "QS23"},
        {"variable": "IMC",    "type": "continuous", "var_original": "IMC"},
        {"variable": "QS907",  "type": "continuous", "var_original": "QS907"},
    ]
    jobs_df = pd.DataFrame(jobs_list)
    needed = ["HV001", "HV022", "PESO_FINAL", "PRESION_ARTERIAL_ELEVADA",
              "DX_HTA_PREVIO_LBL", "QS23", "IMC", "QS907"]
    df_for_r = df[needed].copy()
    for col in df_for_r.columns:
        if pd.api.types.is_integer_dtype(df_for_r[col].dtype):
            df_for_r[col] = df_for_r[col].astype("float64")

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        csv_in = tmp_dir / "data.csv"
        csv_jobs = tmp_dir / "jobs.csv"
        csv_out = tmp_dir / "results.csv"
        r_script = tmp_dir / "compute.R"
        df_for_r.to_csv(csv_in, index=False, encoding="utf-8")
        jobs_df[["variable", "type"]].to_csv(csv_jobs, index=False, encoding="utf-8")
        r_script.write_text(
            _RSCRIPT_FRESH.format(r_lib_user=str(R_LIB_USER).replace("\\", "/")),
            encoding="utf-8",
        )
        cmd = [str(RSCRIPT_EXE), "--vanilla", str(r_script),
               str(csv_in), str(csv_jobs), str(csv_out)]
        proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
        if proc.returncode != 0:
            print("STDOUT:", proc.stdout); print("STDERR:", proc.stderr)
            raise RuntimeError("Rscript falló al computar p-values frescos")
        for line in [l for l in proc.stdout.splitlines() if l.strip()][-6:]:
            print(f"    R| {line}")
        results = pd.read_csv(csv_out)

    var_map = {row["variable"]: row["var_original"] for _, row in jobs_df.iterrows()}
    out: dict[str, float] = {}
    for _, row in results.iterrows():
        out[var_map[row["variable"]]] = float(row["p_value"]) if pd.notna(row["p_value"]) else float("nan")
    return out


# ---------------------------------------------------------------------------
# Formateo
# ---------------------------------------------------------------------------

def _fmt_n_pct(n_unw: int, pct_w: float) -> str:
    n_str = f"{int(n_unw):,}".replace(",", " ")
    pct_str = f"{pct_w:.1f}".replace(".", ",")
    return f"{n_str} ({pct_str})"


def _fmt_mean_sd(mean: float, sd: float) -> str:
    if not np.isfinite(mean):
        return "—"
    m_str = f"{mean:.1f}".replace(".", ",")
    s_str = f"{sd:.1f}".replace(".", ",")
    return f"{m_str} ({s_str})"


def _fmt_p(p: float) -> str:
    if p is None or not np.isfinite(p):
        return "—"
    if p < 0.001:
        return "< 0,001"
    return f"{p:.3f}".replace(".", ",")


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def build_workbook(df: pd.DataFrame, p_pooled: dict, p_fresh: dict) -> Workbook:
    # Totales por (sexo, PAE)
    cell_totals_w = {}
    cell_totals_unw = {}
    for sex in SEX_ORDER:
        for pae in PAE_ORDER:
            mask = (df["SEXO_LBL"] == sex) & (df["PAE_LBL"] == pae)
            cell_totals_w[(sex, pae)] = float(df.loc[mask, "PESO_FINAL"].sum())
            cell_totals_unw[(sex, pae)] = int(mask.sum())

    wb = Workbook()

    # --- Portada ------------------------------------------------------------
    ws = wb.active
    ws.title = "Portada"
    ws["A1"] = "Tabla 1 v3 — Características basales por sexo y presión arterial elevada"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Fecha de generación: 2026-05-29"
    ws["A4"] = "Decisiones aplicadas"
    ws["A4"].font = Font(bold=True)
    lines = [
        "  - Layout: 2 paneles (Hombres, Mujeres), cada uno dividido en Sin PAE / Con PAE.",
        "  - Una sola columna 'p' (variable vs PAE en toda la cohorte, no estratificada por sexo).",
        "  - SEVERIDAD_DEPRESIVA (exposición principal) ahora se muestra como FILA, no estratificador.",
        "  - PRESION_ARTERIAL_ELEVADA (desenlace) es el estratificador de columnas.",
        "  - Continuas: media (DE) ponderada.",
        "  - Categóricas: n no ponderado (% ponderado).",
    ]
    for i, line in enumerate(lines, start=5):
        ws.cell(row=i, column=1, value=line)

    ws.cell(row=12, column=1, value="Origen de los p-values").font = Font(bold=True)
    p_lines = [
        "  9 variables categóricas con p REUSADO de rao_scott_summary.csv (mediana pooled sobre 20 imputaciones):",
        "    SEVERIDAD_DEPRESIVA, HV025, HV270, QS25N, VIOLENCIA_PAREJA, ALCOHOL_PROBLEMATICO,",
        "    QS201, QS109, CALIDAD_DIETA.",
        "  4 variables con p CALCULADO fresco (1.ª imputación) por no estar pre-computados:",
        "    DX_HTA_PREVIO, QS23 (Edad), IMC, QS907 (Circunferencia).",
        "  Motor estadístico común: svychisq F-test (categóricas) / svyglm+regTermTest (continuas).",
        "  Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE).",
    ]
    for i, line in enumerate(p_lines, start=13):
        ws.cell(row=i, column=1, value=line)

    ws.cell(row=21, column=1, value="Coexiste con (no reemplaza)").font = Font(bold=True)
    ws.cell(row=22, column=1, value="  - Tabla_1_v2.xlsx (Auditoría Fase C, layout RPMESP).")
    ws.cell(row=23, column=1, value="  - Tabla_1_v2_5.xlsx (prototipo Sexo × Severidad).")
    ws.cell(row=24, column=1, value="  - Tabla_1.xlsx en Para Publicar Redisenado/Principal/.")

    ws.column_dimensions["A"].width = 110

    # --- Tabla principal ----------------------------------------------------
    ws = wb.create_sheet("Tabla_1_v3")

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    panel_men_fill = PatternFill("solid", fgColor="E7F0F8")
    panel_women_fill = PatternFill("solid", fgColor="FCE7E7")
    block_fill = PatternFill("solid", fgColor="F2F2F2")
    cont_fill = PatternFill("solid", fgColor="FFF6E5")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    indent_left = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center")

    n_total_unw = sum(cell_totals_unw.values())

    # Fila 1: título
    ws.cell(row=1, column=1, value=(
        f"Tabla 1 v3. Características basales según sexo y presión arterial elevada (PAE), "
        f"ENDES 2019-2024. n total no ponderado = {n_total_unw:,} (1.ª imputación MICE)."
    ).replace(",", " "))
    ws.cell(row=1, column=1).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 34

    # Fila 2: panel headers
    ws.cell(row=2, column=1, value="").fill = header_fill
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
    c = ws.cell(row=2, column=2, value="Hombres")
    c.font = bold; c.alignment = center; c.fill = panel_men_fill
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
    c = ws.cell(row=2, column=4, value="Mujeres")
    c.font = bold; c.alignment = center; c.fill = panel_women_fill
    c = ws.cell(row=2, column=6, value="")
    c.fill = header_fill

    # Fila 3: Sin PAE / Con PAE + p
    c = ws.cell(row=3, column=1, value="Característica")
    c.font = bold; c.alignment = left; c.fill = header_fill
    col_idx = 2
    for sex in SEX_ORDER:
        panel_fill = panel_men_fill if sex == "Hombres" else panel_women_fill
        for pae in PAE_ORDER:
            n_unw = cell_totals_unw[(sex, pae)]
            header_text = f"{pae}\nn = {n_unw:,}".replace(",", " ")
            c = ws.cell(row=3, column=col_idx, value=header_text)
            c.font = bold; c.alignment = center; c.fill = panel_fill
            col_idx += 1
    c = ws.cell(row=3, column=6, value="p")
    c.font = bold; c.alignment = center; c.fill = header_fill
    ws.row_dimensions[3].height = 36

    # Cuerpo
    row_cursor = 4

    # Continuas
    for col, label, unidades in CONTINUOUS_VARS:
        c = ws.cell(row=row_cursor, column=1, value=f"{label} ({unidades}), media (DE)")
        c.font = bold; c.alignment = left; c.fill = cont_fill
        col_idx = 2
        for sex in SEX_ORDER:
            for pae in PAE_ORDER:
                mask = (df["SEXO_LBL"] == sex) & (df["PAE_LBL"] == pae)
                m, s = weighted_mean_sd(df.loc[mask, col].to_numpy(), df.loc[mask, "PESO_FINAL"].to_numpy())
                cc = ws.cell(row=row_cursor, column=col_idx, value=_fmt_mean_sd(m, s))
                cc.alignment = right; cc.fill = cont_fill
                col_idx += 1
        # p
        p = p_fresh.get(col, float("nan"))
        cp = ws.cell(row=row_cursor, column=6, value=_fmt_p(p))
        cp.alignment = center; cp.font = bold; cp.fill = cont_fill
        # Marca de origen (asterisco para fresco)
        cp.value = f"{cp.value} *" if np.isfinite(p) else cp.value
        row_cursor += 1

    # Categóricas: bloque header + sub-filas
    for col, label, mapper, levels in CATEGORICAL_VARS:
        # Determinar fuente del p
        if col in P_FROM_POOLED:
            p = p_pooled.get(col, float("nan"))
            p_source = "pooled"
        else:
            p = p_fresh.get(col, float("nan"))
            p_source = "fresh"

        # Header de variable
        c = ws.cell(row=row_cursor, column=1, value=label)
        c.font = bold; c.fill = block_fill; c.alignment = left
        for cc_idx in range(2, 6):
            ws.cell(row=row_cursor, column=cc_idx).fill = block_fill
        suffix = " *" if p_source == "fresh" and np.isfinite(p) else ""
        cp = ws.cell(row=row_cursor, column=6, value=f"{_fmt_p(p)}{suffix}")
        cp.alignment = center; cp.font = bold; cp.fill = block_fill
        row_cursor += 1

        # Sub-filas
        lbl_col = f"{col}_LBL"
        agg_w = df.groupby(["SEXO_LBL", "PAE_LBL", lbl_col], observed=True)["PESO_FINAL"].sum().reset_index()
        agg_unw = df.groupby(["SEXO_LBL", "PAE_LBL", lbl_col], observed=True).size().reset_index(name="n_unw")

        for level in levels:
            c = ws.cell(row=row_cursor, column=1, value=f"   {level}")
            c.alignment = indent_left
            col_idx = 2
            for sex in SEX_ORDER:
                for pae in PAE_ORDER:
                    mw = (agg_w["SEXO_LBL"] == sex) & (agg_w["PAE_LBL"] == pae) & (agg_w[lbl_col] == level)
                    mu = (agg_unw["SEXO_LBL"] == sex) & (agg_unw["PAE_LBL"] == pae) & (agg_unw[lbl_col] == level)
                    n_w = float(agg_w.loc[mw, "PESO_FINAL"].iloc[0]) if mw.any() else 0.0
                    n_unw = int(agg_unw.loc[mu, "n_unw"].iloc[0]) if mu.any() else 0
                    total_w = cell_totals_w[(sex, pae)]
                    pct = (n_w / total_w * 100) if total_w > 0 else 0.0
                    cc = ws.cell(row=row_cursor, column=col_idx, value=_fmt_n_pct(n_unw, pct))
                    cc.alignment = right
                    col_idx += 1
            row_cursor += 1

    # Notas al pie
    row_cursor += 1
    c = ws.cell(row=row_cursor, column=1, value="Notas:")
    c.font = bold; c.alignment = Alignment(horizontal="left", vertical="top")
    row_cursor += 1
    notes = [
        "Continuas: media (desviación estándar) ambas ponderadas por PESO_FINAL multianual.",
        "Categóricas: n no ponderado (porcentaje ponderado dentro de la columna sexo × PAE).",
        "Encabezado de columna: n no ponderado de cada estrato (sexo × PAE).",
        "PAE: presión arterial elevada en la medición actual del estudio. PHQ-9: Patient Health Questionnaire-9.",
        "p sin asterisco: Rao-Scott pooled sobre 20 imputaciones, mediana del p-value (rao_scott_summary.csv).",
        "p con asterisco (*): calculado en la 1.ª imputación únicamente, mismo motor estadístico (svychisq / regTermTest).",
        "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE) con survey.lonely.psu='adjust'.",
        "Niveles 'No aplica' en consumo de alcohol corresponden a saltos estructurales del cuestionario (no imputables).",
    ]
    for note in notes:
        c = ws.cell(row=row_cursor, column=1, value=f"  {note}")
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=6)
        row_cursor += 1

    # Anchos
    ws.column_dimensions["A"].width = 46
    for cc in range(2, 6):
        ws.column_dimensions[get_column_letter(cc)].width = 16
    ws.column_dimensions[get_column_letter(6)].width = 12
    ws.freeze_panes = "B4"

    return wb


def main() -> None:
    print("Tabla 1 v3 (Sexo × PAE, p pooled reusado + 4 frescos):")
    print(f"  cargando {IMP_PATH.relative_to(ROOT)}")
    df = load_and_prepare()
    print(f"  n filas = {len(df):,}")

    print("  totales (n no ponderado) por (sexo, PAE):")
    for sex in SEX_ORDER:
        parts = []
        for pae in PAE_ORDER:
            mask = (df["SEXO_LBL"] == sex) & (df["PAE_LBL"] == pae)
            parts.append(f"{pae}={int(mask.sum()):,}".replace(",", " "))
        print(f"    {sex}: {', '.join(parts)}")

    print("  cargando p-values pooled de rao_scott_summary.csv...")
    p_pooled = load_pooled_pvalues()
    print(f"    {len(p_pooled)} p-values pooled disponibles")

    print("  computando 4 p-values frescos (DX_HTA_PREVIO + 3 continuas)...")
    p_fresh = compute_fresh_pvalues(df)
    print(f"    {len(p_fresh)} p-values frescos calculados")

    wb = build_workbook(df, p_pooled, p_fresh)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"  -> {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
