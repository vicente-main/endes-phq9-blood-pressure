"""Tabla suplementaria S6: comparacion de la cohorte incluida final vs los
excluidos del analisis (STROBE).

No modifica el pipeline central. Reconstruye la base armonizada SIN el filtro
final por anio (reutilizando las funciones de `endes_pipeline.pipeline`), aplica
las mascaras STROBE y compara caracteristicas basales disponibles para ambos
grupos entre:

  - INCLUIDOS: pasan todas las mascaras del flujo (cohorte principal = 164 719).
  - EXCLUIDOS DEL ANALISIS: alcanzan la base estructural (adulto, sexo valido,
    diseno disponible, peso > 0, no gestante/puerperio) pero se excluyen por
    PHQ-9 no disponible o PA no valida.

Comparacion descriptiva NO ponderada (convencion STROBE para tablas de no
respuesta). Reproduce los conteos de `filter_flow_by_year.csv`.

Salida: data/output/Post_Auditoria/Suplementario/Tabla_S6_incluidos_excluidos.xlsx
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from scipy import stats

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))

from endes_pipeline import pipeline as P  # noqa: E402

OUT_DIR = ROOT / "data" / "output" / "Post_Auditoria" / "Suplementario"
QC_DIR = ROOT / "data" / "output" / "qc"

# Estilo v3.1
FILL_HEADER = "D9E1F2"
FILL_SECTION = "F2F2F2"

# Variables comparables (disponibles para incluidos y excluidos).
SEX_MAP = {1: "Hombre", 2: "Mujer"}
AREA_MAP = {1: "Urbano", 2: "Rural"}
RIQUEZA_MAP = {1: "Q1 (mas pobre)", 2: "Q2", 3: "Q3", 4: "Q4", 5: "Q5 (mas rico)"}


def _build_harmonized_with_flags(year: int, archive: Path, divisor: float) -> pd.DataFrame:
    """Replica el camino de _process_year hasta harmonized + mascaras, sin filtrar."""
    csalud, _, _ = P._load_required_module(archive, "CSALUD01", P._csalud_cols_for_year(year), year)
    rech0, _, _ = P._load_required_module(archive, "RECH0", P.RECH0_COLS, year)
    rech23, _, _ = P._load_required_module(archive, "RECH23", P.RECH23_COLS, year)
    rec42, _, _ = P._load_required_module(archive, "REC42", P.REC42_COLS, year)
    re223132, _, _ = P._load_required_module(archive, "RE223132", P.RE223132_COLS, year)

    csalud["HHID"] = csalud["HHID"].astype("string").str.strip()
    rech0["HHID"] = rech0["HHID"].astype("string").str.strip()
    rech23["HHID"] = rech23["HHID"].astype("string").str.strip()

    csalud["CASEID_KEY"] = P._build_caseid_key(csalud["HHID"], csalud["QSNUMERO"])
    rec42["CASEID_KEY"] = P._normalize_caseid(rec42["CASEID"])
    re223132["CASEID_KEY"] = P._normalize_caseid(re223132["CASEID"])

    rech0 = rech0.drop_duplicates(subset=["HHID"], keep="first")
    rech23 = rech23.drop_duplicates(subset=["HHID"], keep="first")
    rec42 = rec42.drop_duplicates(subset=["CASEID_KEY"], keep="first")
    re223132 = re223132.drop_duplicates(subset=["CASEID_KEY"], keep="first")

    merged = csalud.merge(rech0, on="HHID", how="left", validate="many_to_one")
    merged = merged.merge(rech23, on="HHID", how="left", validate="many_to_one")
    merged = merged.merge(rec42[["CASEID_KEY", "V454"]], on="CASEID_KEY", how="left", validate="many_to_one")
    merged = merged.merge(
        re223132[["CASEID_KEY", "V222"]], on="CASEID_KEY", how="left", validate="many_to_one"
    )

    harmonized = P._harmonize_dataset(merged, year, divisor)
    masks = P._build_flow_masks(harmonized)

    included = pd.Series(True, index=harmonized.index)
    for mask in masks.values():
        included &= mask.fillna(False)
    structural = P._running_mask_until(masks, "not_puerperal_or_not_mef")

    harmonized = harmonized.copy()
    harmonized["_incluido"] = included.to_numpy()
    harmonized["_estructural"] = structural.to_numpy()
    return harmonized


def _grp_label(row) -> str:
    if row["_incluido"]:
        return "Incluidos"
    if row["_estructural"]:
        return "Excluidos del analisis"
    return "Excluidos estructurales"


def _cont_row(var, label, inc, exc):
    a = pd.to_numeric(inc[var], errors="coerce").dropna()
    b = pd.to_numeric(exc[var], errors="coerce").dropna()
    p = stats.ttest_ind(a, b, equal_var=False).pvalue if len(a) > 1 and len(b) > 1 else np.nan
    return {
        "Caracteristica": label,
        "Nivel": "media (DE)",
        "Incluidos": f"{a.mean():.1f} ({a.std(ddof=1):.1f})".replace(".", ","),
        "Excluidos del analisis": f"{b.mean():.1f} ({b.std(ddof=1):.1f})".replace(".", ","),
        "p": _fmt_p(p),
        "_is_section": True,
    }


def _cat_rows(var, label, mapping, inc, exc):
    levels = [lv for lv in mapping]
    # Categóricas numéricas (códigos ENDES) vs categóricas de texto (p. ej.
    # ALTITUD_CAT3 = "<1500"/"1500-2499"/">=2500"): comparar en el dtype correcto.
    if pd.api.types.is_numeric_dtype(inc[var].dtype):
        inc_v = pd.to_numeric(inc[var], errors="coerce")
        exc_v = pd.to_numeric(exc[var], errors="coerce")
        keys = list(levels)
    else:
        inc_v = inc[var].astype("string")
        exc_v = exc[var].astype("string")
        keys = [str(lv) for lv in levels]
    table = []
    for k in keys:
        table.append([int((inc_v == k).sum()), int((exc_v == k).sum())])
    table = np.array(table)
    p = np.nan
    if table.shape[0] >= 2 and table.sum(axis=0).min() > 0 and (table.sum(axis=1) > 0).all():
        try:
            p = stats.chi2_contingency(table)[1]
        except Exception:
            p = np.nan
    inc_tot = int(inc_v.notna().sum())
    exc_tot = int(exc_v.notna().sum())
    rows = [{"Caracteristica": label, "Nivel": "", "Incluidos": "", "Excluidos del analisis": "",
             "p": _fmt_p(p), "_is_section": True}]
    for i, lv in enumerate(levels):
        ni, ne = table[i]
        rows.append({
            "Caracteristica": "",
            "Nivel": mapping[lv],
            "Incluidos": f"{ni} ({_pct(ni, inc_tot)})",
            "Excluidos del analisis": f"{ne} ({_pct(ne, exc_tot)})",
            "p": "",
            "_is_section": False,
        })
    return rows


def _pct(n, tot):
    return f"{(100.0 * n / tot):.1f}".replace(".", ",") + " %" if tot else "-"


def _fmt_p(p):
    if p is None or (isinstance(p, float) and np.isnan(p)):
        return "-"
    return "< 0,001" if p < 0.001 else f"{p:.3f}".replace(".", ",")


def main() -> None:
    cfg = json.loads((ROOT / "config" / "pipeline_config.json").read_text(encoding="utf-8-sig"))
    years = cfg["years"]
    divisor = cfg["analysis"]["combined_year_divisor"]
    archives = P._find_outer_archives(ROOT / cfg["paths"]["raw_archives_dir"], years)

    frames = []
    for year in years:
        print(f"  reconstruyendo {year} ...")
        frames.append(_build_harmonized_with_flags(year, archives[year], divisor))
    full = pd.concat(frames, ignore_index=True)

    inc = full[full["_incluido"]]
    exc = full[full["_estructural"] & ~full["_incluido"]]

    # Validacion contra STROBE.
    n_inc = len(inc)
    n_exc = len(exc)
    flow = pd.read_csv(QC_DIR / "filter_flow_by_year.csv")
    expected_exc = int(flow["drop_phq_available"].sum() + flow["drop_bp_available"].sum())
    final_counts = pd.read_csv(QC_DIR / "final_counts_by_year.csv")
    expected_inc = int(final_counts["n_final"].sum())
    print(f"  incluidos={n_inc} (esperado {expected_inc}) | excluidos_analisis={n_exc} (esperado {expected_exc})")
    assert n_inc == expected_inc, f"Incluidos {n_inc} != STROBE {expected_inc}"
    assert n_exc == expected_exc, f"Excluidos {n_exc} != STROBE {expected_exc}"

    rows = []
    rows.append(_cont_row("EDAD", "Edad (anios)", inc, exc))
    rows += _cat_rows("QSSEXO", "Sexo", SEX_MAP, inc, exc)
    rows += _cat_rows("HV025", "Area de residencia", AREA_MAP, inc, exc)
    rows += _cat_rows("HV270", "Quintil de riqueza", RIQUEZA_MAP, inc, exc)
    rows += _cat_rows("ANIO", "Anio de encuesta", {y: str(y) for y in years}, inc, exc)
    rows += _cat_rows("ALTITUD_CAT3", "Altitud (m s.n.m.)",
                      {"<1500": "< 1500", "1500-2499": "1500-2499", ">=2500": ">= 2500"}, inc, exc)
    table = pd.DataFrame(rows)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _write_xlsx(table, n_inc, n_exc, OUT_DIR / "Tabla_S6_incluidos_excluidos.xlsx")
    # CSV de respaldo
    table.drop(columns=["_is_section"]).to_csv(
        OUT_DIR / "Tabla_S6_incluidos_excluidos.csv", index=False, encoding="utf-8-sig"
    )
    print(f"  escrito: {OUT_DIR / 'Tabla_S6_incluidos_excluidos.xlsx'}")


def _write_xlsx(table: pd.DataFrame, n_inc: int, n_exc: int, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla S6"

    header_fill = PatternFill("solid", fgColor=FILL_HEADER)
    section_fill = PatternFill("solid", fgColor=FILL_SECTION)
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    title = "Tabla S6. Comparacion de la cohorte incluida frente a los excluidos del analisis (ENDES 2019-2024)"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    headers = ["Caracteristica", "Nivel",
               f"Incluidos\n(n = {n_inc:,})".replace(",", " "),
               f"Excluidos del analisis\n(n = {n_exc:,})".replace(",", " "),
               "p"]
    for col, htext in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=htext)
        c.fill = header_fill
        c.font = bold
        c.alignment = center
        c.border = border

    r = 3
    for _, row in table.iterrows():
        is_section = bool(row["_is_section"])
        values = [row["Caracteristica"], row["Nivel"], row["Incluidos"], row["Excluidos del analisis"], row["p"]]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            c.alignment = left if col <= 2 else center
            if is_section:
                c.fill = section_fill
                if col in (1, 5):
                    c.font = bold
        r += 1

    footer = (
        "Fuente: ENDES 2019-2024 (INEI), modulo CSALUD01. Comparacion descriptiva NO ponderada "
        "(convencion STROBE para analisis de no respuesta). Continuas: media (DE); categoricas: n (% sobre no faltantes). "
        "p: t de Welch (continuas) o chi-cuadrado de Pearson (categoricas). "
        "Incluidos = cohorte que pasa todas las mascaras del flujo STROBE. "
        "Excluidos del analisis = alcanzan la base estructural pero se excluyen por PHQ-9 no disponible o presion arterial no valida. "
        "Los conteos reproducen filter_flow_by_year.csv y final_counts_by_year.csv."
    )
    ws.cell(row=r + 1, column=1, value=footer).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    ws.row_dimensions[r + 1].height = 80

    widths = [26, 22, 20, 24, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A3"

    wb.save(path)


if __name__ == "__main__":
    main()
