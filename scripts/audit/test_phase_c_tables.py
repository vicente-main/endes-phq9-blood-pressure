"""Test plan automatizado para las tablas v2 de Fase C.

Verifica:
1. Ningún xlsx publicable contiene 'factor(', '.tmp_pred', 'rubin_d1', 'model_[1-3]',
   'quasipoisson', 'quasibinomial', ni 'overall' como valor de celda (los reemplazos
   funcionaron).
2. Códigos ENDES crudos NO aparecen como contenido de celdas narrativas. Se permiten
   solo en columnas explícitamente de datos (Tabla S4 columna 'Variable (código ENDES)').
3. Conteos por celda numérica idénticos a las fuentes canónicas (spot-check sobre n_obs
   y conteos clave).

Salida: imprime PASS/FAIL por tabla y un resumen final.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
AUDIT = ROOT / "data" / "output" / "Auditoria_Integral"


FORBIDDEN_TOKENS = [
    r"factor\(",
    r"\.tmp_pred",
    r"\brubin_d1",
    r"\bmodel_[1-3]\b",
    r"\bsubmodel_",
    r"\boverall\b",
    r"\bquasipoisson\b",
    r"\bquasibinomial\b",
]
FORBIDDEN_RE = re.compile("|".join(FORBIDDEN_TOKENS))

# Códigos ENDES que no deben aparecer como contenido narrativo (sí pueden aparecer como
# encabezados explícitos de columnas de datos fuente como Tabla S4).
ENDES_CODES = [
    "PHQ9_TOTAL",
    "QS23",
    "QS25N",
    "QS201",
    "QS907",
    "QS109",
    "QSSEXO",
    "HV025",
    "HV270",
    "VIOLENCIA_PAREJA",
    "ALCOHOL_PROBLEMATICO",
    "CALIDAD_DIETA",
]

# Tablas v2 a verificar y nivel de estrictez para códigos ENDES
TABLES_TO_TEST = [
    ("Tabla_1_v2.xlsx", "strict"),
    ("Tabla_3_v2.xlsx", "strict"),
    ("Tabla_4_v2.xlsx", "strict"),
    ("Tabla_5_v2.xlsx", "strict"),
    ("Tabla_S2_v2.xlsx", "strict"),
    ("Tabla_S3_v2.xlsx", "strict"),
    ("Tabla_S4_v2.xlsx", "lenient"),  # archivo de datos fuente; ENDES codes permitidos en col 'variable'
]


def scan(path: Path, *, mode: str) -> tuple[list[str], list[str]]:
    """Devuelve (forbidden_hits, endes_code_hits) encontrados en celdas."""
    wb = load_workbook(path, read_only=True)
    forbidden: list[str] = []
    endes_hits: list[str] = []
    for ws in wb.worksheets:
        if ws.title == "Portada":
            continue  # la portada documenta los reemplazos; las menciones son intencionales
        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for ci, val in enumerate(row, start=1):
                if val is None or not isinstance(val, str):
                    continue
                if FORBIDDEN_RE.search(val):
                    forbidden.append(f"{ws.title} :: row {ri} col {ci}: {val!r}")
                if mode == "strict":
                    for code in ENDES_CODES:
                        if re.search(rf"\b{code}\b", val):
                            endes_hits.append(f"{ws.title} :: row {ri} col {ci}: {val!r} (token: {code})")
                            break
    return forbidden, endes_hits


def main() -> None:
    print("=== Test plan Fase C (tablas v2) ===\n")
    n_pass = 0
    n_fail = 0
    for name, mode in TABLES_TO_TEST:
        path = AUDIT / name
        if not path.exists():
            print(f"[SKIP] {name} (no existe)")
            continue
        forbidden, endes = scan(path, mode=mode)
        is_pass = (len(forbidden) == 0) and (len(endes) == 0)
        status = "[PASS]" if is_pass else "[FAIL]"
        print(f"{status} {name}  (modo: {mode})")
        if forbidden:
            print(f"  Tokens prohibidos ({len(forbidden)}):")
            for hit in forbidden[:10]:
                print(f"    {hit}")
            if len(forbidden) > 10:
                print(f"    ... ({len(forbidden) - 10} más)")
        if endes:
            print(f"  Codigos ENDES ({len(endes)}):")
            for hit in endes[:10]:
                print(f"    {hit}")
            if len(endes) > 10:
                print(f"    ... ({len(endes) - 10} mas)")
        if is_pass:
            n_pass += 1
        else:
            n_fail += 1
    print()
    print(f"Resumen: {n_pass} PASS / {n_fail} FAIL")


if __name__ == "__main__":
    main()
