"""E-values (VanderWeele & Ding, Ann Intern Med 2017) para los estimadores clave
de PHQ-9 -> presion arterial elevada / cascada de cuidado.

El E-value es la fuerza minima de asociacion (en escala de razon de riesgo) que un
confusor no medido tendria que tener TANTO con la exposicion COMO con el desenlace,
por encima de los confusores ya medidos, para explicar por completo la asociacion
observada. Cuanto mas cercano a 1, mas fragil/explicable es la senal.

Consume los outputs canonicos pooled (no recomputa modelos):
  - data/output/analysis/models/table3_main_models.csv      (model_2, model_3)
  - data/output/analysis/models/interactions_and_sensitivity_models.csv (S7 sin altitud via hierarchical)
  - data/output/analysis/models/hierarchical_decomposition.csv (h2 sin altitud)
  - data/output/analysis/models/table4_cascade_models.csv   (submodel_domain_bp, submodel_adherence)
  - data/output/analysis/models/logistic_sensitivity_results.csv

Salida: data/output/Post_Auditoria/Suplementario/Tabla_S9_evalues.csv (utf-8-sig)
"""
from __future__ import annotations

import math
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
MODELS = ROOT / "data" / "output" / "analysis" / "models"
OUT = ROOT / "data" / "output" / "Post_Auditoria" / "Suplementario" / "Tabla_S9_evalues.csv"
OUT_XLSX = ROOT / "data" / "output" / "Post_Auditoria" / "Suplementario" / "Tabla_S9_evalues.xlsx"


def _coma(x: float, dec: int) -> str:
    return f"{x:.{dec}f}".replace(".", ",")


def evalue_point(rr: float) -> float:
    """E-value para una estimacion puntual RR (ratio). Maneja RR<1 por inversion."""
    if rr <= 0:
        raise ValueError("RR debe ser > 0")
    if rr < 1:
        rr = 1.0 / rr
    return rr + math.sqrt(rr * (rr - 1.0))


def evalue_ci(lo: float, hi: float) -> float:
    """E-value para el limite del IC mas cercano al nulo (1).

    Si el IC cruza 1, el E-value del IC es 1 (no se requiere confusion para ser
    compatible con el nulo)."""
    if lo <= 1.0 <= hi:
        return 1.0
    # Limite mas cercano al nulo.
    bound = hi if hi < 1.0 else lo
    return evalue_point(bound)


def evalue_or_common(est: float) -> float:
    """E-value de un OR con desenlace COMUN: se aproxima RR ~ sqrt(OR) antes de aplicar
    la formula (VanderWeele y Ding 2017; equivale a EValue::evalues.OR(rare=FALSE))."""
    return evalue_point(math.sqrt(est))


def evalue_ci_or_common(lo: float, hi: float) -> float:
    return evalue_ci(math.sqrt(lo), math.sqrt(hi))


def _phq(df: pd.DataFrame, model: str) -> tuple[float, float, float]:
    r = df[(df["model"] == model) & (df["term"] == "PHQ9_TOTAL")].iloc[0]
    return float(r["exp_estimate"]), float(r["exp_ci_low"]), float(r["exp_ci_high"])


def main() -> None:
    t3 = pd.read_csv(MODELS / "table3_main_models.csv")
    t4 = pd.read_csv(MODELS / "table4_cascade_models.csv")
    hier = pd.read_csv(MODELS / "hierarchical_decomposition.csv")
    logit = pd.read_csv(MODELS / "logistic_sensitivity_results.csv")

    rows = []

    def add(label, rr, lo, hi, note="", or_common=False):
        if or_common:
            ev_pt, ev_ci = evalue_or_common(rr), evalue_ci_or_common(lo, hi)
        else:
            ev_pt, ev_ci = evalue_point(rr), evalue_ci(lo, hi)
        rows.append({
            "estimador": label,
            "PR_OR": round(rr, 4),
            "IC95_inf": round(lo, 4),
            "IC95_sup": round(hi, 4),
            "E_value_estimacion": round(ev_pt, 3),
            "E_value_IC": round(ev_ci, 3),
            "nota": note,
        })

    rr, lo, hi = _phq(t3, "model_2")
    add("Modelo 2 (principal, con altitud) - RP por punto PHQ-9", rr, lo, hi,
        "El IC incluye 1: no se requiere confusion no medida para ser compatible con el nulo.")

    # Modelo 2 preespecificado SIN altitud = h2_estructural_sin_altitud
    r = hier[(hier["model"] == "h2_estructural_sin_altitud") & (hier["term"] == "PHQ9_TOTAL")].iloc[0]
    add("Modelo 2 preespecificado SIN altitud (Tabla S7) - RP por punto PHQ-9",
        float(r["exp_estimate"]), float(r["exp_ci_low"]), float(r["exp_ci_high"]),
        "La altitud (confusor con direcciones opuestas) basta para explicar esta senal inversa.")

    rr, lo, hi = _phq(t3, "model_3")
    add("Modelo 3 (exploratorio, con mediadores+altitud) - RP por punto PHQ-9", rr, lo, hi,
        "Exploratorio; incluye potenciales mediadores.")

    rr, lo, hi = _phq(t4, "submodel_domain_bp")
    add("Cascada: PA elevada en DX HTA previo (submodel_domain_bp) - RP por punto PHQ-9", rr, lo, hi,
        "Subpoblacion diagnosticada; senal inversa (sesgo de deteccion/cuidado).")

    rr, lo, hi = _phq(t4, "submodel_adherence")
    add("Cascada: no adherencia en DX HTA previo (submodel_adherence) - RP por punto PHQ-9", rr, lo, hi,
        "Asociacion nula (p = 0,966); E-value ~ 1.")

    r = logit[(logit["model"] == "model_2_logistic_sensitivity") & (logit["term"] == "PHQ9_TOTAL")].iloc[0]
    add("Sensibilidad logistica (OR) - por punto PHQ-9",
        float(r["exp_estimate"]), float(r["exp_ci_low"]), float(r["exp_ci_high"]),
        "Desenlace comun (~16%): E-value calculado sobre sqrt(OR) (VanderWeele y Ding 2017).",
        or_common=True)

    out = pd.DataFrame(rows)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    out.to_csv(OUT, index=False, encoding="utf-8-sig")
    print(out.to_string(index=False))
    print(f"\n-> {OUT}")

    # Etiquetas y medida (RP/OR) e interpretacion para la version publicable.
    meta = [
        ("Modelo 2 (principal, con altitud)", "RP", "El IC incluye 1: ninguna confusión no medida es necesaria para ser compatible con la ausencia de efecto."),
        ("Modelo 2 preespecificado, sin altitud (Tabla S7)", "RP", "Un confusor tan débil como la altitud basta para explicar la señal inversa."),
        ("Modelo 3 (exploratorio, con mediadores)", "RP", "Exploratorio; incluye potenciales mediadores."),
        ("Cascada: PA elevada en pacientes con dx previo de HTA", "RP", "Subpoblación diagnosticada; señal inversa compatible con sesgo de detección/cuidado."),
        ("Cascada: no adherencia en pacientes con dx previo de HTA", "RP", "Asociación nula (p = 0,97); E-value próximo a 1."),
        ("Sensibilidad logística (OR)", "OR", "Desenlace común (~16 %): E-value calculado sobre la raíz de OR."),
    ]
    _build_xlsx(rows, meta)
    print(f"-> {OUT_XLSX}")


def _build_xlsx(rows, meta):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HEAD = PatternFill("solid", fgColor="D9E1F2")
    GRAY = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla S9"
    ncol = 6

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(1, 1, "Tabla S9. E-values de los estimadores principales de la asociación PHQ-9 → presión "
                      "arterial elevada (ENDES 2019-2024)")
    t.font = Font(bold=True, size=12)
    t.alignment = left

    headers = ["Estimador", "Medida", "RP/OR (IC 95 %)", "E-value\n(estimación)",
               "E-value\n(límite IC)", "Interpretación"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(3, c, h)
        cell.font = Font(bold=True)
        cell.fill = HEAD
        cell.alignment = cen
        cell.border = BORDER

    for i, (row, (label, medida, interp)) in enumerate(zip(rows, meta)):
        r = 4 + i
        pr = _coma(row["PR_OR"], 3)
        ci = f"{_coma(row['IC95_inf'], 3)} a {_coma(row['IC95_sup'], 3)}"
        vals = [label, medida, f"{pr} ({ci})", _coma(row["E_value_estimacion"], 2),
                _coma(row["E_value_IC"], 2), interp]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.alignment = left if c in (1, 6) else cen
            cell.border = BORDER
            cell.font = Font(size=11)
            if i == 0:  # resaltar el modelo principal
                cell.fill = GRAY

    foot_r = 4 + len(rows)
    ws.merge_cells(start_row=foot_r, start_column=1, end_row=foot_r, end_column=ncol)
    foot = ws.cell(foot_r, 1,
        "El E-value es la fuerza mínima de asociación (en escala de razón) que un confusor no medido "
        "debería tener tanto con la exposición (PHQ-9) como con el desenlace, por encima de los confusores "
        "ya ajustados, para explicar por completo la asociación observada; el E-value del límite del IC usa "
        "el extremo más cercano a la unidad (es 1,00 cuando el IC incluye 1). RP por cada punto adicional "
        "del PHQ-9. En la sensibilidad logística, con desenlace común (~16 %), el E-value se calcula sobre "
        "√OR. Método: VanderWeele TJ, Ding P. Ann Intern Med. 2017;167:268-274. Fuente: ENDES 2019-2024 (INEI).")
    foot.font = Font(size=8)
    foot.alignment = left

    widths = {1: 46, 2: 9, 3: 22, 4: 13, 5: 13, 6: 40}
    from openpyxl.utils import get_column_letter
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
