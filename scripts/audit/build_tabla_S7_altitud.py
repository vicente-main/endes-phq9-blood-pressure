"""Tabla S7 — Altitud: aporte (descomposición jerárquica), estratificación y
panel de modificación de efecto.

Lee de los outputs canónicos del pipeline (no recalcula):
  - hierarchical_decomposition.csv        (aporte del PR de PHQ-9 por bloque)
  - table3_main_models.csv                (Modelo 2 final, con altitud)
  - altitude_adjusted_models.csv          (robustez: altitud continua km)
  - altitude_stratified_models.csv        (PR de PHQ-9 por estrato de altitud)
  - altitude_strata_adequacy.csv          (n y eventos por estrato)
  - effect_modification_panel.csv         (interacción por modificador + Holm)

Salida: data/output/Post_Auditoria/Suplementario/Tabla_S7_altitud.xlsx
PR/IC con 3 decimales (los efectos por punto de PHQ-9 son pequeños).
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[2]
MODELS = ROOT / "data" / "output" / "analysis" / "models"
OUT = ROOT / "data" / "output" / "Post_Auditoria" / "Suplementario" / "Tabla_S7_altitud.xlsx"

FILL_HEADER = "D9E1F2"
FILL_SECTION = "F2F2F2"


def _n(x):
    return f"{int(round(float(x))):,}".replace(",", " ")


def _pr(est, lo, hi):
    return f"{est:.3f} ({lo:.3f} a {hi:.3f})".replace(".", ",")


def _p(p):
    if pd.isna(p):
        return "-"
    return "< 0,001" if float(p) < 0.001 else f"{float(p):.3f}".replace(".", ",")


def _phq(df, model):
    r = df[(df["model"] == model) & (df["term"] == "PHQ9_TOTAL")].iloc[0]
    return _pr(r["exp_estimate"], r["exp_ci_low"], r["exp_ci_high"]), _p(r["p_value"])


def main() -> None:
    main_m = pd.read_csv(MODELS / "table3_main_models.csv")
    hier = pd.read_csv(MODELS / "hierarchical_decomposition.csv")
    adj = pd.read_csv(MODELS / "altitude_adjusted_models.csv")
    strat = pd.read_csv(MODELS / "altitude_stratified_models.csv")
    adeq = pd.read_csv(MODELS / "altitude_strata_adequacy.csv").set_index("model")
    panel = pd.read_csv(MODELS / "effect_modification_panel.csv")

    # Bloque A: aporte (descomposición jerárquica del PR de PHQ-9).
    a_labels = [
        ("h0_crudo", "Crudo (PHQ-9 sin ajuste)"),
        ("h1_demografia", "+ Edad y sexo"),
        ("h2_estructural_sin_altitud", "+ Educación, área, riqueza, año (= Modelo 2 preespecificado, SIN altitud)"),
        ("h3_estructural_con_altitud", "+ Altitud (= Modelo 2 final, principal)"),
    ]
    rows_a = [(lab, *_phq(hier, m)) for m, lab in a_labels]
    pr_km = _phq(adj, "model_2_altitud_km_robustez")
    rows_a.append(("Robustez: Modelo 2 con altitud continua (por 1 000 m)", *pr_km))

    # Bloque B: estratificado por altitud.
    strata = [
        ("< 2 500 m", "model_2_estrato_lt2500"),
        ("≥ 2 500 m", "model_2_estrato_ge2500"),
        ("< 1 500 m", "model_2_estrato_cat3_lt1500"),
        ("1 500-2 499 m", "model_2_estrato_cat3_1500_2499"),
        ("≥ 2 500 m ", "model_2_estrato_cat3_ge2500"),
    ]
    rows_b = []
    for label, model in strata:
        pr, p = _phq(strat, model)
        nobs = _n(adeq.loc[model, "mean_n_obs"])
        nev = _n(adeq.loc[model, "mean_n_events"])
        rows_b.append((f"{label} (n = {nobs}; eventos = {nev})", pr, p))

    # Bloque C: panel de modificación de efecto.
    mod_lbl = {"sex": "Sexo", "year": "Año", "area": "Área", "riqueza": "Quintil de riqueza",
               "dxhta": "Dx HTA previo", "altitud": "Altitud"}
    rows_c = []
    for _, r in panel.iterrows():
        holm = r.get("p_holm_exploratorios")
        holm_txt = "—" if (r["tipo"] == "preespecificado" or pd.isna(holm)) else _p(holm)
        rows_c.append((
            mod_lbl.get(r["modificador"], r["modificador"]),
            r["tipo"],
            f"{int(r['df_num'])}",
            f"{float(r['statistic']):.2f}".replace(".", ","),
            _p(r["p_value"]),
            holm_txt,
        ))

    _write_xlsx(rows_a, rows_b, rows_c, OUT)
    print(f"  -> {OUT.relative_to(ROOT)}")


def _write_xlsx(rows_a, rows_b, rows_c, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabla_S7"
    hf = PatternFill("solid", fgColor=FILL_HEADER)
    sf = PatternFill("solid", fgColor=FILL_SECTION)
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

    NCOL = 6
    r = 1
    ws.cell(row=r, column=1, value="Tabla S7. Altitud: aporte al estimado, estratificación y modificación de efecto — PHQ-9 / presión arterial elevada (ENDES 2019-2024)").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    r += 2

    def section(title):
        nonlocal r
        c = ws.cell(row=r, column=1, value=title); c.font = bold
        for cc in range(1, NCOL + 1):
            ws.cell(row=r, column=cc).fill = sf
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        r += 1

    def header(cols):
        nonlocal r
        for i, h in enumerate(cols, start=1):
            c = ws.cell(row=r, column=i, value=h); c.fill = hf; c.font = bold; c.alignment = ctr; c.border = bd
        r += 1

    def datarow(vals):
        nonlocal r
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v); c.border = bd
            c.alignment = lft if i == 1 else ctr
        r += 1

    section("A. Aporte al estimado de PHQ-9 (descomposición jerárquica; PR por punto)")
    header(["Bloque de ajuste", "PR PHQ-9 (IC 95 %)", "p"])
    for lab, pr, p in rows_a:
        datarow([lab, pr, p])
    r += 1

    section("B. PR de PHQ-9 (por punto) estratificado por altitud")
    header(["Estrato (n; eventos)", "PR PHQ-9 (IC 95 %)", "p"])
    for lab, pr, p in rows_b:
        datarow([lab, pr, p])
    r += 1

    section("C. Modificación de efecto (interacción PHQ-9 × modificador; Wald conjunto pooled)")
    header(["Modificador", "Tipo", "gl", "Estadístico", "p", "p (Holm, exploratorios)"])
    for vals in rows_c:
        datarow(list(vals))
    r += 1

    footer = (
        "Fuente: ENDES 2019-2024 (INEI). PR = razón de prevalencia (Poisson, link log); 3 decimales por la magnitud pequeña del efecto por punto. "
        "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE); pooling de Rubin (20 imputaciones MICE). "
        "Bloque A: la asociación inversa aparece al ajustar por edad/sexo y se atenúa al añadir altitud (confusor geográfico incorporado durante el análisis). "
        "Bloque B: dentro de cada estrato el Modelo 2 NO incluye el término de altitud (constante). "
        "Bloque C: preespecificados = sexo y año; exploratorios con corrección de Holm. La modificación por altitud (p = 0,017) y por Dx HTA previo (p = 0,001) son las más marcadas; esta última es consistente con sesgo de detección/cuidado. "
        "Altitud = HV040 (completa; no imputada). Todos los modelos conservan factor(ANIO)."
    )
    fc = ws.cell(row=r, column=1, value=footer)
    fc.alignment = Alignment(wrap_text=True, vertical="top"); fc.font = Font(size=8)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    ws.row_dimensions[r].height = 120

    for col, w in zip("ABCDEF", [46, 22, 10, 12, 10, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    main()
