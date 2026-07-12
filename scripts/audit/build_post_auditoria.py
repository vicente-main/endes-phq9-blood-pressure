"""Build Post_Auditoria/ — Paquete editorial unificado en lenguaje visual v3.1.

Replica la estructura de Para_Publicar_Redisenado/ pero con todas las tablas
y figuras actualizadas al estilo decidido en la fase D (Tabla 1 v3.1):

  data/output/Post_Auditoria/
    Principal/
      Figura_1_STROBE.{svg,png,pdf}   ← Figura_1_STROBE_v2 de Auditoria_Integral
      Leyendas_figuras_RPMESP.txt     ← actualizada con DAG
      Tabla_1.xlsx                    ← Tabla_1_v3_1
      Tabla_2.xlsx                    ← Bivariado pooled
      Tabla_3.xlsx                    ← Modelos 1/2/3
      Tabla_4.xlsx                    ← Submodelos cascada
      Tabla_5.xlsx                    ← Sensibilidad narrativa
    Suplementario/
      Figura_S1_Spline.{svg,png,pdf}  ← Figura_S1_Spline_v2 de Auditoria_Integral
      Figura_S2_DAG.{svg,png}         ← 02_DAG_modelos de Auditoria_Integral
      Leyendas_figuras_suplementarias.txt
      Tabla_S1_flujo_STROBE.xlsx      ← STROBE detallado
      Tabla_S2_soporte_tablas.xlsx    ← DEFF + Diagnósticos + Tabla 2 detalle
      Tabla_S3_sensibilidad_interacciones.xlsx  ← Sensibilidad + interacciones + logística
      Tabla_S4_datos_figuras.xlsx     ← Datos fuente de figuras

Cada tabla mantiene su propósito original pero usa el lenguaje visual común:
- Encabezados en azul claro (D9E1F2), bloques de sección en gris (F2F2F2).
- Continuas: media (DE). Categóricas: n (%).
- p con 3 decimales coma o '< 0,001'.
- Etiquetas legibles (sin factor() ni códigos ENDES crudos).
- Footer con notas técnicas + DEFF donde aplique.
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
ANALYSIS = ROOT / "data" / "output" / "analysis"
AUDIT = ROOT / "data" / "output" / "Auditoria_Integral"
OUT = ROOT / "data" / "output" / "Post_Auditoria"
QC = ROOT / "data" / "output" / "qc"


# ---------------------------------------------------------------------------
# Lenguaje visual común (paleta v3.1)
# ---------------------------------------------------------------------------

FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")        # azul claro, headers
FILL_PANEL_MEN = PatternFill("solid", fgColor="E7F0F8")     # azul más claro
FILL_PANEL_WOMEN = PatternFill("solid", fgColor="FCE7E7")   # rosa claro
FILL_BLOCK = PatternFill("solid", fgColor="F2F2F2")         # gris (sección)
FILL_CONT = PatternFill("solid", fgColor="FFF6E5")          # crema (continuas)
FILL_MAIN_MODEL = PatternFill("solid", fgColor="DCE6F1")    # azul medio (Modelo 2 principal)

FONT_BOLD = Font(bold=True)
FONT_ITALIC9 = Font(italic=True, size=9)
FONT_NOTE = Font(size=9)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_INDENT = Alignment(horizontal="left", vertical="center", indent=1)


# ---------------------------------------------------------------------------
# Etiquetas legibles para términos de modelos
# ---------------------------------------------------------------------------

TERM_LABELS: dict[str, str] = {
    "(Intercept)": "Intercepto",
    "PHQ9_TOTAL": "Puntaje PHQ-9 (por punto adicional)",
    "QS23": "Edad (por año adicional)",
    "IMC": "IMC (por kg/m² adicional)",
    "QS907": "Circunferencia abdominal (por cm adicional)",
    # Sexo
    "factor(QSSEXO)2": "Mujer (vs Hombre)",
    # Área
    "factor(HV025)2": "Rural (vs Urbano)",
    # Riqueza
    "factor(HV270)2": "Q2 — Pobre (vs Q1 más pobre)",
    "factor(HV270)3": "Q3 — Medio (vs Q1)",
    "factor(HV270)4": "Q4 — Rico (vs Q1)",
    "factor(HV270)5": "Q5 — Más rico (vs Q1)",
    # Educación
    "factor(QS25N)1": "Sin educación / inicial",
    "factor(QS25N)2": "Primaria",
    "factor(QS25N)3": "Secundaria",
    "factor(QS25N)4": "Superior no universitaria",
    "factor(QS25N)5": "Superior universitaria",
    # Año
    "factor(ANIO)2020": "Año 2020 (vs 2019)",
    "factor(ANIO)2021": "Año 2021 (vs 2019)",
    "factor(ANIO)2022": "Año 2022 (vs 2019)",
    "factor(ANIO)2023": "Año 2023 (vs 2019)",
    "factor(ANIO)2024": "Año 2024 (vs 2019)",
    # VPAR
    "factor(VIOLENCIA_PAREJA)1": "Violencia leve / única",
    "factor(VIOLENCIA_PAREJA)2": "Violencia múltiple / severa",
    # Alcohol problemático
    "factor(ALCOHOL_PROBLEMATICO)1": "Consumo problemático",
    "factor(ALCOHOL_PROBLEMATICO)NaN": "No aplica (no consume)",
    # QS201 consumo de tabaco (últimos 30 días; ENDES Sí=1 / No=2)
    "factor(QS201)2": "No fumó en últimos 30 días",
    "factor(QS201)NaN": "No aplica (no fumó en últimos 12 meses; salto)",
    # QS109 diagnóstico de diabetes (ENDES Sí=1 / No=2)
    "factor(QS109)2": "Sin diagnóstico de diabetes",
    # Calidad dieta
    "factor(CALIDAD_DIETA)1": "Dieta adecuada",
    # Altitud (confusor geográfico; vs < 1 500 m)
    "factor(ALTITUD_CAT3)1500-2499": "Altitud 1 500-2 499 m (vs < 1 500)",
    "factor(ALTITUD_CAT3)>=2500": "Altitud ≥ 2 500 m (vs < 1 500)",
    # Diagnóstico previo de HTA
    "factor(DX_HTA_PREVIO)1": "Diagnóstico previo de HTA (vs sin diagnóstico)",
    # Interacciones
    "PHQ9_TOTAL:factor(QSSEXO)2": "Interacción PHQ-9 × Mujer",
    "PHQ9_TOTAL:factor(ANIO)2020": "Interacción PHQ-9 × Año 2020",
    "PHQ9_TOTAL:factor(ANIO)2021": "Interacción PHQ-9 × Año 2021",
    "PHQ9_TOTAL:factor(ANIO)2022": "Interacción PHQ-9 × Año 2022",
    "PHQ9_TOTAL:factor(ANIO)2023": "Interacción PHQ-9 × Año 2023",
    "PHQ9_TOTAL:factor(ANIO)2024": "Interacción PHQ-9 × Año 2024",
    "PHQ9_TOTAL:factor(HV025)2": "Interacción PHQ-9 × Rural",
    "PHQ9_TOTAL:factor(HV270)2": "Interacción PHQ-9 × Q2",
    "PHQ9_TOTAL:factor(HV270)3": "Interacción PHQ-9 × Q3",
    "PHQ9_TOTAL:factor(HV270)4": "Interacción PHQ-9 × Q4",
    "PHQ9_TOTAL:factor(HV270)5": "Interacción PHQ-9 × Q5",
    "PHQ9_TOTAL:factor(DX_HTA_PREVIO)1": "Interacción PHQ-9 × Dx HTA previo",
    "PHQ9_TOTAL:factor(ALTITUD_CAT3)1500-2499": "Interacción PHQ-9 × Altitud 1 500-2 499 m",
    "PHQ9_TOTAL:factor(ALTITUD_CAT3)>=2500": "Interacción PHQ-9 × Altitud ≥ 2 500 m",
}

MODEL_LABELS: dict[str, str] = {
    "model_1": "Modelo 1 — Crudo",
    "model_2": "Modelo 2 — Estructural (principal)",
    "model_3": "Modelo 3 — Exploratorio (con mediadores)",
    "submodel_adherence": "No adherencia terapéutica (dx HTA previo, n = 14 943)",
    "submodel_domain_bp": "Descontrol tensional (dx HTA previo, n = 14 956)",
    "sensitivity_no_2020": "Sensibilidad: cohorte sin 2020 (n = 144 456)",
    "sensitivity_second_bp_measure": "Sensibilidad: segunda toma de PA",
    "interaction_sex": "Interacción PHQ-9 × Sexo",
    "interaction_year": "Interacción PHQ-9 × Año",
    "interaction_area": "Interacción PHQ-9 × Área",
    "interaction_riqueza": "Interacción PHQ-9 × Quintil de riqueza",
    "interaction_dxhta": "Interacción PHQ-9 × Dx HTA previo",
    "interaction_altitud": "Interacción PHQ-9 × Altitud",
    "model_2_logistic_sensitivity": "Sensibilidad logística (OR, link logit)",
}


# Orden semántico canónico de términos para las tablas wide. La derivación es
# DINÁMICA: a este orden preferido se le anexa cualquier término presente en los
# datos que no figure aquí (evita que una variable nueva —p. ej. altitud— se caiga
# silenciosamente de una tabla wide, como ocurría antes).
TERM_ORDER_PREF = [
    "PHQ9_TOTAL",
    "QS23", "IMC", "QS907",
    "factor(QSSEXO)2",
    "factor(QS25N)1", "factor(QS25N)2", "factor(QS25N)3", "factor(QS25N)4", "factor(QS25N)5",
    "factor(HV025)2",
    "factor(HV270)2", "factor(HV270)3", "factor(HV270)4", "factor(HV270)5",
    "factor(VIOLENCIA_PAREJA)1", "factor(VIOLENCIA_PAREJA)2",
    "factor(DX_HTA_PREVIO)1",
    "factor(ANIO)2020", "factor(ANIO)2021", "factor(ANIO)2022", "factor(ANIO)2023", "factor(ANIO)2024",
    "factor(ALTITUD_CAT3)1500-2499", "factor(ALTITUD_CAT3)>=2500",
    "factor(QS201)2", "factor(QS201)NaN",
    "factor(ALCOHOL_PROBLEMATICO)1", "factor(ALCOHOL_PROBLEMATICO)NaN",
    "factor(CALIDAD_DIETA)1",
    "factor(QS109)2",
    # Términos de interacción
    "PHQ9_TOTAL:factor(QSSEXO)2",
    "PHQ9_TOTAL:factor(ANIO)2020", "PHQ9_TOTAL:factor(ANIO)2021", "PHQ9_TOTAL:factor(ANIO)2022",
    "PHQ9_TOTAL:factor(ANIO)2023", "PHQ9_TOTAL:factor(ANIO)2024",
    "PHQ9_TOTAL:factor(HV025)2",
    "PHQ9_TOTAL:factor(HV270)2", "PHQ9_TOTAL:factor(HV270)3", "PHQ9_TOTAL:factor(HV270)4", "PHQ9_TOTAL:factor(HV270)5",
    "PHQ9_TOTAL:factor(DX_HTA_PREVIO)1",
    "PHQ9_TOTAL:factor(ALTITUD_CAT3)1500-2499", "PHQ9_TOTAL:factor(ALTITUD_CAT3)>=2500",
]


def humanize_term(term: str) -> str:
    """Traduce un término de modelo a su etiqueta legible. Si no hay mapeo, lo deja igual."""
    return TERM_LABELS.get(term, term)


def _build_wide_table(out_path: Path, sheet_name: str, title: str, df: pd.DataFrame,
                      models: list[str], main_model: str | None, value_label: str,
                      portada_titulo: str, tabla_destino: str, fuente: str,
                      notas_portada: list[str], footer: list[str]) -> None:
    """Constructor genérico de tabla 'wide' (modelos en columnas, términos en filas).

    El conjunto de términos se deriva DINÁMICAMENTE de los datos (unión de términos
    presentes en los modelos pedidos, excluyendo el intercepto), ordenado según
    TERM_ORDER_PREF y anexando cualquier término extra. Así la wide siempre contiene
    los mismos términos que la versión stacked.
    """
    df = df[df["model"].isin(models)].copy()
    present = [t for t in dict.fromkeys(df["term"].tolist()) if t != "(Intercept)"]
    ordered = [t for t in TERM_ORDER_PREF if t in present] + [t for t in present if t not in TERM_ORDER_PREF]
    by_key = {(r["model"], r["term"]): r for _, r in df.iterrows()}

    def cell_for(model_name, term):
        r = by_key.get((model_name, term))
        if r is None:
            return "—", "—"
        return f"{fmt_decimal(r['exp_estimate'], 2)} ({fmt_ci(r['exp_ci_low'], r['exp_ci_high'], 2)})", fmt_p(r["p_value"])

    wb = Workbook(); wb.remove(wb.active)
    add_portada(wb, titulo=portada_titulo, tabla_destino=tabla_destino, fuente=fuente, notas=notas_portada)
    ws = wb.create_sheet(sheet_name)
    ncol = 1 + 2 * len(models)
    add_title_row(ws, 1, title, ncol)

    # Doble encabezado: fila 2 = modelos (merge de 2 columnas); fila 3 = métrica / p
    ws.cell(row=2, column=1, value="").fill = FILL_HEADER
    for i, m in enumerate(models):
        base = 2 + i * 2
        ws.merge_cells(start_row=2, start_column=base, end_row=2, end_column=base + 1)
        c = ws.cell(row=2, column=base, value=MODEL_LABELS.get(m, m))
        c.font = FONT_BOLD; c.alignment = ALIGN_CENTER
        c.fill = FILL_MAIN_MODEL if m == main_model else FILL_HEADER

    headers = ["Término"]
    for _ in models:
        headers += [f"{value_label} (IC 95%)", "p"]
    style_header_row(ws, 3, headers)
    if main_model in models:
        i = models.index(main_model)
        for col in (2 + i * 2, 3 + i * 2):
            ws.cell(row=3, column=col).fill = FILL_MAIN_MODEL

    row = 4
    for term in ordered:
        ccl = ws.cell(row=row, column=1, value=humanize_term(term))
        ccl.alignment = ALIGN_LEFT if term == "PHQ9_TOTAL" else ALIGN_INDENT
        if term == "PHQ9_TOTAL":
            ccl.font = FONT_BOLD
        for i, m in enumerate(models):
            pr_ci, p = cell_for(m, term)
            base = 2 + i * 2
            c1 = ws.cell(row=row, column=base, value=pr_ci); c1.alignment = ALIGN_RIGHT
            c2 = ws.cell(row=row, column=base + 1, value=p); c2.alignment = ALIGN_CENTER
            if m == main_model:
                c1.fill = FILL_MAIN_MODEL; c2.fill = FILL_MAIN_MODEL
            if term == "PHQ9_TOTAL":
                c1.font = FONT_BOLD; c2.font = FONT_BOLD
        row += 1

    add_footer_notes(ws, row + 1, footer, ncol)
    ws.column_dimensions["A"].width = 46
    for i in range(len(models)):
        ws.column_dimensions[get_column_letter(2 + i * 2)].width = 22
        ws.column_dimensions[get_column_letter(3 + i * 2)].width = 11
    ws.freeze_panes = "B4"
    wb.save(out_path)
    print(f"  -> {out_path.name}")


# ---------------------------------------------------------------------------
# Formatters
# ---------------------------------------------------------------------------

def fmt_p(p) -> str:
    if p is None or (isinstance(p, float) and (np.isnan(p) or not np.isfinite(p))):
        return "—"
    p = float(p)
    if p < 0.001:
        return "< 0,001"
    return f"{p:.3f}".replace(".", ",")


def fmt_decimal(x, decimals: int = 2) -> str:
    if x is None or (isinstance(x, float) and (np.isnan(x) or not np.isfinite(x))):
        return "—"
    return f"{float(x):.{decimals}f}".replace(".", ",")


def fmt_ci(lo, hi, decimals: int = 2) -> str:
    return f"{fmt_decimal(lo, decimals)} a {fmt_decimal(hi, decimals)}"


def fmt_int_es(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "—"
    return f"{int(x):,}".replace(",", " ")


# ---------------------------------------------------------------------------
# Helpers comunes para construir hojas
# ---------------------------------------------------------------------------

def add_portada(wb: Workbook, *, titulo: str, tabla_destino: str, fuente: str, notas: list[str]) -> None:
    ws = wb.create_sheet("Portada", 0)
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Fecha: 2026-05-29"
    ws["A4"] = "Tabla destino"
    ws["A4"].font = FONT_BOLD
    ws["A5"] = f"  {tabla_destino}"
    ws["A7"] = "Fuente de datos"
    ws["A7"].font = FONT_BOLD
    ws["A8"] = f"  {fuente}"
    ws["A10"] = "Convenciones (lenguaje visual v3.1)"
    ws["A10"].font = FONT_BOLD
    ws["A11"] = "  - Encabezados azul claro (D9E1F2), secciones gris (F2F2F2)."
    ws["A12"] = "  - Continuas: media (DE) ponderadas. Categóricas: n no ponderado (% ponderado)."
    ws["A13"] = "  - PR/OR con 2 decimales y coma; IC 95% como 'X,XX a X,XX'; p con 3 decimales o '< 0,001'."
    ws["A14"] = "  - Etiquetas legibles sin factor(), .tmp_pred*, ni códigos ENDES crudos."
    ws["A15"] = "  - Diseño svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE)."
    ws["A16"] = "  - Pooling sobre 20 imputaciones MICE (Rubin) donde aplique."
    if notas:
        ws["A18"] = "Notas específicas"
        ws["A18"].font = FONT_BOLD
        for i, n in enumerate(notas, start=19):
            ws.cell(row=i, column=1, value=f"  - {n}")
    ws.column_dimensions["A"].width = 115


def add_title_row(ws, row: int, title: str, ncols: int) -> None:
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30


def add_footer_notes(ws, row_start: int, notes: list[str], ncols: int) -> int:
    c = ws.cell(row=row_start, column=1, value="Notas:")
    c.font = FONT_BOLD; c.alignment = Alignment(horizontal="left", vertical="top")
    row = row_start + 1
    for note in notes:
        c = ws.cell(row=row, column=1, value=f"  {note}")
        c.font = FONT_NOTE
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        row += 1
    return row


def style_header_row(ws, row: int, labels: list[str], fills: list[PatternFill] = None) -> None:
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = FONT_BOLD
        c.alignment = ALIGN_CENTER if i > 1 else ALIGN_LEFT
        c.fill = (fills[i-1] if fills else FILL_HEADER)


# ---------------------------------------------------------------------------
# 1) Tabla 1 — copia desde la tabla de DATOS OBSERVADOS (Opción B).
#    Fuente: analysis/tables/tabla1_datos_observados.xlsx
#    (generada por scripts/audit/build_tabla1_observados.py).
#    NO se usa la versión imputada Tabla_1_v3_1.xlsx (decisión editorial 2026-06:
#    la Tabla 1 describe la muestra observada; la imputación se reserva a los modelos).
# ---------------------------------------------------------------------------

def build_tabla_1(out_dir: Path) -> None:
    src = ANALYSIS / "tables" / "tabla1_datos_observados.xlsx"
    if not src.exists():
        raise FileNotFoundError(
            f"No existe {src}. Genera primero la Tabla 1 de datos observados:\n"
            f"  .\\.venv\\Scripts\\python.exe .\\scripts\\audit\\build_tabla1_observados.py"
        )
    dst = out_dir / "Tabla_1.xlsx"
    shutil.copy(src, dst)
    wb = load_workbook(dst)
    if "Tabla_1_observados" in wb.sheetnames:
        wb["Tabla_1_observados"].title = "Tabla_1"
    # El footer correcto (datos observados, faltantes, p complete-case) ya viene
    # del builder fuente; no se reescribe aquí.
    wb.save(dst)
    print(f"  -> Tabla_1.xlsx (datos observados)")


# ---------------------------------------------------------------------------
# 2) Tabla 2 — Bivariado pooled (Rao-Scott)
# ---------------------------------------------------------------------------

VAR_DISPLAY_NAMES = {
    "SEVERIDAD_DEPRESIVA": "Severidad PHQ-9 (5 niveles)",
    "QSSEXO": "Sexo",
    "QS25N": "Nivel educativo",
    "HV025": "Área de residencia",
    "HV270": "Quintil de riqueza",
    "VIOLENCIA_PAREJA": "Violencia de pareja",
    "QS201": "Consumo de tabaco (últimos 30 días)",
    "ALCOHOL_PROBLEMATICO": "Consumo problemático de alcohol",
    "CALIDAD_DIETA": "Calidad de la dieta",
    "QS109": "Diagnóstico de diabetes",
    "ALTITUD_CAT3": "Altitud (m s.n.m.)",
}


def build_tabla_2(out_dir: Path) -> None:
    # Fuente: bivariado sobre DATOS OBSERVADOS (Opción B, complete-case),
    # generado por scripts/audit/build_tabla2_observados.py. NO pooled sobre imputaciones.
    src = ANALYSIS / "tables" / "rao_scott_observado.csv"
    if not src.exists():
        raise FileNotFoundError(
            f"No existe {src}. Genera primero el bivariado de datos observados:\n"
            f"  .\\.venv\\Scripts\\python.exe .\\scripts\\audit\\build_tabla2_observados.py"
        )
    rs = pd.read_csv(src)
    rs["Variable"] = rs["variable"].map(VAR_DISPLAY_NAMES).fillna(rs["variable"])

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla 2 — Asociaciones bivariadas con presión arterial elevada (Rao-Scott pooled)",
        tabla_destino="Tabla_2.xlsx (Principal)",
        fuente="data/output/analysis/tables/rao_scott_summary.csv (pooled sobre 20 imputaciones)",
        notas=[
            "Asociación bivariada de cada característica categórica con PA elevada.",
            "Estadístico: F de Rao-Scott (corregido por diseño complejo).",
            "p: calculado por análisis de casos completos sobre los datos observados (no imputado).",
            "Continuas (Edad, PHQ-9, IMC, Cintura) tienen su prueba bivariada en Tabla 1 (columna 'p').",
        ],
    )
    ws = wb.create_sheet("Tabla_2")
    add_title_row(ws,
        1,
        "Tabla 2. Asociación bivariada de las características con presión arterial elevada, sobre datos observados. ENDES 2019-2024.",
        4,
    )
    style_header_row(ws, 2, ["Característica", "Estadístico F (Rao-Scott)", "p", "n (casos completos)"])

    # Excluir SEVERIDAD_DEPRESIVA — está como exposición principal; mejor mostrar arriba?
    # Mantenerla; lectores quieren ver su asociación bivariada en una mirada.
    order = ["SEVERIDAD_DEPRESIVA", "QSSEXO", "QS25N", "HV025", "HV270",
             "VIOLENCIA_PAREJA", "ALCOHOL_PROBLEMATICO", "QS201",
             "QS109", "CALIDAD_DIETA", "ALTITUD_CAT3"]
    row = 3
    for var in order:
        sub = rs[rs["variable"] == var]
        if sub.empty:
            continue
        r = sub.iloc[0]
        ws.cell(row=row, column=1, value=VAR_DISPLAY_NAMES.get(var, var)).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=fmt_decimal(r["statistic"], 2)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=3, value=fmt_p(r["p_value"])).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4, value=f"{int(r['n_complete']):,}".replace(",", " ")).alignment = ALIGN_CENTER
        row += 1

    add_footer_notes(ws, row + 1, [
        "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE) con survey.lonely.psu='adjust'.",
        "Estadístico F de Rao-Scott con corrección por efecto de diseño, sobre datos observados (análisis de casos completos).",
        "n (casos completos): registros con la variable no faltante usados en la prueba. Para variables continuas (Edad, PHQ-9, IMC, cintura), ver columna 'p' en Tabla 1.",
        "Una p baja indica asociación bivariada cruda; los modelos ajustados (Tabla 3) son los que evalúan asociación con PA controlando por confusores.",
    ], 4)

    ws.column_dimensions["A"].width = 42
    for c in range(2, 5):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.freeze_panes = "A3"
    wb.save(out_dir / "Tabla_2.xlsx")
    print(f"  -> Tabla_2.xlsx")


# ---------------------------------------------------------------------------
# 3, 4, S3 — Tablas de modelos (estilo común)
# ---------------------------------------------------------------------------

def _format_model_row(ws, row: int, term_label: str, pr, ci_lo, ci_hi, p,
                       highlight: bool = False, indent: bool = True) -> None:
    """Una fila de coeficiente: Término | PR | IC95% | p (sin columna n; n va en footer)."""
    fill = FILL_MAIN_MODEL if highlight else None
    cells = [
        (1, term_label, ALIGN_INDENT if indent else ALIGN_LEFT),
        (2, fmt_decimal(pr, 2), ALIGN_RIGHT),
        (3, fmt_ci(ci_lo, ci_hi, 2), ALIGN_RIGHT),
        (4, fmt_p(p), ALIGN_CENTER),
    ]
    for col, val, align in cells:
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = align
        if fill:
            c.fill = fill


def _write_model_block(ws, start_row: int, df_model: pd.DataFrame, model_name: str,
                        skip_intercept: bool = True, highlight_phq: bool = True,
                        ncols: int = 4) -> int:
    """Escribe un bloque de modelo: header de sección + filas de coeficientes."""
    header_label = MODEL_LABELS.get(model_name, model_name)
    c = ws.cell(row=start_row, column=1, value=header_label)
    c.font = FONT_BOLD; c.alignment = ALIGN_LEFT; c.fill = FILL_BLOCK
    for col in range(2, ncols + 1):
        ws.cell(row=start_row, column=col).fill = FILL_BLOCK
    row = start_row + 1

    for _, r in df_model.iterrows():
        term = str(r["term"])
        if skip_intercept and term == "(Intercept)":
            continue
        label = humanize_term(term)
        is_phq = "PHQ9_TOTAL" in term and ":" not in term
        highlight = highlight_phq and is_phq
        _format_model_row(
            ws, row, label,
            r["exp_estimate"], r["exp_ci_low"], r["exp_ci_high"],
            r["p_value"],
            highlight=highlight,
        )
        row += 1
    return row


def build_tabla_3(out_dir: Path) -> None:
    """Tabla 3 — Modelos 1, 2, 3 con PHQ-9 como exposición."""
    df = pd.read_csv(ANALYSIS / "models" / "table3_main_models.csv")

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla 3 — Modelos de razón de prevalencia para PA elevada según PHQ-9",
        tabla_destino="Tabla_3.xlsx (Principal)",
        fuente="data/output/analysis/models/table3_main_models.csv (pooled sobre 20 imputaciones, reglas de Rubin)",
        notas=[
            "Cuasi-Poisson con link log → razón de prevalencia (PR).",
            "Modelo 2 (estructural) es el resultado principal del estudio.",
            "Modelo 3 incluye covariables que pueden ser mediadores; se interpreta como exploratorio.",
            "Filas de PHQ-9 resaltadas en azul medio para destacar la exposición principal.",
        ],
    )
    ws = wb.create_sheet("Tabla_3")
    add_title_row(ws, 1,
        "Tabla 3. Razón de prevalencia (PR) de presión arterial elevada según puntaje PHQ-9 y covariables, "
        "ENDES 2019-2024 (n = 164 719, 20 imputaciones MICE).",
        4,
    )
    style_header_row(ws, 2, ["Término", "PR", "IC 95%", "p"])

    row = 3
    for model_name in ["model_1", "model_2", "model_3"]:
        df_m = df[df["model"] == model_name]
        row = _write_model_block(ws, row, df_m, model_name, ncols=4)

    add_footer_notes(ws, row + 1, [
        "PR: razón de prevalencia (cuasi-Poisson, link log). IC 95% y p combinados por reglas de Rubin sobre 20 imputaciones MICE. n = 164 719 (constante en los tres modelos).",
        "Imputación: las 10 covariables MAR fueron imputadas; la exposición PHQ-9 y el desenlace PA elevada NO se imputaron. Detalle en Métodos del manuscrito.",
        "Modelo 1: PHQ-9 crudo. Modelo 2 (principal): + edad, sexo, educación, área, riqueza, VPAR, año y altitud. Modelo 3 (exploratorio): + IMC, cintura, consumo de tabaco, consumo problemático de alcohol, dieta y diagnóstico de diabetes (potenciales mediadores).",
        "Altitud (factor(ALTITUD_CAT3)) incorporada al ajuste estructural como confusor geográfico (enmienda 2026-06-01). Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE), survey.lonely.psu='adjust'. Intercepto omitido (disponible en CSV fuente).",
    ], 4)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 11
    ws.freeze_panes = "A3"
    wb.save(out_dir / "Tabla_3.xlsx")
    print(f"  -> Tabla_3.xlsx")


def build_tabla_3_wide(out_dir: Path) -> None:
    """Tabla 3 wide — Modelos 1, 2 y 3 lado a lado (estilo JAMA/Lancet). Términos dinámicos."""
    df = pd.read_csv(ANALYSIS / "models" / "table3_main_models.csv")
    _build_wide_table(
        out_dir / "Tabla_3_wide.xlsx", "Tabla_3_wide",
        "Tabla 3 (wide). Razón de prevalencia (PR) según PHQ-9 y covariables, 3 modelos en columnas. ENDES 2019-2024, n = 164 719.",
        df, ["model_1", "model_2", "model_3"], "model_2", "PR",
        "Tabla 3 (formato wide) — Comparación lado a lado de Modelos 1, 2 y 3",
        "Tabla_3_wide.xlsx (Principal, alternativa)",
        "data/output/analysis/models/table3_main_models.csv (pooled 20 imputaciones)",
        [
            "Versión alternativa de Tabla 3 con los 3 modelos en columnas (estilo JAMA / Lancet).",
            "Contiene EXACTAMENTE los mismos términos que la versión stacked (Tabla_3.xlsx); orden de términos dinámico.",
            "Modelo 2 (principal) incluye altitud como confusor estructural. Términos no incluidos en un modelo: '—'.",
        ],
        [
            "PR: razón de prevalencia (cuasi-Poisson, link log). IC 95% y p combinados por Rubin sobre 20 imputaciones MICE. n = 164 719 (constante).",
            "Modelo 2 (principal) destacado en azul medio. Términos en negrita: exposición principal PHQ-9. Altitud incorporada como confusor (enmienda 2026-06-01).",
            "Imputación: 10 covariables MAR; PHQ-9 y PA elevada NO. '—' indica término no incluido en el modelo.",
        ],
    )


def build_tabla_4_wide(out_dir: Path) -> None:
    """Tabla 4 wide — Submodelos de cascada (no adherencia y descontrol) lado a lado."""
    df = pd.read_csv(ANALYSIS / "models" / "table4_cascade_models.csv")
    _build_wide_table(
        out_dir / "Tabla_4_wide.xlsx", "Tabla_4_wide",
        "Tabla 4 (wide). Cascada de cuidado: no adherencia y descontrol de PA, lado a lado. ENDES 2019-2024 (DX HTA previo).",
        df, ["submodel_adherence", "submodel_domain_bp"], None, "PR",
        "Tabla 4 (formato wide) — Cascada de cuidado lado a lado",
        "Tabla_4_wide.xlsx (Principal, alternativa)",
        "data/output/analysis/models/table4_cascade_models.csv (pooled 20 imputaciones)",
        [
            "Versión wide de Tabla 4: los dos submodelos de cascada en columnas para comparar coeficientes.",
            "No adherencia (NO_ADHERENCIA_HTA, n = 14 943) vs descontrol tensional (PA elevada, n = 14 956), en personas con DX previo de HTA.",
            "Mismas covariables que el Modelo 2 (incluida altitud). Términos dinámicos.",
        ],
        [
            "PR: razón de prevalencia (cuasi-Poisson, link log), pooled sobre 20 imputaciones MICE (Rubin).",
            "Subgrupo con DX previo de HTA. Mismas covariables que el Modelo 2 estructural (incluida altitud).",
            "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE). '—' = término no incluido.",
        ],
    )


def build_tabla_S3_wide(out_dir: Path) -> None:
    """Tabla S3 wide — Sensibilidades e interacciones (PR) lado a lado."""
    df = pd.read_csv(ANALYSIS / "models" / "interactions_and_sensitivity_models.csv")
    models = ["sensitivity_no_2020", "sensitivity_second_bp_measure",
              "interaction_sex", "interaction_year", "interaction_area",
              "interaction_riqueza", "interaction_dxhta", "interaction_altitud"]
    _build_wide_table(
        out_dir / "Tabla_S3_wide.xlsx", "Tabla_S3_wide",
        "Tabla S3 (wide). Sensibilidades e interacciones (PR), modelos en columnas. ENDES 2019-2024.",
        df, models, None, "PR",
        "Tabla S3 (formato wide) — Sensibilidades e interacciones lado a lado",
        "Tabla_S3_wide.xlsx (Suplementario, alternativa)",
        "data/output/analysis/models/interactions_and_sensitivity_models.csv (pooled 20 imputaciones)",
        [
            "Versión wide del panel de sensibilidades y modificación de efecto (PR cuasi-Poisson).",
            "Incluye el panel completo de modificación de efecto: sexo, año, área, riqueza, Dx HTA previo y altitud.",
            "El test conjunto de interacción (Wald pooled + Holm) está en Tabla S7-C. La sensibilidad logística (OR) está en Tabla_S3 (hoja b).",
        ],
        [
            "PR: razón de prevalencia (cuasi-Poisson, log), pooled sobre 20 imputaciones MICE (Rubin). Todos ajustados por altitud.",
            "Sensibilidad sin 2020: n = 144 456; resto n = 164 719. Modificadores exploratorios (área, riqueza, Dx HTA, altitud) corregidos por multiplicidad en Tabla S7-C.",
            "'—' indica término no incluido en ese modelo. La hilera de interacción aparece solo en su modelo correspondiente.",
        ],
    )


def build_tabla_4(out_dir: Path) -> None:
    """Tabla 4 — Submodelos de cascada de cuidado."""
    df = pd.read_csv(ANALYSIS / "models" / "table4_cascade_models.csv")

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla 4 — Cascada de cuidado: no adherencia y descontrol de PA en personas con DX previo de HTA",
        tabla_destino="Tabla_4.xlsx (Principal)",
        fuente="data/output/analysis/models/table4_cascade_models.csv (pooled sobre 20 imputaciones)",
        notas=[
            "Subanálisis en el subgrupo con diagnóstico previo de HTA.",
            "No adherencia terapéutica: NO_ADHERENCIA_HTA como desenlace (n = 14 943).",
            "Descontrol tensional: PRESION_ARTERIAL_ELEVADA como desenlace (n = 14 956).",
            "Mismas covariables de ajuste que el Modelo 2 estructural.",
        ],
    )
    ws = wb.create_sheet("Tabla_4")
    add_title_row(ws, 1,
        "Tabla 4. Razón de prevalencia (PR) en cascada de cuidado: no adherencia y descontrol de PA "
        "en personas con DX previo de HTA. ENDES 2019-2024.",
        4,
    )
    style_header_row(ws, 2, ["Término", "PR", "IC 95%", "p"])

    row = 3
    for model_name in ["submodel_adherence", "submodel_domain_bp"]:
        df_m = df[df["model"] == model_name]
        row = _write_model_block(ws, row, df_m, model_name, ncols=4)

    add_footer_notes(ws, row + 1, [
        "PR: razón de prevalencia (cuasi-Poisson, link log). Pooled sobre 20 imputaciones MICE (Rubin).",
        "No adherencia: desenlace = NO_ADHERENCIA_HTA en personas con DX previo (n = 14 943). Descontrol: desenlace = PA elevada en personas con DX previo (n = 14 956).",
        "Mismas covariables que Modelo 2 estructural. Las 10 covariables MAR fueron imputadas; el desenlace NO. Ver Métodos del manuscrito.",
        "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE). Intercepto omitido (CSV fuente).",
    ], 4)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 11
    ws.freeze_panes = "A3"
    wb.save(out_dir / "Tabla_4.xlsx")
    print(f"  -> Tabla_4.xlsx")


def build_tabla_5(out_dir: Path) -> None:
    """Tabla 5 — Resumen de sensibilidades y robustez."""
    df_main = pd.read_csv(ANALYSIS / "models" / "table3_main_models.csv")
    df_int = pd.read_csv(ANALYSIS / "models" / "interactions_and_sensitivity_models.csv")
    df_spline = pd.read_csv(ANALYSIS / "figures" / "spline_nonlinearity_summary.csv")
    df_log = pd.read_csv(ANALYSIS / "models" / "logistic_sensitivity_results.csv")

    def _phq_row(df, model_name):
        sub = df[(df["model"] == model_name) & (df["term"] == "PHQ9_TOTAL")]
        if sub.empty:
            return None
        r = sub.iloc[0]
        return float(r["exp_estimate"]), float(r["exp_ci_low"]), float(r["exp_ci_high"]), float(r["p_value"]), float(r["mean_n_obs"])

    scenarios = []
    # Principal
    r = _phq_row(df_main, "model_2")
    scenarios.append(("Definición principal del desenlace",
                       "Modelo 2 (estructural)", "PR por punto adicional de PHQ-9", *r,
                       "La asociación inversa pequeña se mantiene tras ajustar por confusores estructurales."))
    # No 2020
    r = _phq_row(df_int, "sensitivity_no_2020")
    scenarios.append(("Sin cohorte 2020 (COVID)",
                       "Modelo 2 sin 2020", "PR por punto adicional de PHQ-9", *r,
                       "La dirección y magnitud se preservan al excluir el año 2020 con sub-muestreo COVID."))
    # Second BP
    r = _phq_row(df_int, "sensitivity_second_bp_measure")
    scenarios.append(("Definición con segunda toma de PA",
                       "Sensibilidad segunda toma", "PR por punto adicional de PHQ-9", *r,
                       "Misma dirección con definición alternativa del desenlace."))
    # Logística
    r = _phq_row(df_log, "model_2_logistic_sensitivity")
    scenarios.append(("Sensibilidad logística (OR)",
                       "Modelo 2 con link logit", "OR por punto adicional de PHQ-9", *r,
                       "OR consistente con la PR; descarta sesgo de especificación cuasi-Poisson."))
    # Spline
    spline_p = float(df_spline.iloc[0]["median_p_value"])
    scenarios.append(("Spline cúbica restringida (linealidad)",
                       "Curva spline (PHQ-9)", "p de no-linealidad pooled", None, None, None, spline_p, None,
                       f"No se rechaza linealidad (p = {fmt_p(spline_p)}); el coeficiente lineal del Modelo 2 es interpretable."))

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla 5 — Resumen narrativo de sensibilidades y robustez",
        tabla_destino="Tabla_5.xlsx (Principal)",
        fuente="model_2 (principal) + sensibilidades + spline (todos pooled sobre 20 imputaciones)",
        notas=[
            "Vista ejecutiva del resultado principal y su robustez a especificaciones alternativas.",
            "Si el manuscrito tiene límite editorial estricto, esta tabla puede reemplazar Tabla 4 + Tabla S3 como display único de sensibilidades.",
            "El detalle por covariable de cada modelo está en Tabla 3, Tabla 4 y Tabla S3.",
        ],
    )
    ws = wb.create_sheet("Tabla_5")
    add_title_row(ws, 1,
        "Tabla 5. Resumen de sensibilidades para la asociación entre PHQ-9 y presión arterial elevada. "
        "ENDES 2019-2024 (n principal = 164 719).",
        6,
    )
    style_header_row(ws, 2, ["Escenario", "Modelo", "Parámetro", "Estimación", "IC 95%", "p"])

    row = 3
    for sc in scenarios:
        if len(sc) == 9:
            scenario, model, param, pr, ci_lo, ci_hi, p, n_obs, lectura = sc
        elif len(sc) == 8:
            scenario, model, param, pr, ci_lo, ci_hi, p, lectura = sc[:7] + ("",)
            n_obs = None
            lectura = sc[-1]
        else:
            continue
        ws.cell(row=row, column=1, value=scenario).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=model).alignment = ALIGN_LEFT
        ws.cell(row=row, column=3, value=param).alignment = ALIGN_LEFT
        ws.cell(row=row, column=4, value=fmt_decimal(pr, 2) if pr is not None else "—").alignment = ALIGN_RIGHT
        ws.cell(row=row, column=5, value=fmt_ci(ci_lo, ci_hi, 2) if ci_lo is not None else "—").alignment = ALIGN_RIGHT
        ws.cell(row=row, column=6, value=fmt_p(p)).alignment = ALIGN_CENTER
        row += 1
        # Fila de lectura
        c = ws.cell(row=row, column=1, value=f"   Lectura: {lectura}")
        c.font = FONT_ITALIC9
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 24
        row += 1

    add_footer_notes(ws, row + 1, [
        "Todos los modelos pooled sobre 20 imputaciones MICE (reglas de Rubin).",
        "PR: razón de prevalencia (cuasi-Poisson, log link). OR: odds ratio (logística).",
        "Spline: 4 nodos en PHQ-9 = 0, 4, 9, 14. p de no-linealidad mediana pooled.",
        "Detalle por covariable: Tabla 3 (modelos principales), Tabla 4 (cascada), Tabla S3 (sensibilidades e interacciones).",
    ], 6)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 11
    ws.freeze_panes = "A3"
    wb.save(out_dir / "Tabla_5.xlsx")
    print(f"  -> Tabla_5.xlsx")


# ---------------------------------------------------------------------------
# Tabla S1 — STROBE detallado
# ---------------------------------------------------------------------------

def build_tabla_S1(out_dir: Path) -> None:
    df = pd.read_csv(ROOT / "data" / "output" / "Para Enviar" / "07_figura_1_strobe_datos.csv")

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla S1 — Flujo de selección muestral STROBE",
        tabla_destino="Tabla_S1_flujo_STROBE.xlsx (Suplementario)",
        fuente="data/output/Para Enviar/07_figura_1_strobe_datos.csv (alias de filter_flow_by_year.csv agregado)",
        notas=[
            "Detalle numérico de Figura 1 (Principal).",
            "Las exclusiones reproductivas (gestación, puerperio) aplican únicamente a mujeres de 15-49 años.",
            "2020 con sub-muestreo COVID-19: 8 828 exclusiones por PA no medible (vs 5-314 los demás años).",
        ],
    )
    ws = wb.create_sheet("Tabla_S1")
    add_title_row(ws, 1,
        "Tabla S1. Flujo de selección STROBE detallado. ENDES 2019-2024.", 6,
    )
    style_header_row(ws, 2, ["Paso", "Etapa", "n antes", "n excluidos", "n después", "% retenido"])

    row = 3
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=int(r["order"])).alignment = ALIGN_CENTER
        ws.cell(row=row, column=2, value=str(r["stage"])).alignment = ALIGN_LEFT
        ws.cell(row=row, column=3, value=fmt_int_es(r["n_before"])).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=4, value=fmt_int_es(r["n_excluded"])).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=5, value=fmt_int_es(r["n_after"])).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=6, value=fmt_decimal(r["pct_remaining_from_start"], 1)).alignment = ALIGN_RIGHT
        row += 1

    add_footer_notes(ws, row + 1, [
        "Cohorte estructural: 174 282 (paso 9).",
        "Cohorte principal de análisis: 164 719 (paso 11).",
        "Análisis de sensibilidad sin 2020 (Tabla 5): n = 144 456.",
        "Exclusiones reproductivas: condicionales a mujeres de 15-49 años (FLAG_MEF).",
    ], 6)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    for c in range(3, 7):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "A3"
    wb.save(out_dir / "Tabla_S1_flujo_STROBE.xlsx")
    print(f"  -> Tabla_S1_flujo_STROBE.xlsx")


# ---------------------------------------------------------------------------
# Tabla S2 — Soporte: DEFF + Diagnósticos + Tabla 2 detallada
# ---------------------------------------------------------------------------

def build_tabla_S2(out_dir: Path) -> None:
    deff_df = pd.read_csv(ANALYSIS / "tables" / "table1_deff_summary.csv") if (ANALYSIS / "tables" / "table1_deff_summary.csv").exists() else None
    diag_df = pd.read_csv(ANALYSIS / "models" / "model_diagnostics_summary.csv")
    rao_per = pd.read_csv(ANALYSIS / "tables" / "rao_scott_per_imputation.csv")

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla S2 — Soporte técnico (DEFF, diagnósticos de modelos, bivariados por imputación)",
        tabla_destino="Tabla_S2_soporte_tablas.xlsx (Suplementario)",
        fuente="rao_scott_per_imputation.csv, model_diagnostics_summary.csv, valores DEFF de Tabla 1 v2",
        notas=[
            "DEFF: efecto de diseño, n efectivo equivalente a muestreo aleatorio simple ≈ n / DEFF.",
            "Diagnósticos: familia, link, gl residuales y deviance escalada de cada modelo.",
            "Detalle por imputación de Rao-Scott: 10 variables × 20 imputaciones = 200 filas.",
        ],
    )

    # Hoja DEFF
    ws = wb.create_sheet("DEFF")
    add_title_row(ws, 1, "DEFF — Efectos de diseño globales de la Tabla 1", 3)
    style_header_row(ws, 2, ["Indicador", "Valor", "n efectivo SRS"])
    deff_values = [
        ("DEFF promedio global (Tabla 1)", 3.91, 164719 / 3.91),
        ("DEFF prevalencia de PA elevada", 3.66, 164719 / 3.66),
        ("DEFF media de PHQ-9", 3.17, 164719 / 3.17),
    ]
    row = 3
    for ind, val, neff in deff_values:
        ws.cell(row=row, column=1, value=ind).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=fmt_decimal(val, 2)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=3, value=fmt_int_es(round(neff))).alignment = ALIGN_RIGHT
        row += 1
    add_footer_notes(ws, row + 1, [
        "DEFF = Var(estimador bajo diseño complejo) / Var(estimador bajo muestreo aleatorio simple del mismo n).",
        "DEFF > 1 refleja la pérdida de precisión por conglomeración y ponderación (esperado en ENDES y similares).",
        "n efectivo SRS = n nominal / DEFF; aproxima el tamaño 'efectivo' bajo SRS equivalente.",
    ], 3)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22

    # Hoja Diagnósticos
    ws = wb.create_sheet("Diagnosticos_modelos")
    add_title_row(ws, 1, "Diagnósticos de los modelos ajustados (pooled sobre 20 imputaciones)", 5)
    style_header_row(ws, 2, ["Modelo", "Familia / link", "gl residuales (medio)", "Deviance escalada (mediana)", "Imputaciones"])
    family_map = {"quasipoisson": "Cuasi-Poisson / log", "quasibinomial": "Cuasi-binomial / logit"}
    row = 3
    for _, r in diag_df.iterrows():
        model = MODEL_LABELS.get(r["model"], r["model"])
        fam = family_map.get(r.get("family", ""), r.get("family", ""))
        ws.cell(row=row, column=1, value=model).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=fam).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3, value=fmt_decimal(r.get("median_df_resid", r.get("mean_df_resid")), 0)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=4, value=fmt_decimal(r.get("median_scale_deviance"), 2)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=5, value=int(r.get("imputations", 20))).alignment = ALIGN_CENTER
        row += 1
    add_footer_notes(ws, row + 1, [
        "Familia cuasi-Poisson con link log → razón de prevalencia (PR).",
        "Familia cuasi-binomial con link logit → odds ratio (OR), usado solo como sensibilidad.",
        "Deviance escalada (mediana sobre 20 imputaciones) cerca de 1 indica buen ajuste.",
    ], 5)
    ws.column_dimensions["A"].width = 50
    for c in range(2, 6):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # Hoja Bivariados detalle (compacto: solo mediana, min, max)
    ws = wb.create_sheet("Bivariados_detalle")
    add_title_row(ws, 1,
        "Detalle de pruebas Rao-Scott bivariadas sobre 20 imputaciones MICE.", 6,
    )
    style_header_row(ws, 2, ["Característica", "Estadístico F (medio)", "p mediana", "p mínima", "p máxima", "Imputaciones"])
    rs_summary = pd.read_csv(ANALYSIS / "tables" / "rao_scott_summary.csv")
    rs_summary["Variable"] = rs_summary["variable"].map(VAR_DISPLAY_NAMES).fillna(rs_summary["variable"])
    order = ["SEVERIDAD_DEPRESIVA", "QSSEXO", "QS25N", "HV025", "HV270",
             "VIOLENCIA_PAREJA", "ALCOHOL_PROBLEMATICO", "QS201",
             "QS109", "CALIDAD_DIETA", "ALTITUD_CAT3"]
    row = 3
    for var in order:
        sub = rs_summary[rs_summary["variable"] == var]
        if sub.empty:
            continue
        r = sub.iloc[0]
        ws.cell(row=row, column=1, value=VAR_DISPLAY_NAMES.get(var, var)).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2, value=fmt_decimal(r["mean_statistic"], 2)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=3, value=fmt_p(r["median_p_value"])).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4, value=fmt_p(r["min_p_value"])).alignment = ALIGN_CENTER
        ws.cell(row=row, column=5, value=fmt_p(r["max_p_value"])).alignment = ALIGN_CENTER
        ws.cell(row=row, column=6, value=int(r["imputations"])).alignment = ALIGN_CENTER
        row += 1
    add_footer_notes(ws, row + 1, [
        "Cada fila agrega 20 pruebas Rao-Scott (una por imputación MICE). p mediana es la mostrada en Tabla 2.",
        "min ≈ mediana ≈ máx indica baja variabilidad entre imputaciones: típicamente variables NO imputadas (la prueba se ejecuta sobre datos idénticos) o MICE targets con muy baja proporción de faltantes.",
        "Las variaciones reales de p entre imputaciones ocurren en MICE targets con missingness no trivial (ej. QS25N, CALIDAD_DIETA).",
        "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE).",
    ], 6)
    ws.column_dimensions["A"].width = 42
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 16

    wb.save(out_dir / "Tabla_S2_soporte_tablas.xlsx")
    print(f"  -> Tabla_S2_soporte_tablas.xlsx")


# ---------------------------------------------------------------------------
# Tabla S3 — Sensibilidad e interacciones + Sensibilidad logística
# ---------------------------------------------------------------------------

def build_tabla_S3(out_dir: Path) -> None:
    df_int = pd.read_csv(ANALYSIS / "models" / "interactions_and_sensitivity_models.csv")
    df_log = pd.read_csv(ANALYSIS / "models" / "logistic_sensitivity_results.csv")

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla S3 — Sensibilidades, interacciones y modelo logístico",
        tabla_destino="Tabla_S3_sensibilidad_interacciones.xlsx (Suplementario)",
        fuente="interactions_and_sensitivity_models.csv + logistic_sensitivity_results.csv (pooled 20 imp)",
        notas=[
            "Sensibilidad sin 2020: descarta sesgo por sub-muestreo COVID.",
            "Sensibilidad segunda toma: definición alternativa del desenlace (solo 2.ª medición de PA).",
            "Interacciones por sexo y año: prueban modificación de efecto del PHQ-9.",
            "Sensibilidad logística: misma fórmula del Modelo 2 pero link logit → OR; descarta sesgo de especificación.",
        ],
    )

    # Hoja Sensibilidad + Interacciones
    ws = wb.create_sheet("Sensibilidad_interacciones")
    add_title_row(ws, 1,
        "Tabla S3a. Modelos de sensibilidad e interacción (PR). ENDES 2019-2024.", 4,
    )
    style_header_row(ws, 2, ["Término", "PR", "IC 95%", "p"])
    row = 3
    n_subset_map = {
        "sensitivity_no_2020": "144 456",
        "sensitivity_second_bp_measure": "164 719",
        "interaction_sex": "164 719",
        "interaction_year": "164 719",
    }
    for model_name in ["sensitivity_no_2020", "sensitivity_second_bp_measure",
                       "interaction_sex", "interaction_year", "interaction_area",
                       "interaction_riqueza", "interaction_dxhta", "interaction_altitud"]:
        df_m = df_int[df_int["model"] == model_name]
        row = _write_model_block(ws, row, df_m, model_name, ncols=4)

    add_footer_notes(ws, row + 1, [
        "PR: cuasi-Poisson log link, pooled sobre 20 imputaciones MICE (Rubin). n: sin 2020 = 144 456; resto = 164 719.",
        "Interacción PHQ-9 × Mujer significativa (p = 0,001): asociación inversa más pronunciada en hombres.",
        "Interacción PHQ-9 × Año: sin evidencia consistente de modificación temporal.",
        "Imputación: 10 covariables MAR; PHQ-9 y PA elevada NO. Ver Métodos del manuscrito.",
    ], 4)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 11
    ws.freeze_panes = "A3"

    # Hoja Sensibilidad logística
    ws = wb.create_sheet("Sensibilidad_logistica")
    add_title_row(ws, 1,
        "Tabla S3b. Sensibilidad logística (OR). Misma fórmula del Modelo 2 con link logit.", 4,
    )
    style_header_row(ws, 2, ["Término", "OR", "IC 95%", "p"])
    row = 3
    row = _write_model_block(ws, row, df_log, "model_2_logistic_sensitivity", ncols=4)

    add_footer_notes(ws, row + 1, [
        "OR: odds ratio (cuasi-binomial, logit). Pooled sobre 20 imputaciones (Rubin). n = 164 719.",
        "Sensibilidad para descartar que el resultado del Modelo 2 dependa de la elección cuasi-Poisson.",
        "OR ≈ PR cuando la prevalencia es baja; aquí ~12 % → OR y PR cercanos pero no idénticos. Ver Métodos.",
    ], 4)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 11
    ws.freeze_panes = "A3"

    wb.save(out_dir / "Tabla_S3_sensibilidad_interacciones.xlsx")
    print(f"  -> Tabla_S3_sensibilidad_interacciones.xlsx")


# ---------------------------------------------------------------------------
# Tabla S4 — Datos fuente de figuras
# ---------------------------------------------------------------------------

def build_tabla_S4(out_dir: Path) -> None:
    strobe = pd.read_csv(ROOT / "data" / "output" / "Para Enviar" / "07_figura_1_strobe_datos.csv")
    spline = pd.read_csv(ANALYSIS / "figures" / "spline_curve_pooled.csv")
    spline_nl = pd.read_csv(ANALYSIS / "figures" / "spline_nonlinearity_summary.csv")

    wb = Workbook(); wb.remove(wb.active)
    add_portada(
        wb,
        titulo="Tabla S4 — Datos fuente de figuras",
        tabla_destino="Tabla_S4_datos_figuras.xlsx (Suplementario)",
        fuente="07_figura_1_strobe_datos.csv + spline_curve_pooled.csv + spline_nonlinearity_summary.csv",
        notas=[
            "Datos numéricos exactos que sostienen las Figuras 1 (STROBE) y S1 (Spline).",
            "Permite reproducir las figuras desde cero o auditar valor por valor.",
            "Coma decimal RPMESP.",
        ],
    )

    # Hoja Figura 1 STROBE
    ws = wb.create_sheet("Figura_1_STROBE")
    add_title_row(ws, 1, "Datos de Figura 1 — Flujo STROBE", 6)
    style_header_row(ws, 2, ["Paso", "Etapa", "n antes", "n excluidos", "n después", "% retenido"])
    row = 3
    for _, r in strobe.iterrows():
        ws.cell(row=row, column=1, value=int(r["order"])).alignment = ALIGN_CENTER
        ws.cell(row=row, column=2, value=str(r["stage"])).alignment = ALIGN_LEFT
        ws.cell(row=row, column=3, value=fmt_int_es(r["n_before"])).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=4, value=fmt_int_es(r["n_excluded"])).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=5, value=fmt_int_es(r["n_after"])).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=6, value=fmt_decimal(r["pct_remaining_from_start"], 1)).alignment = ALIGN_RIGHT
        row += 1
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    for c in range(3, 7):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Hoja Figura S1 Spline (curva)
    ws = wb.create_sheet("Figura_S1_curva")
    add_title_row(ws, 1, "Datos de Figura S1 — Curva spline PHQ-9 (pooled 20 imputaciones)", 6)
    style_header_row(ws, 2, ["PHQ-9", "Estimación (log-PR)", "EE", "IC 95% inferior", "IC 95% superior", "Prevalencia ajustada (%)"])
    spline_sorted = spline.sort_values("label").reset_index(drop=True)
    row = 3
    for _, r in spline_sorted.iterrows():
        ws.cell(row=row, column=1, value=int(float(r["label"]))).alignment = ALIGN_CENTER
        ws.cell(row=row, column=2, value=fmt_decimal(r["estimate"], 4)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=3, value=fmt_decimal(r["std_error"], 4)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=4, value=fmt_decimal(r["ci_low"], 4)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=5, value=fmt_decimal(r["ci_high"], 4)).alignment = ALIGN_RIGHT
        ws.cell(row=row, column=6, value=fmt_decimal(np.exp(r["estimate"]) * 100, 2)).alignment = ALIGN_RIGHT
        row += 1
    ws.column_dimensions["A"].width = 8
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # Hoja Figura S1 Spline (no-linealidad pooled)
    ws = wb.create_sheet("Figura_S1_nolinealidad")
    add_title_row(ws, 1, "Datos de Figura S1 — Resumen de no-linealidad pooled", 4)
    style_header_row(ws, 2, ["Estadístico Wald medio", "p mediana", "p mínima", "p máxima"])
    r = spline_nl.iloc[0]
    ws.cell(row=3, column=1, value=fmt_decimal(r["mean_wald_statistic"], 2)).alignment = ALIGN_RIGHT
    ws.cell(row=3, column=2, value=fmt_p(r["median_p_value"])).alignment = ALIGN_CENTER
    ws.cell(row=3, column=3, value=fmt_p(r["min_p_value"])).alignment = ALIGN_CENTER
    ws.cell(row=3, column=4, value=fmt_p(r["max_p_value"])).alignment = ALIGN_CENTER
    add_footer_notes(ws, 5, [
        "p > 0,05 indica que no hay evidencia de no-linealidad → el coeficiente lineal de PHQ-9 en Modelo 2 es interpretable.",
        "Wald con 2 gl (3 nodos internos del spline cúbico restringido: 0, 4, 9, 14).",
    ], 4)
    for c in range(1, 5):
        ws.column_dimensions[get_column_letter(c)].width = 20

    wb.save(out_dir / "Tabla_S4_datos_figuras.xlsx")
    print(f"  -> Tabla_S4_datos_figuras.xlsx")


# ---------------------------------------------------------------------------
# Figuras y leyendas
# ---------------------------------------------------------------------------

def copy_figures(out_principal: Path, out_suplementario: Path) -> None:
    # STROBE v2 → Figura 1 Principal
    for ext in ("svg", "png", "pdf"):
        src = AUDIT / f"Figura_1_STROBE_v2.{ext}"
        if src.exists():
            shutil.copy(src, out_principal / f"Figura_1_STROBE.{ext}")
            print(f"  -> Principal/Figura_1_STROBE.{ext}")
    # Spline v2 → Figura S1 Suplementario
    for ext in ("svg", "png", "pdf"):
        src = AUDIT / f"Figura_S1_Spline_v2.{ext}"
        if src.exists():
            shutil.copy(src, out_suplementario / f"Figura_S1_Spline.{ext}")
            print(f"  -> Suplementario/Figura_S1_Spline.{ext}")
    # DAG → Figura S2 Suplementario (versión rediseñada con matplotlib;
    # 02_DAG_modelos.R sigue siendo la especificación formal)
    for ext in ("svg", "png", "pdf"):
        src = AUDIT / f"Figura_S2_DAG.{ext}"
        if src.exists():
            shutil.copy(src, out_suplementario / f"Figura_S2_DAG.{ext}")
            print(f"  -> Suplementario/Figura_S2_DAG.{ext}")


def write_leyendas(out_principal: Path, out_suplementario: Path) -> None:
    principal_text = (
        "Leyendas de figuras principales (RPMESP)\n"
        "\n"
        "Figura 1. Diagrama de flujo STROBE para la selección muestral del análisis multianual ENDES 2019-2024. "
        "STROBE: Strengthening the Reporting of Observational Studies in Epidemiology; ENDES: Encuesta Demográfica "
        "y de Salud Familiar; MEF: mujeres en edad fértil (15-49 años); PHQ-9: Patient Health Questionnaire-9; "
        "PA: presión arterial. La exclusión reproductiva (gestación V454=1 y puerperio temprano V222<2) se aplicó "
        "únicamente a mujeres en edad fértil. En 2020, las restricciones operativas COVID-19 ocasionaron 8 828 "
        "exclusiones adicionales por PA no medible; el análisis de sensibilidad sin 2020 (Tabla 5) incluye n = 144 456.\n"
    )
    (out_principal / "Leyendas_figuras_RPMESP.txt").write_text(principal_text, encoding="utf-8")
    print("  -> Principal/Leyendas_figuras_RPMESP.txt")

    supl_text = (
        "Leyendas de figuras suplementarias\n"
        "\n"
        "Figura S1. Curva spline cúbica restringida de la prevalencia ajustada de presión arterial elevada según el "
        "puntaje PHQ-9, pooled sobre 20 imputaciones MICE. Bandas: intervalo de confianza al 95% (Rubin). Nodos en "
        "PHQ-9 = 0, 4, 9, 14. La región sombreada en gris (PHQ-9 > 20) contiene < 1,2 % de la muestra; el "
        "histograma marginal inferior muestra la densidad de PHQ-9. La p mediana de no-linealidad pooled fue 0,251 "
        "(no se rechaza linealidad). La figura se conserva como verificación de la forma funcional del Modelo 2 "
        "estructural, no como hallazgo principal. PHQ-9: Patient Health Questionnaire-9; IC 95%: intervalo de "
        "confianza del 95%.\n"
        "\n"
        "Figura S2. Grafo dirigido acíclico (DAG) que ilustra la estructura causal del estudio, con flujo izquierda → "
        "derecha en el eje principal exposición → mediadores → desenlace. La flecha negra gruesa entre PHQ-9 y PA "
        "elevada representa el efecto total a estimar. Los confusores estructurales (verde) están incluidos en el "
        "Modelo 2 (principal): edad, sexo, educación, área de residencia, quintil de riqueza, violencia de pareja y "
        "año del estudio; los arcos dashed bidireccionales entre EDUC ↔ RIQUEZA y EDAD ↔ AÑO indican correlaciones "
        "esperadas entre confusores. Los mediadores potenciales (naranja) están en caminos causales dirigidos PHQ-9 "
        "→ PA elevada y se añaden al Modelo 3 (exploratorio): IMC y circunferencia abdominal (vía metabólica), "
        "consumo de tabaco (últimos 30 días) y consumo problemático de alcohol (vía conductual), calidad de la dieta "
        "y diagnóstico de diabetes (vía metabólica). Se grafican aristas confusor → mediador fundamentadas en la "
        "literatura (p. ej., edad → IMC; sexo → circunferencia; riqueza/educación → dieta; edad → diabetes). El nodo U "
        "(violeta dashed) representa factores no medidos (estrés crónico, sueño, dislipidemia / ERC / enfermedad "
        "coronaria, adherencia previa, sesgo de "
        "detección) que pueden actuar como confusores residuales sobre PHQ-9 y PA elevada. Las variables de diseño "
        "muestral (HV001, HV022, PESO_FINAL) no se grafican porque entran al svydesign() y no al ajuste causal. "
        "Para mantener la legibilidad, se omiten en la figura algunas aristas mediador → mediador "
        "(tabaco ↔ alcohol problemático; DIETA → IMC; IMC → CINTURA) y confusor → mediador secundarias; la especificación "
        "formal completa está en el archivo 02_DAG_modelos.R. El DAG fue construido con dagitty (R) y verificado "
        "acíclico (isAcyclic = TRUE); el conjunto mínimo de ajuste para el efecto total coincide con las "
        "covariables del Modelo 2.\n"
    )
    (out_suplementario / "Leyendas_figuras_suplementarias.txt").write_text(supl_text, encoding="utf-8")
    print("  -> Suplementario/Leyendas_figuras_suplementarias.txt")


def write_readme(out_dir: Path) -> None:
    readme = (
        "Post_Auditoria/\n"
        "================\n"
        "\n"
        "Paquete editorial unificado en lenguaje visual v3.1 (Tabla 1 v3.1 → resto de tablas).\n"
        "Generado el 2026-05-29 desde scripts/audit/build_post_auditoria.py.\n"
        "\n"
        "NO reemplaza Para Enviar/Para Publicar Redisenado/. Coexiste como propuesta de paquete\n"
        "final para presentar a la revista.\n"
        "\n"
        "Estructura\n"
        "----------\n"
        "  Principal/\n"
        "    Figura_1_STROBE.{svg,png,pdf}   Diagrama STROBE (replica Figura_1_STROBE_v2 de la auditoría).\n"
        "    Tabla_1.xlsx                    16 variables, Sexo × PAE, p Rao-Scott pooled, DEFF en pie.\n"
        "    Tabla_2.xlsx                    Asociaciones bivariadas Rao-Scott (pooled).\n"
        "    Tabla_3.xlsx                    Modelos 1 (crudo), 2 (principal estructural), 3 (exploratorio).\n"
        "    Tabla_4.xlsx                    Submodelos de cascada en personas con DX previo de HTA.\n"
        "    Tabla_5.xlsx                    Resumen de sensibilidades y robustez.\n"
        "    Leyendas_figuras_RPMESP.txt\n"
        "\n"
        "  Suplementario/\n"
        "    Figura_S1_Spline.{svg,png,pdf}  Spline PHQ-9 → PA (pooled, IC 95% y región sparse).\n"
        "    Figura_S2_DAG.{svg,png}         DAG conceptual del estudio (dagitty).\n"
        "    Tabla_S1_flujo_STROBE.xlsx      Detalle numérico del flujo STROBE.\n"
        "    Tabla_S2_soporte_tablas.xlsx    DEFF, diagnósticos de modelos, bivariados por imputación.\n"
        "    Tabla_S3_sensibilidad_interacciones.xlsx  Sensibilidad + interacciones + logística.\n"
        "    Tabla_S4_datos_figuras.xlsx     Datos fuente exactos de Figura 1 y S1.\n"
        "    Leyendas_figuras_suplementarias.txt\n"
        "\n"
        "Lenguaje visual común\n"
        "---------------------\n"
        "  - Encabezados azul claro (D9E1F2), secciones gris (F2F2F2).\n"
        "  - PHQ-9 destacado en azul medio (DCE6F1) por ser exposición principal.\n"
        "  - Continuas: media (DE). Categóricas: n no ponderado (% ponderado).\n"
        "  - PR/OR con 2 decimales y coma; IC 95% como 'X,XX a X,XX'; p con 3 decimales o '< 0,001'.\n"
        "  - Etiquetas legibles: sin factor(), .tmp_pred*, ni códigos ENDES crudos.\n"
        "  - Pooled sobre 20 imputaciones MICE donde aplique (Rubin / mediana de p).\n"
        "\n"
        "Sobre el pooling sobre 20 imputaciones (importante)\n"
        "---------------------------------------------------\n"
        "Cada análisis (modelo, prueba bivariada, spline) se ejecutó en los 20 datasets\n"
        "imputados MICE y se combinó por reglas de Rubin (coeficientes) o mediana del p\n"
        "(pruebas Rao-Scott). 'Pooled sobre 20 imputaciones' se refiere a este procedimiento,\n"
        "NO a que cada variable haya sido imputada.\n"
        "\n"
        "  Variables imputadas (MAR, 10 covariables listadas en mice_manifest.json):\n"
        "    edad, sexo, educación, área de residencia, riqueza, violencia de pareja,\n"
        "    IMC, circunferencia abdominal, calidad de dieta, diagnóstico de diabetes (QS109).\n"
        "\n"
        "  Variables NO imputadas:\n"
        "    - Desenlace: PA elevada (los casos sin medición válida se excluyen en STROBE).\n"
        "    - Exposición: PHQ-9 (se prorratea cuando faltan 1-2 ítems; ≥3 ítems faltantes excluye).\n"
        "    - Saltos estructurales del cuestionario: QS201, ALCOHOL_PROBLEMATICO\n"
        "      (NMAR por diseño; mantienen 'No aplica' como categoría explícita).\n"
        "    - TIEMPO_DX_HTA_MESES (excluida formalmente del estudio por ausencia de datos crudos).\n"
        "\n"
        "Regeneración\n"
        "------------\n"
        "  .\\.venv\\Scripts\\python.exe scripts\\audit\\build_post_auditoria.py\n"
    )
    (out_dir / "LEEME.txt").write_text(readme, encoding="utf-8")
    print("  -> LEEME.txt")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Post_Auditoria — paquete editorial v3.1:")
    out_principal = OUT / "Principal"
    out_suplementario = OUT / "Suplementario"
    out_principal.mkdir(parents=True, exist_ok=True)
    out_suplementario.mkdir(parents=True, exist_ok=True)

    print("\n[Principal]")
    build_tabla_1(out_principal)
    build_tabla_2(out_principal)
    build_tabla_3(out_principal)
    build_tabla_3_wide(out_principal)
    build_tabla_4(out_principal)
    build_tabla_4_wide(out_principal)
    build_tabla_5(out_principal)

    print("\n[Suplementario]")
    build_tabla_S1(out_suplementario)
    build_tabla_S2(out_suplementario)
    build_tabla_S3(out_suplementario)
    build_tabla_S3_wide(out_suplementario)
    build_tabla_S4(out_suplementario)
    # Tabla S5 (justificación bibliográfica del DAG) se genera con su propio
    # script para mantener separable la documentación causal:
    #   python scripts/audit/build_tabla_S5_dag_justificacion.py

    print("\n[Figuras]")
    copy_figures(out_principal, out_suplementario)

    print("\n[Leyendas y README]")
    write_leyendas(out_principal, out_suplementario)
    write_readme(OUT)

    print(f"\nListo. Paquete en {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
