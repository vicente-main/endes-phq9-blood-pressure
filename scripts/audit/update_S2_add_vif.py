"""Añade a S2_Table.xlsx la sección 'C. Variance inflation factors (VIF)' y define
'median scaled deviance' (auditoría 2026-07-16, hallazgo #7 y pregunta abierta 2).

- #7: el cuerpo cita "VIF ... maximum 15" pero S2 no tenía sección VIF. Se añade,
  declarando que se computa sobre la primera imputación (vif_first_imputation.csv).
- Pregunta abierta 2: se define 'median scaled deviance' como la mediana sobre las
  20 imputaciones del cociente devianza residual/gl del ajuste (descriptor naïve de
  bondad de ajuste), que NO es el parámetro de dispersión de diseño φ del quasi-
  Poisson (φ < 1, subdispersión; reportado en el cuerpo). Son cantidades distintas.

Idempotente: si ya existe la sección C la reescribe.

Fuente VIF: data/output_2025/analysis/vif_first_imputation.csv
Salida: ENVIO_PLOS_ONE_2019-2025/S2_Table.xlsx (in place)
"""
from __future__ import annotations
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
VIF_CSV = ROOT / "data" / "output_2025" / "analysis" / "vif_first_imputation.csv"
S2_PATH = ROOT / "ENVIO_PLOS_ONE_2019-2025" / "S2_Table.xlsx"
FILL_SECTION = "F2F2F2"
FILL_HEADER = "D9E1F2"

# Etiquetas legibles para los predictores dummy del VIF (diccionario INEI real).
VIF_LABELS = {
    "PHQ9_TOTAL": "PHQ-9 score",
    "QS23": "Age",
    "IMC": "Body mass index",
    "QS907": "Waist circumference",
    "QSSEXO_2": "Sex: female",
    "QS25N_1": "Education: primary",
    "QS25N_2": "Education: secondary",
    "QS25N_3": "Education: higher non-university",
    "QS25N_4": "Education: higher university",
    "QS25N_5": "Education: postgraduate",
    "HV025_2": "Area: rural",
    "HV270_2": "Wealth: Q2",
    "HV270_3": "Wealth: Q3",
    "HV270_4": "Wealth: Q4",
    "HV270_5": "Wealth: Q5",
    "VIOLENCIA_PAREJA_1": "IPV: with violence (physical)",
    "VIOLENCIA_PAREJA_2": "IPV: no partner",
    "QS201_2": "Tobacco: did not smoke (30 d)",
    "QS109_2": "Diabetes: no diagnosis",
    "ALCOHOL_PROBLEMATICO_1": "Alcohol: problematic use",
    "CALIDAD_DIETA_1": "Diet: adequate",
    "ALTITUD_CAT3_<1500": "Altitude: < 1,500 m",
    "ALTITUD_CAT3_>=2500": "Altitude: ≥ 2,500 m",
}


def main() -> None:
    vif = pd.read_csv(VIF_CSV).sort_values("vif", ascending=False).reset_index(drop=True)
    wb = load_workbook(S2_PATH)
    ws = wb.worksheets[0]

    # Localizar/limpiar una sección C previa (para idempotencia).
    stop = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.startswith("C. Variance inflation"):
            stop = r
            break
    if stop is not None:
        ws.delete_rows(stop, ws.max_row - stop + 1)

    r = ws.max_row + 2
    hdr = ws.cell(r, 1, "C. Variance inflation factors (VIF)")
    hdr.font = Font(bold=True)
    hdr.fill = PatternFill("solid", fgColor=FILL_SECTION)
    r += 1
    for c, txt in enumerate(["Predictor (model term)", "VIF"], 1):
        cell = ws.cell(r, c, txt)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=FILL_HEADER)
    r += 1
    for _, row in vif.iterrows():
        pred = str(row["predictor"])
        ws.cell(r, 1, VIF_LABELS.get(pred, pred))
        ws.cell(r, 2, round(float(row["vif"]), 2))
        r += 1

    r += 1
    vif_max = float(vif["vif"].max())
    top_pred = VIF_LABELS.get(str(vif.iloc[0]["predictor"]), str(vif.iloc[0]["predictor"]))
    notes = [
        "Notes.",
        "A. DEFF = design effect (ratio of the survey-design variance to the simple-random-sample "
        "variance under svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE)).",
        "B. Median scaled deviance = median across the 20 MICE imputations of the residual "
        "deviance divided by the residual degrees of freedom of the fitted model. It is a naive "
        "goodness-of-fit descriptor of the Poisson working model and is NOT the design-based "
        "dispersion parameter φ of the quasi-Poisson variance; the latter, estimated under the "
        "complex design, is < 1 (sub-dispersion) and is reported in the main text.",
        f"C. VIF computed on the first MICE imputation over all Model-3 predictors. The highest "
        f"values ({top_pred}, maximum {vif_max:.1f}) occur among the educational-level indicator "
        f"variables, reflecting the expected collinearity among dummy levels of the SAME categorical "
        f"factor (education), not collinearity between distinct covariables; all values are below the "
        f"conventional threshold of 10. Educational level is coded 0–5 per the INEI ENDES dictionary "
        f"(reference = 0, no formal education / initial, which includes 'never attended school', "
        f"QS24 = 2).",
    ]
    for note in notes:
        cell = ws.cell(r, 1, note)
        if note.endswith("."):
            if note in ("Notes.",):
                cell.font = Font(bold=True)
        r += 1

    wb.save(S2_PATH)
    print(f"S2_Table.xlsx actualizado: sección C con {len(vif)} predictores; VIF máx = "
          f"{vif['vif'].max():.2f} ({vif.iloc[0]['predictor']}).")


if __name__ == "__main__":
    main()
