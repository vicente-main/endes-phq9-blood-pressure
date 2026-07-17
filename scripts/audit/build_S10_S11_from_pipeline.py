"""S10/S11: modelos completos ajustados (Tabla 3 y Tabla 4) generados DESDE EL
PIPELINE (autoritativo), no copiados del .docx.

Responde al pendiente §5.3 del traspaso de Rodrigo (2026-07-12): regenerar
S10_Table.xlsx (Modelos 1/2/3 completos) y S11_Table.xlsx (cascada: no adherencia
y PA no controlada) con el mismo formato y etiquetas que las versiones del .docx,
tomando los valores de:
  - data/output_2025/analysis/models/table3_main_models.csv   -> S10
  - data/output_2025/analysis/models/table4_cascade_models.csv -> S11

Ademas VALIDA celda-a-celda contra las versiones de Rodrigo (S10_Table.xlsx /
S11_Table.xlsx en la raiz del proyecto) e imprime cualquier discrepancia.

Formato: PR/CI a 2 decimales, salvo la fila de exposicion PHQ-9 a 3 decimales;
p a 3 decimales o "< 0.001". Intercepto omitido.

Salida: ENVIO_PLOS_ONE_2019-2025/S10_Table.xlsx y S11_Table.xlsx
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
MODELS = ROOT / "data" / "output_2025" / "analysis" / "models"
OUT_DIR = ROOT / "ENVIO_PLOS_ONE_2019-2025"
RODRIGO_DIR = ROOT  # sus S10/S11 estan en la raiz

FILL_HEADER = "D9E1F2"
FILL_SECTION = "F2F2F2"

YEARS = [("Year %d (vs 2019)" % y, "factor(ANIO)%d" % y) for y in range(2020, 2026)]
WEALTH = [("Q2 — Poor (vs Q1 poorest)", "factor(HV270)2"),
          ("Q3 — Middle (vs Q1)", "factor(HV270)3"),
          ("Q4 — Rich (vs Q1)", "factor(HV270)4"),
          ("Q5 — Richest (vs Q1)", "factor(HV270)5")]
# Etiquetas corregidas (auditoria 2026-07-16): diccionario INEI real 0..5 —
# referencia = 0 (no formal/initial); 5 = Postgraduate (antes omitido y todo
# el bloque desplazado una posicion). Fuente unica: etiquetas_educacion.py.
EDU = [("Primary", "factor(QS25N)1"), ("Secondary", "factor(QS25N)2"),
       ("Higher non-university", "factor(QS25N)3"), ("Higher university", "factor(QS25N)4"),
       ("Postgraduate", "factor(QS25N)5")]
ALT = [("Altitude 1,500-2,499 m (vs < 1,500)", "factor(ALTITUD_CAT3)1500-2499"),
       ("Altitude ≥ 2,500 m (vs < 1,500)", "factor(ALTITUD_CAT3)>=2500")]
# IPV: QS710/QS711 son actos FISICOS (auditoria 2026-07-16, hallazgo #6) — el
# constructo no incluye violencia psicologica ni sexual; coherente con Metodos.
SEX_VP = [("Female (vs Male)", "factor(QSSEXO)2"),
          ("With violence (physical)", "factor(VIOLENCIA_PAREJA)1"),
          ("No partner", "factor(VIOLENCIA_PAREJA)2")]
PHQ = ("PHQ-9 score (per additional point)", "PHQ9_TOTAL")
AGE = ("Age (per additional year)", "QS23")

STRUCT = [PHQ, AGE, *ALT, *YEARS, ("Rural (vs Urban)", "factor(HV025)2"), *WEALTH, *EDU, *SEX_VP]

M3_ORDER = [
    ("BMI (per additional kg/m²)", "IMC"), PHQ, AGE,
    ("Waist circumference (per additional cm)", "QS907"),
    ("Problematic use", "factor(ALCOHOL_PROBLEMATICO)1"),
    ("Not applicable (does not drink)", "factor(ALCOHOL_PROBLEMATICO)NaN"),
    *ALT, *YEARS,
    ("Adequate diet", "factor(CALIDAD_DIETA)1"),
    ("Rural (vs Urban)", "factor(HV025)2"), *WEALTH,
    ("No diabetes diagnosis", "factor(QS109)2"),
    ("Did not smoke in last 30 days", "factor(QS201)2"),
    ("Not applicable (did not smoke in last 12 months; skip)", "factor(QS201)NaN"),
    *EDU, *SEX_VP,
]

S10_SECTIONS = [
    ("Model 1 — Crude", "model_1", [PHQ]),
    ("Model 2 — Structural (main)", "model_2", STRUCT),
    ("Model 3 — Exploratory (with mediators)", "model_3", M3_ORDER),
]
S11_SECTIONS = [
    ("Therapeutic non-adherence (prior HTN dx, n = 17,454)", "submodel_adherence", STRUCT),
    ("Uncontrolled blood pressure (prior HTN dx, n = 17,469)", "submodel_domain_bp", STRUCT),
]

S10_TITLE = ("S10 Table. Full adjusted models (Model 1 crude, Model 2 structural, Model 3 "
             "exploratory) for elevated blood pressure by PHQ-9 and all covariables. "
             "ENDES 2019-2025 (n = 191,757; 20 MICE imputations).")
S10_NOTES = [
    "Notes:",
    "PR: prevalence ratio (quasi-Poisson, log link). 95% CI and p combined by Rubin's rules over "
    "20 MICE imputations. n = 191,757 (constant across the three models).",
    "Imputation: the MAR covariables were imputed; the PHQ-9 exposure, the elevated-BP outcome, "
    "altitude, and educational level were NOT imputed (education missingness is the structural "
    "QS24 questionnaire skip, resolved deterministically). Details in the manuscript Methods.",
    "Education reference category: no formal education / initial (QS25N = 0; includes 'never "
    "attended school', QS24 = 2).",
    "Model 1: crude PHQ-9. Model 2 (main): + age, sex, education, area, wealth, IPV, year, and altitude. "
    "Model 3 (exploratory): + BMI, waist, tobacco use, problematic alcohol use, diet, and diabetes "
    "diagnosis (potential mediators).",
    "Altitude incorporated into the structural adjustment as a geographic confounder. Intercept omitted.",
    "Exposure row (PHQ-9) shown to 3 decimals; remaining coefficients to 2 decimals.",
    "Source: table3_main_models.csv (pipeline, 2019-2025 run).",
]
S11_TITLE = ("S11 Table. Full adjusted care-cascade models (non-adherence and uncontrolled blood "
             "pressure) by PHQ-9 and all covariables in persons with a prior hypertension diagnosis. "
             "ENDES 2019-2025.")
S11_NOTES = [
    "Notes:",
    "PR: prevalence ratio (quasi-Poisson, log link). Pooled over 20 MICE imputations (Rubin).",
    "Non-adherence: outcome = NO_ADHERENCIA_HTA in persons with prior Dx (n = 17,454). "
    "Uncontrolled: outcome = elevated BP in persons with prior Dx (n = 17,469).",
    "Same covariables as the structural Model 2. The MAR covariables were imputed; the outcome, "
    "altitude, and educational level were NOT (education missingness is the structural QS24 skip, "
    "resolved deterministically). See the manuscript Methods.",
    "Education reference category: no formal education / initial (QS25N = 0; includes 'never "
    "attended school', QS24 = 2).",
    "Exposure row (PHQ-9) shown to 3 decimals; remaining coefficients to 2 decimals. Intercept omitted.",
    "Source: table4_cascade_models.csv (pipeline, 2019-2025 run).",
]


def _fp(p):
    return "< 0.001" if p < 0.001 else f"{p:.3f}"


def _fmt(row, dec):
    pr = f"{row['exp_estimate']:.{dec}f}"
    ci = f"{row['exp_ci_low']:.{dec}f} to {row['exp_ci_high']:.{dec}f}"
    return pr, ci, _fp(float(row["p_value"]))


def _load(csv):
    d = pd.read_csv(csv)
    return {(r["model"], str(r["term"])): r for _, r in d.iterrows()}


def _build_rows(sections, idx):
    rows = []  # (kind, label, pr, ci, p)
    missing = []
    for header, model, order in sections:
        rows.append(("section", header, "", "", ""))
        for label, term in order:
            key = (model, term)
            if key not in idx:
                missing.append(key)
                continue
            dec = 3 if term == "PHQ9_TOTAL" else 2
            pr, ci, p = _fmt(idx[key], dec)
            rows.append(("data", label, pr, ci, p))
    return rows, missing


def _validate(rows, rodrigo_file):
    """Compara filas de datos (label,pr,ci,p) contra el archivo de Rodrigo, en orden."""
    if not rodrigo_file.exists():
        return [f"(no se encontro {rodrigo_file.name} para validar)"]
    d = pd.read_excel(rodrigo_file, header=None)
    ref = []
    for i in range(len(d)):
        lab = d.iloc[i, 0]
        if not isinstance(lab, str):
            continue
        pr = d.iloc[i, 1] if d.shape[1] > 1 else None
        if pd.isna(pr):  # encabezados de seccion / notas / titulo
            continue
        if lab.strip() == "Term" or str(pr).strip() == "PR":  # fila de encabezado
            continue
        ref.append((lab.strip(), str(d.iloc[i, 1]).strip(),
                    str(d.iloc[i, 2]).strip(), str(d.iloc[i, 3]).strip()))
    mine = [(l, pr, ci, p) for (k, l, pr, ci, p) in rows if k == "data"]
    diffs = []
    if len(mine) != len(ref):
        diffs.append(f"conteo de filas de datos: pipeline={len(mine)} vs Rodrigo={len(ref)}")
    for i, (m, r) in enumerate(zip(mine, ref)):
        # normalizar 'to' vs 'a', y ≥
        if m[1] != r[1] or m[2].replace(" to ", " to ") != r[2] or m[3] != r[3]:
            if not (m[1] == r[1] and m[2] == r[2] and m[3] == r[3]):
                diffs.append(f"[{i}] '{m[0]}': pipeline PR={m[1]} CI={m[2]} p={m[3]}  |  "
                             f"Rodrigo PR={r[1]} CI={r[2]} p={r[3]}")
    return diffs


def _write(path, title, rows, notes):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    wb = Workbook(); ws = wb.active; ws.title = path.stem.replace("_Table", " Table")
    hfill = PatternFill("solid", fgColor=FILL_HEADER)
    sfill = PatternFill("solid", fgColor=FILL_SECTION)
    bold = Font(bold=True); thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1, title).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    for c, h in enumerate(["Term", "PR", "95% CI", "p"], 1):
        cell = ws.cell(3, c, h); cell.fill = hfill; cell.font = bold; cell.border = border; cell.alignment = center
    r = 4
    for kind, label, pr, ci, p in rows:
        if kind == "section":
            cell = ws.cell(r, 1, label); cell.font = bold; cell.fill = sfill
            for c in range(1, 5):
                ws.cell(r, c).fill = sfill; ws.cell(r, c).border = border
        else:
            ws.cell(r, 1, label); ws.cell(r, 2, pr); ws.cell(r, 3, ci); ws.cell(r, 4, p)
            for c in range(1, 5):
                ws.cell(r, c).border = border
                ws.cell(r, c).alignment = left if c == 1 else center
        r += 1
    r += 1
    for note in notes:
        cell = ws.cell(r, 1, note); cell.alignment = left
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        if note == "Notes:":
            cell.font = bold
        r += 1
    for col, w in zip("ABCD", [46, 10, 16, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(path)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    idx3 = _load(MODELS / "table3_main_models.csv")
    idx4 = _load(MODELS / "table4_cascade_models.csv")

    s10, miss10 = _build_rows(S10_SECTIONS, idx3)
    s11, miss11 = _build_rows(S11_SECTIONS, idx4)
    for name, miss in [("S10", miss10), ("S11", miss11)]:
        if miss:
            print(f"  ⚠ {name}: terminos no encontrados en el pipeline: {miss}")

    print("=== Validacion S10 vs S10_Table.xlsx de Rodrigo ===")
    d10 = _validate(s10, RODRIGO_DIR / "S10_Table.xlsx")
    print("  " + ("\n  ".join(d10) if d10 else "SIN DISCREPANCIAS ✓"))
    print("=== Validacion S11 vs S11_Table.xlsx de Rodrigo ===")
    d11 = _validate(s11, RODRIGO_DIR / "S11_Table.xlsx")
    print("  " + ("\n  ".join(d11) if d11 else "SIN DISCREPANCIAS ✓"))

    _write(OUT_DIR / "S10_Table.xlsx", S10_TITLE, s10, S10_NOTES)
    _write(OUT_DIR / "S11_Table.xlsx", S11_TITLE, s11, S11_NOTES)
    print(f"\nEscritos:\n  {OUT_DIR / 'S10_Table.xlsx'}\n  {OUT_DIR / 'S11_Table.xlsx'}")


if __name__ == "__main__":
    main()
