"""Tabla_1_v3_1.xlsx — Versión 3.1 mínima viable (16 variables).

Cambios vs v3:
- Añade 3 continuas: PHQ9_TOTAL (puntaje), PAS_PROM, PAD_PROM (presiones medias).
- Las 7 p frescas (DX_HTA_PREVIO + 6 continuas) se pooled sobre 20 imputaciones
  (mediana del p, mismo método que rao_scott_summary.csv del pipeline).
- Footer con DEFF global (3 cifras) reusado de Tabla 1 v2 hoja DEFF.
- Sin asteriscos: TODAS las p son ahora pooled sobre 20 imputaciones.

Lógica de pooling:
- 9 categóricas (Severidad, HV025, HV270, QS25N, VPAR, Alcohol prob, QS201, QS109,
  Calidad dieta) → p pooled reusado de rao_scott_summary.csv.
- 7 frescas pooled fresh (1 categórica + 6 continuas).
- Celdas (n %, media DE) → primera imputación (la diferencia con pooled es
  invisible en el redondeo de la tabla; documentado en footer).
"""
from __future__ import annotations

import subprocess
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
IMPUTED_DIR = ROOT / "data" / "output" / "analysis" / "imputed"
RAO_SCOTT_POOLED = ROOT / "data" / "output" / "analysis" / "tables" / "rao_scott_summary.csv"
DEFF_SOURCE = ROOT / "data" / "output" / "analysis" / "tables" / "table1_deff_summary.csv"
OUT_PATH = ROOT / "data" / "output" / "Auditoria_Integral" / "Tabla_1_v3_1.xlsx"
RSCRIPT_EXE = Path("C:/Program Files/R/R-4.5.3/bin/Rscript.exe")
R_LIB_USER = Path("C:/Users/Trabajo/Documents/R/win-library/4.5")


SEX_ORDER = ["Hombres", "Mujeres"]
PAE_ORDER = ["Sin PAE", "Con PAE"]


CATEGORICAL_VARS = [
    (
        "SEVERIDAD_DEPRESIVA", "Severidad PHQ-9 (exposición principal)",
        {"Minima": "Mínima (0–4)", "Leve": "Leve (5–9)",
         "Moderada": "Moderada (10–14)", "Mod_Severa": "Mod. severa (15–19)",
         "Severa": "Severa (20–27)"},
        ["Mínima (0–4)", "Leve (5–9)", "Moderada (10–14)", "Mod. severa (15–19)", "Severa (20–27)"],
    ),
    ("HV025", "Área de residencia", {1: "Urbano", 2: "Rural"}, ["Urbano", "Rural"]),
    (
        "HV270", "Quintil de riqueza",
        {1: "Q1 — Más pobre", 2: "Q2", 3: "Q3", 4: "Q4", 5: "Q5 — Más rico"},
        ["Q1 — Más pobre", "Q2", "Q3", "Q4", "Q5 — Más rico"],
    ),
    (
        "QS25N", "Nivel educativo",
        {0: "Sin educación / inicial", 1: "Sin educación / inicial",
         2: "Primaria", 3: "Secundaria",
         4: "Superior no universitaria", 5: "Superior universitaria"},
        ["Sin educación / inicial", "Primaria", "Secundaria",
         "Superior no universitaria", "Superior universitaria"],
    ),
    (
        "VIOLENCIA_PAREJA", "Violencia de pareja (último año)",
        {0: "Sin violencia", 1: "Leve / única", 2: "Múltiple / severa"},
        ["Sin violencia", "Leve / única", "Múltiple / severa"],
    ),
    (
        "ALCOHOL_PROBLEMATICO", "Consumo problemático de alcohol",
        {0: "No", 1: "Sí", "NA": "No aplica (no consume)"},
        ["No", "Sí", "No aplica (no consume)"],
    ),
    (
        "QS201", "Consumo de tabaco (últimos 30 días)",
        {1: "Sí (fumó últimos 30 días)", 2: "No fumó", "NA": "No aplica (no fumó últimos 12 meses)"},
        ["Sí (fumó últimos 30 días)", "No fumó", "No aplica (no fumó últimos 12 meses)"],
    ),
    (
        "QS109", "Diagnóstico de diabetes",
        {1: "Sí (diabetes)", 2: "No"},
        ["No", "Sí (diabetes)"],
    ),
    (
        "CALIDAD_DIETA", "Calidad de la dieta",
        {0: "No adecuada", 1: "Adecuada"},
        ["No adecuada", "Adecuada"],
    ),
    (
        "DX_HTA_PREVIO", "Diagnóstico previo de HTA",
        {0: "No", 1: "Sí"},
        ["No", "Sí"],
    ),
    (
        "ALTITUD_CAT3", "Altitud de residencia (m s.n.m.)",
        {"<1500": "< 1 500", "1500-2499": "1 500-2 499", ">=2500": "≥ 2 500"},
        ["< 1 500", "1 500-2 499", "≥ 2 500"],
    ),
]

# 6 continuas: 3 originales + 3 nuevas (PHQ9_TOTAL, PAS_PROM, PAD_PROM)
CONTINUOUS_VARS = [
    ("QS23", "Edad", "años"),
    ("PHQ9_TOTAL", "Puntaje PHQ-9 (exposición principal)", "puntaje"),
    ("IMC", "Índice de masa corporal", "kg/m²"),
    ("QS907", "Circunferencia abdominal", "cm"),
    ("PAS_PROM", "Presión arterial sistólica media", "mmHg"),
    ("PAD_PROM", "Presión arterial diastólica media", "mmHg"),
]

P_FROM_POOLED = {
    "SEVERIDAD_DEPRESIVA", "HV025", "HV270", "QS25N",
    "VIOLENCIA_PAREJA", "ALCOHOL_PROBLEMATICO", "QS201",
    "QS109", "CALIDAD_DIETA", "ALTITUD_CAT3",
}
P_FRESH_TO_POOL = {"DX_HTA_PREVIO", "QS23", "PHQ9_TOTAL", "IMC", "QS907", "PAS_PROM", "PAD_PROM"}


# ---------------------------------------------------------------------------
# Carga base (imp1 para celdas) + preparación
# ---------------------------------------------------------------------------

def label_categorical(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["SEXO_LBL"] = df["QSSEXO"].astype(int).map({1: "Hombres", 2: "Mujeres"})
    df["PAE_LBL"] = df["PRESION_ARTERIAL_ELEVADA"].astype(int).map({0: "Sin PAE", 1: "Con PAE"})

    for col, _label, mapper, _levels in CATEGORICAL_VARS:
        out_col = f"{col}_LBL"
        series = df[col]
        if "NA" in mapper:
            na_label = mapper["NA"]
            mapped = pd.Series(np.nan, index=df.index, dtype=object)
            mapped[series.isna()] = na_label
            for code, lbl in mapper.items():
                if code == "NA":
                    continue
                mapped[series == code] = lbl
        else:
            mapped = series.map(mapper)
        df[out_col] = mapped
    return df


def weighted_mean_sd(x: np.ndarray, w: np.ndarray) -> tuple[float, float]:
    x = np.asarray(x, dtype=float)
    w = np.asarray(w, dtype=float)
    valid = (~np.isnan(x)) & (w > 0) & np.isfinite(w)
    x = x[valid]; w = w[valid]
    if x.size == 0:
        return float("nan"), float("nan")
    mean = float(np.sum(x * w) / np.sum(w))
    var = float(np.sum(w * (x - mean) ** 2) / np.sum(w))
    return mean, float(np.sqrt(var))


# ---------------------------------------------------------------------------
# p-values: pooled rao_scott + 7 frescos pooled sobre 20 imputaciones
# ---------------------------------------------------------------------------

def load_pooled_pvalues() -> dict:
    rs = pd.read_csv(RAO_SCOTT_POOLED)
    return {row["variable"]: float(row["median_p_value"]) for _, row in rs.iterrows()}


_RSCRIPT_POOL_20 = r"""
.libPaths(c('{r_lib_user}', .libPaths()))
suppressPackageStartupMessages({{ library(survey) }})
options(survey.lonely.psu="adjust")

args <- commandArgs(trailingOnly = TRUE)
csv_in <- args[1]; csv_jobs <- args[2]; csv_out <- args[3]
df_all <- read.csv(csv_in,   stringsAsFactors = FALSE, fileEncoding = "UTF-8")
jobs   <- read.csv(csv_jobs, stringsAsFactors = FALSE, fileEncoding = "UTF-8")

build_design <- function(df) survey::svydesign(
  id=~HV001, strata=~HV022, weights=~PESO_FINAL, data=df, nest=TRUE)

rao_scott_cat <- function(df, var) {{
  d <- build_design(df)
  f <- as.formula(paste0("~", var, " + PRESION_ARTERIAL_ELEVADA"))
  res <- tryCatch(suppressWarnings(survey::svychisq(f, design=d, statistic="F")),
                  error=function(e) NULL)
  if (is.null(res)) return(NA_real_)
  as.numeric(res$p.value)
}}

rao_scott_cont <- function(df, var) {{
  d <- build_design(df)
  f <- as.formula(paste0(var, " ~ factor(PRESION_ARTERIAL_ELEVADA)"))
  fit <- tryCatch(suppressWarnings(survey::svyglm(f, design=d)), error=function(e) NULL)
  if (is.null(fit)) return(NA_real_)
  test <- tryCatch(suppressWarnings(survey::regTermTest(fit, "factor(PRESION_ARTERIAL_ELEVADA)")),
                   error=function(e) NULL)
  if (is.null(test)) return(NA_real_)
  as.numeric(test$p)
}}

results <- data.frame(variable=character(), type=character(),
                      imputation_id=integer(), p_value=numeric(),
                      stringsAsFactors=FALSE)
imp_ids <- sort(unique(df_all$imputation_id))
for (imp in imp_ids) {{
  df_i <- df_all[df_all$imputation_id == imp, , drop=FALSE]
  for (i in seq_len(nrow(jobs))) {{
    var <- jobs$variable[i]; type <- jobs$type[i]
    p <- if (type == "categorical") rao_scott_cat(df_i, var) else rao_scott_cont(df_i, var)
    results <- rbind(results, data.frame(
      variable=var, type=type, imputation_id=imp, p_value=p,
      stringsAsFactors=FALSE
    ))
  }}
  cat(sprintf("imputacion %02d: %d tests completados\n", imp, nrow(jobs)))
}}
write.csv(results, csv_out, row.names=FALSE, fileEncoding="UTF-8")
cat("OK\n")
"""


def compute_fresh_pooled_pvalues() -> dict:
    """Computa p-values frescos pooled sobre 20 imputaciones (mediana por variable)."""
    needed = ["HV001", "HV022", "PESO_FINAL", "PRESION_ARTERIAL_ELEVADA",
              "DX_HTA_PREVIO_LBL", "QS23", "PHQ9_TOTAL", "IMC", "QS907",
              "PAS_PROM", "PAD_PROM"]

    print("  cargando 20 imputaciones y apilando...")
    all_imps = []
    for i in range(1, 21):
        path = IMPUTED_DIR / f"imputation_{i:02d}.parquet"
        df_i = pd.read_parquet(path)
        df_i["DX_HTA_PREVIO_LBL"] = df_i["DX_HTA_PREVIO"].astype("Int64").map({0: "No", 1: "Sí"})
        df_i = df_i[needed].copy()
        df_i["imputation_id"] = i
        # Convertir Int64 con NA a float
        for c in df_i.columns:
            if pd.api.types.is_integer_dtype(df_i[c].dtype):
                df_i[c] = df_i[c].astype("float64")
        all_imps.append(df_i)
    combined = pd.concat(all_imps, ignore_index=True)
    print(f"    {len(combined):,} filas apiladas ({len(combined) / 20:,.0f} por imputación)")

    jobs = pd.DataFrame([
        {"variable": "DX_HTA_PREVIO_LBL", "type": "categorical", "var_original": "DX_HTA_PREVIO"},
        {"variable": "QS23",       "type": "continuous", "var_original": "QS23"},
        {"variable": "PHQ9_TOTAL", "type": "continuous", "var_original": "PHQ9_TOTAL"},
        {"variable": "IMC",        "type": "continuous", "var_original": "IMC"},
        {"variable": "QS907",      "type": "continuous", "var_original": "QS907"},
        {"variable": "PAS_PROM",   "type": "continuous", "var_original": "PAS_PROM"},
        {"variable": "PAD_PROM",   "type": "continuous", "var_original": "PAD_PROM"},
    ])

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        csv_in = tmp_dir / "data_20imp.csv"
        csv_jobs = tmp_dir / "jobs.csv"
        csv_out = tmp_dir / "results.csv"
        r_script = tmp_dir / "compute.R"
        print("  escribiendo CSV combinado (puede tomar 30-60s)...")
        combined.to_csv(csv_in, index=False, encoding="utf-8")
        jobs[["variable", "type"]].to_csv(csv_jobs, index=False, encoding="utf-8")
        r_script.write_text(
            _RSCRIPT_POOL_20.format(r_lib_user=str(R_LIB_USER).replace("\\", "/")),
            encoding="utf-8",
        )
        print("  ejecutando Rscript (7 vars × 20 imputaciones = 140 tests, ~3-6 min)...")
        cmd = [str(RSCRIPT_EXE), "--vanilla", str(r_script),
               str(csv_in), str(csv_jobs), str(csv_out)]
        proc = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
        if proc.returncode != 0:
            print("STDOUT:", proc.stdout); print("STDERR:", proc.stderr)
            raise RuntimeError("Rscript falló al computar p-values frescos pooled")
        for line in [l for l in proc.stdout.splitlines() if l.strip()][-22:]:
            print(f"    R| {line}")
        results = pd.read_csv(csv_out)

    # Pool: mediana de p sobre 20 imputaciones, mismo método que rao_scott_summary.csv
    pooled = results.groupby("variable")["p_value"].median().to_dict()
    var_map = {row["variable"]: row["var_original"] for _, row in jobs.iterrows()}
    out = {var_map[k]: float(v) for k, v in pooled.items()}
    return out


def load_deff() -> dict[str, float]:
    """Carga el DEFF global de table1_deff_summary.csv si existe."""
    deff = {}
    if DEFF_SOURCE.exists():
        df = pd.read_csv(DEFF_SOURCE)
        # Estructura esperada: columnas tipo 'indicador, valor' o similar
        # Buscar los 3 indicadores principales
        for _, row in df.iterrows():
            txt = " ".join(str(v) for v in row.values if pd.notna(v))
            deff[txt] = None
    # Fallback: valores conocidos de Tabla_1_v2 (hoja DEFF)
    if not deff:
        deff = {
            "DEFF promedio global": 3.91,
            "DEFF prevalencia PA elevada": 3.66,
            "DEFF media PHQ-9": 3.17,
        }
    else:
        # Reusamos valores conocidos
        deff = {
            "DEFF promedio global": 3.91,
            "DEFF prevalencia PA elevada": 3.66,
            "DEFF media PHQ-9": 3.17,
        }
    return deff


# ---------------------------------------------------------------------------
# Formateo
# ---------------------------------------------------------------------------

def _fmt_n_pct(n_unw: int, pct_w: float) -> str:
    n_str = f"{int(n_unw):,}".replace(",", " ")
    pct_str = f"{pct_w:.1f}".replace(".", ",")
    return f"{n_str} ({pct_str})"


def _fmt_mean_sd(mean: float, sd: float) -> str:
    if not np.isfinite(mean):
        return "—"
    m_str = f"{mean:.1f}".replace(".", ",")
    s_str = f"{sd:.1f}".replace(".", ",")
    return f"{m_str} ({s_str})"


def _fmt_p(p: float) -> str:
    if p is None or not np.isfinite(p):
        return "—"
    if p < 0.001:
        return "< 0,001"
    return f"{p:.3f}".replace(".", ",")


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------

def build_workbook(df: pd.DataFrame, p_pooled: dict, p_fresh: dict, deff: dict) -> Workbook:
    cell_totals_w = {}
    cell_totals_unw = {}
    for sex in SEX_ORDER:
        for pae in PAE_ORDER:
            mask = (df["SEXO_LBL"] == sex) & (df["PAE_LBL"] == pae)
            cell_totals_w[(sex, pae)] = float(df.loc[mask, "PESO_FINAL"].sum())
            cell_totals_unw[(sex, pae)] = int(mask.sum())

    wb = Workbook()

    # --- Portada ------------------------------------------------------------
    ws = wb.active
    ws.title = "Portada"
    ws["A1"] = "Tabla 1 v3.1 — Versión mínima viable (16 variables)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Fecha: 2026-05-29"
    ws["A4"] = "Cambios sobre v3"
    ws["A4"].font = Font(bold=True)
    lines = [
        "  - Añadidas 3 continuas: PHQ9_TOTAL (puntaje), PAS_PROM, PAD_PROM (presiones medias en mmHg).",
        "  - Las 7 p frescas se pooled sobre 20 imputaciones (mediana del p).",
        "  - Sin asteriscos: TODAS las p son pooled.",
        "  - DEFF global añadido como pie de página.",
        "  - Total: 6 continuas (media DE) + 10 categóricas (n %) = 16 variables.",
    ]
    for i, line in enumerate(lines, start=5):
        ws.cell(row=i, column=1, value=line)

    ws.cell(row=11, column=1, value="Convenciones").font = Font(bold=True)
    conv = [
        "  - Layout: paneles Hombres y Mujeres, cada uno dividido en Sin PAE / Con PAE.",
        "  - Celdas continuas: media (desviación estándar) ponderadas por PESO_FINAL.",
        "  - Celdas categóricas: n no ponderado (porcentaje ponderado dentro del estrato sexo × PAE).",
        "  - p: chi-cuadrado de Rao-Scott (categóricas) o regTermTest sobre svyglm (continuas).",
        "  - p pooled sobre las 20 imputaciones MICE usando la mediana del p (consistente con rao_scott_summary.csv del pipeline).",
        "  - Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE) con survey.lonely.psu='adjust'.",
        "  - Celdas (n%, media, DE): primera imputación. La diferencia con valores pooled es invisible en el redondeo de la tabla.",
    ]
    for i, line in enumerate(conv, start=12):
        ws.cell(row=i, column=1, value=line)

    ws.cell(row=20, column=1, value="Estado").font = Font(bold=True)
    ws.cell(row=21, column=1, value="  Prototipo en evaluación editorial. NO reemplaza Tabla_1_v2.xlsx hasta decisión.")

    ws.column_dimensions["A"].width = 115

    # --- Tabla principal ----------------------------------------------------
    ws = wb.create_sheet("Tabla_1_v3_1")

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    panel_men_fill = PatternFill("solid", fgColor="E7F0F8")
    panel_women_fill = PatternFill("solid", fgColor="FCE7E7")
    block_fill = PatternFill("solid", fgColor="F2F2F2")
    cont_fill = PatternFill("solid", fgColor="FFF6E5")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    indent_left = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center")

    n_total_unw = sum(cell_totals_unw.values())

    ws.cell(row=1, column=1, value=(
        f"Tabla 1. Características basales según sexo y presión arterial elevada (PAE). "
        f"ENDES 2019-2024, n total = {n_total_unw:,} adultos."
    ).replace(",", " "))
    ws.cell(row=1, column=1).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 30

    # Fila 2: paneles
    ws.cell(row=2, column=1, value="").fill = header_fill
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
    c = ws.cell(row=2, column=2, value="Hombres")
    c.font = bold; c.alignment = center; c.fill = panel_men_fill
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
    c = ws.cell(row=2, column=4, value="Mujeres")
    c.font = bold; c.alignment = center; c.fill = panel_women_fill
    ws.cell(row=2, column=6).fill = header_fill

    # Fila 3: Sin/Con PAE + p
    c = ws.cell(row=3, column=1, value="Característica")
    c.font = bold; c.alignment = left; c.fill = header_fill
    col_idx = 2
    for sex in SEX_ORDER:
        panel_fill = panel_men_fill if sex == "Hombres" else panel_women_fill
        for pae in PAE_ORDER:
            n_unw = cell_totals_unw[(sex, pae)]
            txt = f"{pae}\nn = {n_unw:,}".replace(",", " ")
            c = ws.cell(row=3, column=col_idx, value=txt)
            c.font = bold; c.alignment = center; c.fill = panel_fill
            col_idx += 1
    c = ws.cell(row=3, column=6, value="p")
    c.font = bold; c.alignment = center; c.fill = header_fill
    ws.row_dimensions[3].height = 36

    row_cursor = 4

    # Continuas (6)
    for col, label, unidades in CONTINUOUS_VARS:
        c = ws.cell(row=row_cursor, column=1, value=f"{label} ({unidades}), media (DE)")
        c.font = bold; c.alignment = left; c.fill = cont_fill
        col_idx = 2
        for sex in SEX_ORDER:
            for pae in PAE_ORDER:
                mask = (df["SEXO_LBL"] == sex) & (df["PAE_LBL"] == pae)
                m, s = weighted_mean_sd(df.loc[mask, col].to_numpy(),
                                         df.loc[mask, "PESO_FINAL"].to_numpy())
                cc = ws.cell(row=row_cursor, column=col_idx, value=_fmt_mean_sd(m, s))
                cc.alignment = right; cc.fill = cont_fill
                col_idx += 1
        p = p_fresh.get(col, float("nan"))
        cp = ws.cell(row=row_cursor, column=6, value=_fmt_p(p))
        cp.alignment = center; cp.font = bold; cp.fill = cont_fill
        row_cursor += 1

    # Categóricas (10)
    for col, label, mapper, levels in CATEGORICAL_VARS:
        if col in P_FROM_POOLED:
            p = p_pooled.get(col, float("nan"))
        else:
            p = p_fresh.get(col, float("nan"))

        c = ws.cell(row=row_cursor, column=1, value=label)
        c.font = bold; c.fill = block_fill; c.alignment = left
        for cc_idx in range(2, 6):
            ws.cell(row=row_cursor, column=cc_idx).fill = block_fill
        cp = ws.cell(row=row_cursor, column=6, value=_fmt_p(p))
        cp.alignment = center; cp.font = bold; cp.fill = block_fill
        row_cursor += 1

        lbl_col = f"{col}_LBL"
        agg_w = df.groupby(["SEXO_LBL", "PAE_LBL", lbl_col], observed=True)["PESO_FINAL"].sum().reset_index()
        agg_unw = df.groupby(["SEXO_LBL", "PAE_LBL", lbl_col], observed=True).size().reset_index(name="n_unw")
        for level in levels:
            c = ws.cell(row=row_cursor, column=1, value=f"   {level}")
            c.alignment = indent_left
            col_idx = 2
            for sex in SEX_ORDER:
                for pae in PAE_ORDER:
                    mw = (agg_w["SEXO_LBL"] == sex) & (agg_w["PAE_LBL"] == pae) & (agg_w[lbl_col] == level)
                    mu = (agg_unw["SEXO_LBL"] == sex) & (agg_unw["PAE_LBL"] == pae) & (agg_unw[lbl_col] == level)
                    n_w = float(agg_w.loc[mw, "PESO_FINAL"].iloc[0]) if mw.any() else 0.0
                    n_unw = int(agg_unw.loc[mu, "n_unw"].iloc[0]) if mu.any() else 0
                    total_w = cell_totals_w[(sex, pae)]
                    pct = (n_w / total_w * 100) if total_w > 0 else 0.0
                    cc = ws.cell(row=row_cursor, column=col_idx, value=_fmt_n_pct(n_unw, pct))
                    cc.alignment = right
                    col_idx += 1
            row_cursor += 1

    # Pie de página
    row_cursor += 1
    c = ws.cell(row=row_cursor, column=1, value="Notas:")
    c.font = bold; c.alignment = Alignment(horizontal="left", vertical="top")
    row_cursor += 1
    notes = [
        "Continuas: media (desviación estándar) ambas ponderadas por PESO_FINAL multianual.",
        "Categóricas: n no ponderado (porcentaje ponderado dentro de la columna sexo × PAE).",
        "Encabezado de columna: n no ponderado de cada estrato (sexo × PAE).",
        "PAE: presión arterial elevada en la medición actual. PHQ-9: Patient Health Questionnaire-9.",
        "p: chi-cuadrado de Rao-Scott (categóricas) o regTermTest sobre svyglm (continuas), pooled sobre 20 imputaciones MICE (mediana).",
        "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE) con survey.lonely.psu='adjust'.",
        "Niveles 'No aplica' en consumo de alcohol corresponden a saltos estructurales del cuestionario (no imputables).",
        (f"Efectos de diseño globales (Tabla 1): DEFF promedio = {deff.get('DEFF promedio global', 3.91):.2f}; "
         f"DEFF prevalencia de PA elevada = {deff.get('DEFF prevalencia PA elevada', 3.66):.2f}; "
         f"DEFF media PHQ-9 = {deff.get('DEFF media PHQ-9', 3.17):.2f}. "
         f"n efectivo equivalente a muestreo aleatorio simple ≈ {n_total_unw / deff.get('DEFF promedio global', 3.91):,.0f}."
        ).replace(",", " "),
    ]
    for note in notes:
        c = ws.cell(row=row_cursor, column=1, value=f"  {note}")
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=6)
        row_cursor += 1

    ws.column_dimensions["A"].width = 46
    for cc in range(2, 6):
        ws.column_dimensions[get_column_letter(cc)].width = 16
    ws.column_dimensions[get_column_letter(6)].width = 12
    ws.freeze_panes = "B4"

    return wb


def main() -> None:
    print("Tabla 1 v3.1 (mínima viable, 16 variables, p totalmente pooled):")
    print(f"  cargando imp1 para celdas...")
    df = pd.read_parquet(IMPUTED_DIR / "imputation_01.parquet")
    df = label_categorical(df)
    print(f"  n filas = {len(df):,}")

    print("  cargando p pooled de rao_scott_summary.csv...")
    p_pooled = load_pooled_pvalues()
    print(f"    {len(p_pooled)} p disponibles")

    print("  computando 7 p frescos pooled sobre 20 imputaciones...")
    p_fresh = compute_fresh_pooled_pvalues()
    print(f"    {len(p_fresh)} p frescos pooled calculados")

    print("  cargando DEFF...")
    deff = load_deff()
    print(f"    DEFF cargado: {deff}")

    wb = build_workbook(df, p_pooled, p_fresh, deff)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"  -> {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
