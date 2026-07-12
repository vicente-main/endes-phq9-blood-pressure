"""Tabla 1 — versión Opción B (DATOS OBSERVADOS).

Decisión editorial (revisor 2026-06): la Tabla 1 describe la muestra REAL.
- Celdas (n %, media DE) calculadas sobre la base OBSERVADA (sin imputar).
- Se reporta el número de valores faltantes por variable.
- La imputación múltiple se reserva para los modelos multivariables.
- p-values: Rao-Scott (svychisq) / regTermTest sobre svyglm, complete-case sobre
  los DATOS OBSERVADOS (un solo dataset), NO pooled sobre imputaciones.

No sobrescribe Tabla_1_v3_1.xlsx (histórico). Salida nueva.
"""
from __future__ import annotations

import subprocess
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
BASE_PARQUET = ROOT / "data" / "output" / "endes_hta_depresion_2019_2024.parquet"
OUT_PATH = ROOT / "data" / "output" / "analysis" / "tables" / "tabla1_datos_observados.xlsx"
RSCRIPT_EXE = Path("C:/Program Files/R/R-4.5.3/bin/Rscript.exe")
R_LIB_USER = Path("C:/Users/Trabajo/Documents/R/win-library/4.5")

SEX_ORDER = ["Hombres", "Mujeres"]
PAE_ORDER = ["Sin PAE", "Con PAE"]

CATEGORICAL_VARS = [
    ("SEVERIDAD_DEPRESIVA", "Severidad PHQ-9 (exposición principal)",
     {"Minima": "Mínima (0–4)", "Leve": "Leve (5–9)", "Moderada": "Moderada (10–14)",
      "Mod_Severa": "Mod. severa (15–19)", "Severa": "Severa (20–27)"},
     ["Mínima (0–4)", "Leve (5–9)", "Moderada (10–14)", "Mod. severa (15–19)", "Severa (20–27)"]),
    ("HV025", "Área de residencia", {1: "Urbano", 2: "Rural"}, ["Urbano", "Rural"]),
    ("HV270", "Quintil de riqueza",
     {1: "Q1 — Más pobre", 2: "Q2", 3: "Q3", 4: "Q4", 5: "Q5 — Más rico"},
     ["Q1 — Más pobre", "Q2", "Q3", "Q4", "Q5 — Más rico"]),
    ("QS25N", "Nivel educativo",
     {0: "Sin educación / inicial", 1: "Sin educación / inicial", 2: "Primaria", 3: "Secundaria",
      4: "Superior no universitaria", 5: "Superior universitaria"},
     ["Sin educación / inicial", "Primaria", "Secundaria",
      "Superior no universitaria", "Superior universitaria"]),
    # Etiquetas de VIOLENCIA corregidas a la codificación real del pipeline:
    # 0 = sin violencia ; 1 = con violencia (física/psicológica) ; 2 = sin pareja
    ("VIOLENCIA_PAREJA", "Violencia de pareja (último año)",
     {0: "Sin violencia", 1: "Con violencia (física/psicológica)", 2: "Sin pareja"},
     ["Sin violencia", "Con violencia (física/psicológica)", "Sin pareja"]),
    ("ALCOHOL_PROBLEMATICO", "Consumo problemático de alcohol",
     {0: "No", 1: "Sí", "NA": "No aplica (no consume)"},
     ["No", "Sí", "No aplica (no consume)"]),
    ("QS201", "Consumo de tabaco (últimos 30 días)",
     {1: "Sí (fumó últimos 30 días)", 2: "No fumó", "NA": "No aplica (no fumó últimos 12 meses)"},
     ["Sí (fumó últimos 30 días)", "No fumó", "No aplica (no fumó últimos 12 meses)"]),
    ("QS109", "Diagnóstico de diabetes", {1: "Sí (diabetes)", 2: "No"}, ["No", "Sí (diabetes)"]),
    ("CALIDAD_DIETA", "Calidad de la dieta", {0: "No adecuada", 1: "Adecuada"}, ["No adecuada", "Adecuada"]),
    ("DX_HTA_PREVIO", "Diagnóstico previo de HTA", {0: "No", 1: "Sí"}, ["No", "Sí"]),
    ("ALTITUD_CAT3", "Altitud del conglomerado de residencia (m s.n.m.)",
     {"<1500": "< 1 500", "1500-2499": "1 500-2 499", ">=2500": "≥ 2 500"},
     ["< 1 500", "1 500-2 499", "≥ 2 500"]),
]

CONTINUOUS_VARS = [
    ("QS23", "Edad", "años"),
    ("PHQ9_TOTAL", "Puntaje PHQ-9 (exposición principal)", "puntaje"),
    ("IMC", "Índice de masa corporal", "kg/m²"),
    ("QS907", "Circunferencia abdominal", "cm"),
    ("PAS_PROM", "Presión arterial sistólica media", "mmHg"),
    ("PAD_PROM", "Presión arterial diastólica media", "mmHg"),
]

# Variables cuyos NaN son saltos estructurales mostrados como nivel "No aplica"
# (NO se cuentan como datos faltantes).
STRUCTURAL_NA = {"ALCOHOL_PROBLEMATICO", "QS201"}

DEFF = {"DEFF promedio global": 3.91, "DEFF prevalencia PA elevada": 3.66, "DEFF media PHQ-9": 3.17}


def label_categorical(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["SEXO_LBL"] = df["QSSEXO"].astype(int).map({1: "Hombres", 2: "Mujeres"})
    df["PAE_LBL"] = df["PRESION_ARTERIAL_ELEVADA"].astype(int).map({0: "Sin PAE", 1: "Con PAE"})
    for col, _lab, mapper, _lv in CATEGORICAL_VARS:
        series = df[col]
        if "NA" in mapper:
            mapped = pd.Series(np.nan, index=df.index, dtype=object)
            mapped[series.isna()] = mapper["NA"]
            for code, lbl in mapper.items():
                if code == "NA":
                    continue
                mapped[series == code] = lbl
        else:
            mapped = series.map(mapper)
        df[f"{col}_LBL"] = mapped
    return df


def missing_counts(df: pd.DataFrame) -> dict:
    out = {}
    for col, _l, _m, _lv in CATEGORICAL_VARS:
        out[col] = 0 if col in STRUCTURAL_NA else int(df[col].isna().sum())
    for col, _l, _u in CONTINUOUS_VARS:
        out[col] = int(df[col].isna().sum())
    return out


def weighted_mean_sd(x, w):
    x = np.asarray(x, dtype=float); w = np.asarray(w, dtype=float)
    valid = (~np.isnan(x)) & (w > 0) & np.isfinite(w)
    x = x[valid]; w = w[valid]
    if x.size == 0:
        return float("nan"), float("nan")
    mean = float(np.sum(x * w) / np.sum(w))
    var = float(np.sum(w * (x - mean) ** 2) / np.sum(w))
    return mean, float(np.sqrt(var))


_RSCRIPT = r"""
.libPaths(c('{r_lib_user}', .libPaths()))
suppressPackageStartupMessages({{ library(survey) }})
options(survey.lonely.psu="adjust")
args <- commandArgs(trailingOnly = TRUE)
csv_in <- args[1]; csv_jobs <- args[2]; csv_out <- args[3]
df <- read.csv(csv_in, stringsAsFactors = FALSE, fileEncoding = "UTF-8")
jobs <- read.csv(csv_jobs, stringsAsFactors = FALSE, fileEncoding = "UTF-8")
des <- function(d) survey::svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, data=d, nest=TRUE)
cat_p <- function(var) {{
  sub <- df[!is.na(df[[var]]), , drop=FALSE]
  sub[[var]] <- as.factor(sub[[var]])
  sub$PRESION_ARTERIAL_ELEVADA <- as.factor(sub$PRESION_ARTERIAL_ELEVADA)
  d <- des(sub)
  f <- as.formula(paste0("~`", var, "` + PRESION_ARTERIAL_ELEVADA"))
  r <- tryCatch(suppressWarnings(survey::svychisq(f, design=d, statistic="F")), error=function(e) NULL)
  if (is.null(r)) NA_real_ else as.numeric(r$p.value)
}}
cont_p <- function(var) {{
  sub <- df[!is.na(df[[var]]), , drop=FALSE]
  d <- des(sub)
  f <- as.formula(paste0(var, " ~ factor(PRESION_ARTERIAL_ELEVADA)"))
  fit <- tryCatch(suppressWarnings(survey::svyglm(f, design=d)), error=function(e) NULL)
  if (is.null(fit)) return(NA_real_)
  t <- tryCatch(suppressWarnings(survey::regTermTest(fit, "factor(PRESION_ARTERIAL_ELEVADA)")), error=function(e) NULL)
  if (is.null(t)) NA_real_ else as.numeric(t$p)
}}
res <- data.frame(variable=character(), p_value=numeric(), stringsAsFactors=FALSE)
for (i in seq_len(nrow(jobs))) {{
  v <- jobs$variable[i]; ty <- jobs$type[i]
  p <- if (ty == "categorical") cat_p(v) else cont_p(v)
  res <- rbind(res, data.frame(variable=v, p_value=p, stringsAsFactors=FALSE))
  cat(sprintf("%s (%s): p=%.5g\n", v, ty, p))
}}
write.csv(res, csv_out, row.names=FALSE, fileEncoding="UTF-8")
cat("OK\n")
"""


def compute_observed_pvalues(df: pd.DataFrame) -> dict:
    cat_cols = [c for c, *_ in CATEGORICAL_VARS]
    cont_cols = [c for c, *_ in CONTINUOUS_VARS]
    needed = ["HV001", "HV022", "PESO_FINAL", "PRESION_ARTERIAL_ELEVADA"] + cat_cols + cont_cols
    sub = df[needed].copy()
    for c in sub.columns:
        if pd.api.types.is_integer_dtype(sub[c].dtype):
            sub[c] = sub[c].astype("float64")
    jobs = pd.DataFrame(
        [{"variable": c, "type": "categorical"} for c in cat_cols] +
        [{"variable": c, "type": "continuous"} for c in cont_cols]
    )
    with tempfile.TemporaryDirectory() as tmp:
        t = Path(tmp)
        ci, cj, co, rs = t / "data.csv", t / "jobs.csv", t / "out.csv", t / "p.R"
        sub.to_csv(ci, index=False, encoding="utf-8")
        jobs.to_csv(cj, index=False, encoding="utf-8")
        rs.write_text(_RSCRIPT.format(r_lib_user=str(R_LIB_USER).replace("\\", "/")), encoding="utf-8")
        print("  ejecutando Rscript (16 tests sobre datos observados)...")
        proc = subprocess.run([str(RSCRIPT_EXE), "--vanilla", str(rs), str(ci), str(cj), str(co)],
                              capture_output=True, text=True, encoding="utf-8", errors="replace")
        if proc.returncode != 0:
            print("STDOUT:", proc.stdout); print("STDERR:", proc.stderr)
            raise RuntimeError("Rscript falló")
        for line in [l for l in proc.stdout.splitlines() if l.strip()][-20:]:
            print(f"    R| {line}")
        res = pd.read_csv(co)
    return {row["variable"]: float(row["p_value"]) for _, row in res.iterrows()}


def _miss_suffix(n):
    return f"  [faltantes: {n:,}]".replace(",", " ") if n > 0 else ""

def _fmt_n_pct(n, pct):
    return f"{int(n):,}".replace(",", " ") + f" ({pct:.1f})".replace(".", ",")

def _fmt_mean_sd(m, s):
    if not np.isfinite(m):
        return "—"
    return f"{m:.1f}".replace(".", ",") + f" ({s:.1f})".replace(".", ",")

def _fmt_p(p):
    if p is None or not np.isfinite(p):
        return "—"
    return "< 0,001" if p < 0.001 else f"{p:.3f}".replace(".", ",")


def build_workbook(df, pvals, miss):
    cell_w, cell_unw = {}, {}
    for sex in SEX_ORDER:
        for pae in PAE_ORDER:
            m = (df["SEXO_LBL"] == sex) & (df["PAE_LBL"] == pae)
            cell_w[(sex, pae)] = float(df.loc[m, "PESO_FINAL"].sum())
            cell_unw[(sex, pae)] = int(m.sum())
    n_total = sum(cell_unw.values())

    wb = Workbook(); ws = wb.active; ws.title = "Tabla_1_observados"
    header = PatternFill("solid", fgColor="D9E1F2"); men = PatternFill("solid", fgColor="E7F0F8")
    women = PatternFill("solid", fgColor="FCE7E7"); block = PatternFill("solid", fgColor="F2F2F2")
    cont = PatternFill("solid", fgColor="FFF6E5")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    leftA = Alignment(horizontal="left", vertical="center")
    indent = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center")

    ws.cell(row=1, column=1, value=(
        f"Tabla 1. Características basales según sexo y presión arterial elevada (PAE), "
        f"sobre datos observados. ENDES 2019-2024, n total = {n_total:,} adultos.").replace(",", " "))
    ws.cell(row=1, column=1).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1).fill = header
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
    c = ws.cell(row=2, column=2, value="Hombres"); c.font = bold; c.alignment = center; c.fill = men
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
    c = ws.cell(row=2, column=4, value="Mujeres"); c.font = bold; c.alignment = center; c.fill = women
    ws.cell(row=2, column=6).fill = header

    c = ws.cell(row=3, column=1, value="Característica"); c.font = bold; c.alignment = leftA; c.fill = header
    col = 2
    for sex in SEX_ORDER:
        pf = men if sex == "Hombres" else women
        for pae in PAE_ORDER:
            c = ws.cell(row=3, column=col, value=f"{pae}\nn = {cell_unw[(sex,pae)]:,}".replace(",", " "))
            c.font = bold; c.alignment = center; c.fill = pf; col += 1
    c = ws.cell(row=3, column=6, value="p"); c.font = bold; c.alignment = center; c.fill = header
    ws.row_dimensions[3].height = 36

    r = 4
    for col_name, label, unit in CONTINUOUS_VARS:
        c = ws.cell(row=r, column=1, value=f"{label} ({unit}), media (DE){_miss_suffix(miss[col_name])}")
        c.font = bold; c.alignment = leftA; c.fill = cont
        col = 2
        for sex in SEX_ORDER:
            for pae in PAE_ORDER:
                m = (df["SEXO_LBL"] == sex) & (df["PAE_LBL"] == pae)
                mean, sd = weighted_mean_sd(df.loc[m, col_name].to_numpy(), df.loc[m, "PESO_FINAL"].to_numpy())
                cc = ws.cell(row=r, column=col, value=_fmt_mean_sd(mean, sd)); cc.alignment = right; cc.fill = cont
                col += 1
        cp = ws.cell(row=r, column=6, value=_fmt_p(pvals.get(col_name))); cp.alignment = center; cp.font = bold; cp.fill = cont
        r += 1

    for col_name, label, mapper, levels in CATEGORICAL_VARS:
        c = ws.cell(row=r, column=1, value=f"{label}{_miss_suffix(miss[col_name])}")
        c.font = bold; c.fill = block; c.alignment = leftA
        for ci in range(2, 6):
            ws.cell(row=r, column=ci).fill = block
        cp = ws.cell(row=r, column=6, value=_fmt_p(pvals.get(col_name))); cp.alignment = center; cp.font = bold; cp.fill = block
        r += 1
        lbl = f"{col_name}_LBL"
        agg_w = df.groupby(["SEXO_LBL", "PAE_LBL", lbl], observed=True)["PESO_FINAL"].sum().reset_index()
        agg_u = df.groupby(["SEXO_LBL", "PAE_LBL", lbl], observed=True).size().reset_index(name="n")
        for level in levels:
            ws.cell(row=r, column=1, value=f"   {level}").alignment = indent
            col = 2
            for sex in SEX_ORDER:
                for pae in PAE_ORDER:
                    mw = (agg_w["SEXO_LBL"]==sex)&(agg_w["PAE_LBL"]==pae)&(agg_w[lbl]==level)
                    mu = (agg_u["SEXO_LBL"]==sex)&(agg_u["PAE_LBL"]==pae)&(agg_u[lbl]==level)
                    nw = float(agg_w.loc[mw,"PESO_FINAL"].iloc[0]) if mw.any() else 0.0
                    nu = int(agg_u.loc[mu,"n"].iloc[0]) if mu.any() else 0
                    # % ponderado sobre NO faltantes del estrato
                    denom_mask = (df["SEXO_LBL"]==sex)&(df["PAE_LBL"]==pae)&(df[lbl].notna())
                    denom = float(df.loc[denom_mask,"PESO_FINAL"].sum())
                    pct = (nw/denom*100) if denom > 0 else 0.0
                    cc = ws.cell(row=r, column=col, value=_fmt_n_pct(nu, pct)); cc.alignment = right
                    col += 1
            r += 1

    r += 1
    ws.cell(row=r, column=1, value="Notas:").font = bold; r += 1
    notes = [
        "Características descritas sobre los DATOS OBSERVADOS (sin imputación); la imputación múltiple se reservó para los modelos multivariables.",
        "Continuas: media (desviación estándar), ponderadas por PESO_FINAL. Categóricas: n no ponderado (porcentaje ponderado sobre los casos no faltantes del estrato sexo × PAE).",
        "«[faltantes: n]» indica el número de valores perdidos observados por variable. Los niveles «No aplica» (alcohol, tabaco) corresponden a saltos estructurales del cuestionario y NO se cuentan como faltantes.",
        "p: chi-cuadrado de Rao-Scott (categóricas) o regTermTest sobre svyglm (continuas), calculados por análisis de casos completos sobre los datos observados.",
        "Diseño: svydesign(id=~HV001, strata=~HV022, weights=~PESO_FINAL, nest=TRUE) con survey.lonely.psu='adjust'.",
        (f"Efecto de diseño global: DEFF promedio = {DEFF['DEFF promedio global']:.2f}; "
         f"n efectivo equivalente a muestreo aleatorio simple ≈ {n_total/DEFF['DEFF promedio global']:,.0f}.").replace(",", " "),
    ]
    for note in notes:
        c = ws.cell(row=r, column=1, value=f"  {note}"); c.font = Font(size=9)
        c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1

    ws.column_dimensions["A"].width = 52
    for cc in range(2, 6):
        ws.column_dimensions[get_column_letter(cc)].width = 16
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "B4"
    return wb


def main():
    print("Tabla 1 (datos observados, Opción B):")
    df = pd.read_parquet(BASE_PARQUET)
    print(f"  base observada: {len(df):,} filas")
    df = label_categorical(df)
    miss = missing_counts(df)
    print(f"  faltantes: { {k:v for k,v in miss.items() if v>0} }")
    pvals = compute_observed_pvalues(df)
    print(f"  p calculados: {len(pvals)}")
    wb = build_workbook(df, pvals, miss)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"  -> {OUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
