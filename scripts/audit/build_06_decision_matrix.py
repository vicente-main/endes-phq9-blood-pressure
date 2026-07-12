"""Construye 06_decision_principal_suplementario.xlsx, la matriz de decisión
del gate B en formato spreadsheet.

Lee la lista de decisiones inline y exporta a xlsx con coloreado por decisión.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data" / "output" / "Auditoria_Integral" / "06_decision_principal_suplementario.xlsx"


# (id, artefacto, fuente_actual, hallazgo_A1_A7, decision, ubicacion, justificacion)
MATRIX = [
    ("B01", "Estrategia de imputación", "MICE_TARGETS en src/endes_pipeline/analysis.py", "A1: estrategia más rigurosa que estándar ENDES (complete-case); 3 categorías de faltantes bien gestionadas", "mantener", "métodos (texto)", "Sólida; explicitar en el manuscrito con texto sugerido en A1 §4"),
    ("B02", "DAG conceptual", "(no existía) → 02_DAG_modelos.svg/.png", "A2: render funciona, acíclico (TRUE), adjustmentSet = Modelo 2", "crear_nuevo", "suplementario (Figura S2)", "Justifica de un vistazo la elección Modelo 2 vs Modelo 3"),
    ("B03", "Figura STROBE", "Para Publicar Redisenado/Principal/Figura_1_STROBE.{svg,png,pdf}", "A3: conteos correctos; faltan notas explícitas (COVID 2020, MEF reproductiva)", "regenerar_v2", "principal (Figura 1)", "Notas de exclusión MEF y sub-muestreo COVID 2020 son críticos para interpretación"),
    ("B04", "Tabla 1 (basal weighted)", "Para Publicar Redisenado/Principal/Tabla_1.xlsx", "A4 P05: categorías PHQ9_TOTAL/QS23/QS907 y grupo 'overall' en celdas publicables", "regenerar_v2", "principal", "Lector no técnico no entiende códigos crudos"),
    ("B05", "Tabla 2 (bivariado)", "Para Publicar Redisenado/Principal/Tabla_2.xlsx", "A4: encabezados y categorías ya legibles; sin tokens raw", "mantener", "principal", "Sin cambios necesarios"),
    ("B06", "Tabla 3 (modelos principales)", "Para Publicar Redisenado/Principal/Tabla_3.xlsx", "A4 P01, P11-P13: factor(ANIO)20XX, factor(HV270)2..5, factor(QS25N)1..5, interacciones con sintaxis interna", "regenerar_v2", "principal", "Bloqueante: factor() raw en tabla principal"),
    ("B07", "Tabla 4 (cascada cuidado)", "Para Publicar Redisenado/Principal/Tabla_4.xlsx", "A4 P02: factor(ANIO)20XX", "regenerar_v2", "principal", "Mismo bloqueo que Tabla 3"),
    ("B08", "Tabla 5 (sensibilidad principal)", "Para Publicar Redisenado/Principal/Tabla_5.xlsx", "A4 P16: celdas None en fila spline (menor)", "renombrar", "principal", "Reemplazar None por '—' y ajustar la fila spline"),
    ("B09", "Tabla S1 (flujo STROBE)", "Para Publicar Redisenado/Suplementario/Tabla_S1_flujo_STROBE.xlsx", "A3: conteos correctos", "mantener", "suplementario", "Sin cambios"),
    ("B10", "Tabla S2 (soporte tablas)", "Para Publicar Redisenado/Suplementario/Tabla_S2_soporte_tablas.xlsx", "A4 P06-P08: PHQ9_TOTAL/QS23/QS907, rubin_d1_chisq_approx, .tmp_pred*", "regenerar_v2", "suplementario", "Bloqueante: nombres internos R en tabla publicable"),
    ("B11", "Tabla S3 (sensibilidad/interacciones)", "Para Publicar Redisenado/Suplementario/Tabla_S3_sensibilidad_interacciones.xlsx", "A4 P03-P04: factor(ANIO)20XX", "regenerar_v2", "suplementario", "Mismo bloqueo que Tablas 3-4"),
    ("B12", "Tabla S4 (datos figuras)", "Para Publicar Redisenado/Suplementario/Tabla_S4_datos_figuras.xlsx", "A4 P09: variable=PHQ9_TOTAL (menor)", "renombrar", "suplementario", "Cambiar encabezado/etiquetado de la col 'variable'"),
    ("B13", "Figura S1 (spline)", "Para Publicar Redisenado/Suplementario/Figura_S1_Spline.{svg,png,pdf}", "A5: pred_se correcto, p estable; mejoras opcionales (eje Y 'Razón de prevalencia', truncar X en 20 o histograma marginal, documentar nodos [0,4,9,14])", "regenerar_v2", "suplementario", "Resultado robusto; mejora pedagógica de la lectura"),
    ("B14", "Datos fuente spline", "data/output/analysis/figures/spline_curve_pooled.csv", "A5: sin pred_se=0, IC reales", "mantener", "suplementario (datos)", "Datos correctos"),
    ("B15", "Figura S3 (forest robustez subgrupos)", "(no existe)", "A7: opcional; información ya está en Tabla S3 e interactions_and_sensitivity_models.csv", "no_crear", "—", "Default conservador: no añadir display adicional sin demanda editorial explícita"),
]


COLOR_BY_DECISION = {
    "mantener": "E2EFDA",
    "renombrar": "FFF2CC",
    "regenerar_v2": "FCE4D6",
    "crear_nuevo": "BDD7EE",
    "descartar": "F4B084",
    "no_crear": "D9D9D9",
}


def write_xlsx(path: Path) -> None:
    wb = Workbook()

    # --- Hoja 0: Portada ----------------------------------------------------
    ws = wb.active
    ws.title = "Portada"
    ws["A1"] = "Matriz de decisión — Gate B (Auditoría Integral)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Fecha: 2026-05-27"
    ws["A4"] = "Estructura:"
    ws["A5"] = "  1. Matriz: una fila por artefacto, con hallazgo, decisión, ubicación y justificación"
    ws["A6"] = "  2. Resumen por decisión: cuántos artefactos en cada categoría"
    ws["A7"] = "  3. Plan de ejecución C: orden de producción de v2"
    ws["A8"] = "  4. Test plan: criterios que cada v2 debe pasar"
    ws["A10"] = "Reglas:"
    ws["A11"] = "  - Solo artefactos marcados 'regenerar_v2' o 'crear_nuevo' pasan a Fase C."
    ws["A12"] = "  - Toda producción de v2 vive en data/output/Auditoria_Integral/."
    ws["A13"] = "  - Para Publicar Redisenado/ NO se sobrescribe."
    ws["A14"] = "  - Cada v2 cita la fila B0X que lo autoriza."
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 90

    # --- Hoja 1: Matriz ------------------------------------------------------
    ws = wb.create_sheet("Matriz")
    headers = ["ID", "Artefacto", "Fuente actual", "Hallazgo (A1-A7)", "Decisión", "Ubicación editorial", "Justificación"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in MATRIX:
        ws.append(row)
        last = ws.max_row
        color = COLOR_BY_DECISION.get(row[4])
        if color:
            ws.cell(row=last, column=5).fill = PatternFill("solid", fgColor=color)
        for c in range(1, 8):
            ws.cell(row=last, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, width in enumerate([6, 32, 50, 70, 16, 24, 70], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    # --- Hoja 2: Resumen ----------------------------------------------------
    ws = wb.create_sheet("Resumen")
    counts: dict[str, int] = {}
    for row in MATRIX:
        counts[row[4]] = counts.get(row[4], 0) + 1
    ws.append(["Decisión", "Cantidad", "Artefactos"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for dec, n in sorted(counts.items(), key=lambda x: -x[1]):
        ids = ", ".join(r[0] for r in MATRIX if r[4] == dec)
        ws.append([dec, n, ids])
        color = COLOR_BY_DECISION.get(dec)
        if color:
            ws.cell(row=ws.max_row, column=1).fill = PatternFill("solid", fgColor=color)
    for col_idx, width in enumerate([18, 12, 60], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Hoja 3: Plan_C ----------------------------------------------------
    ws = wb.create_sheet("Plan_C")
    ws.append(["Paso", "Artefacto v2", "ID matriz", "Esfuerzo", "Notas"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    plan_c = [
        ("C1", "Tabla 1 v2", "B04", "bajo", "Mapeo de códigos ENDES → etiquetas; 'overall' → 'General'"),
        ("C2", "Tabla 3 v2", "B06", "medio", "Mapeo factor() → etiquetas legibles; interacciones"),
        ("C3", "Tabla 4 v2", "B07", "bajo", "Reusar mapeo de C2"),
        ("C4", "Tabla S2 v2", "B10", "medio", "Reemplazar rubin_d1_chisq_approx y .tmp_pred*; reusar mapeo C2"),
        ("C5", "Tabla S3 v2", "B11", "bajo", "Reusar mapeo C2"),
        ("C6", "Tabla 5 renombrar", "B08", "trivial", "None → '—'; fila spline rephrase"),
        ("C7", "Tabla S4 renombrar", "B12", "trivial", "Encabezado col 'variable'"),
        ("C8", "STROBE v2", "B03", "medio", "Añadir notas COVID 2020 y MEF al pie; preservar conteos exactos"),
        ("C9", "Spline v2", "B13", "medio", "Eje Y 'Razón de prevalencia'; opciones X (truncar o histograma); leyenda con nodos"),
    ]
    for row in plan_c:
        ws.append(row)
    for col_idx, width in enumerate([8, 28, 14, 12, 80], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Hoja 4: Test_Plan --------------------------------------------------
    ws = wb.create_sheet("Test_Plan")
    ws.append(["Test", "Aplicable a"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    tests = [
        ("grep -E 'factor\\(|\\.tmp_pred|rubin_d1|model_[1-3]' archivo.xlsx → 0 matches", "tablas v2 (C1-C5)"),
        ("Sin códigos ENDES crudos (QS23/QS25N/QS201/QS907/QS109/HV025/HV270/QSSEXO/VIOLENCIA_PAREJA/ALCOHOL_PROBLEMATICO/CALIDAD_DIETA/PHQ9_TOTAL) en encabezados publicables", "tablas v2 publicables"),
        ("Conteos numéricos celda a celda idénticos a fuente canónica (analysis_plan_*, model_diagnostics_summary.csv)", "todas las v2"),
        ("SVG/PNG/PDF abren sin texto cortado, sin overlap de etiquetas", "figuras v2 (C8, C9)"),
        ("STROBE v2: match celda a celda con 07_figura_1_strobe_datos.csv", "C8"),
        ("Spline v2: cada punto mantiene pred_se > 0", "C9"),
        ("UTF-8 sin BOM en CSV; xlsx con caracteres especiales correctos", "todas las v2"),
        ("p con 3 decimales; estimadores con 2; porcentajes con 1; coma decimal (RPMESP)", "tablas v2"),
    ]
    for row in tests:
        ws.append(row)
    for col_idx, width in enumerate([90, 24], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Escrito: {path}")
    for dec, n in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"  {dec}: {n}")


if __name__ == "__main__":
    write_xlsx(OUT)
