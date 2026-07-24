"""Corrige terminología y numeración en los suplementos no reconstruidos."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[2]
SUBMISSION = ROOT / "ENVIO_PLOS_ONE_2019-2025"


def backup(path: Path) -> None:
    target = path.with_name(f"{path.stem}.pre_guia_2026-07-23{path.suffix}")
    if not target.exists():
        shutil.copy2(path, target)


def replace_all(path: Path, replacements: dict[str, str]) -> None:
    backup(path)
    workbook = load_workbook(path)
    counts = {source: 0 for source in replacements}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                value = cell.value
                for source, target in replacements.items():
                    if source in value:
                        value = value.replace(source, target)
                        counts[source] += 1
                cell.value = value
    final_text = "\n".join(
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    )
    missing = [
        source
        for source, count in counts.items()
        if count == 0 and replacements[source] not in final_text
    ]
    if missing:
        raise RuntimeError(f"{path.name}: no se encontraron reemplazos esperados: {missing}")
    workbook.save(path)


def main() -> None:
    replace_all(
        SUBMISSION / "S1_Table.xlsx",
        {
            "Final analytical cohort": "Final analytic sample",
            "Sensitivity cohort without 2020 (COVID)": (
                "Sensitivity sample without the 2020 survey year (COVID-19)"
            ),
        },
    )
    replace_all(
        SUBMISSION / "S5_Table.xlsx",
        {
            "each DAG edge (S2 Fig)": "each DAG edge (S1 Fig)",
            (
                "Geographic confounder incorporated into the adjustment set during the analysis "
                "(amendment 2026-06-01). Together with edge B8 (altitude → LOWER hypertension), "
                "the opposite directions induce the spurious inverse PHQ-9 ↔ elevated BP association "
                "that attenuates when adjusting for altitude."
            ): (
                "Altitude was incorporated as a geographic and contextual adjustment variable "
                "(analysis-framework amendment 2026-06-01). Together with edge B8 (altitude → "
                "lower hypertension prevalence), the hypothesized opposite directions could "
                "contribute to the inverse PHQ-9–elevated-BP pattern; the observed attenuation "
                "does not establish causal confounding."
            ),
            (
                "Geographic confounder (amendment 2026-06-01). The opposite directions of A8 "
                "(altitude → +depression) and B8 (altitude → −hypertension) are the causal basis "
                "of the spurious inverse association; when altitude enters the adjustment, the "
                "PHQ-9 effect attenuates toward the null."
            ): (
                "Geographic/contextual adjustment variable (analysis-framework amendment "
                "2026-06-01). The hypothesized opposite directions of A8 and B8 motivated "
                "adjustment. The estimate moved toward the null after adjustment, but this "
                "model contrast does not identify a causal mechanism."
            ),
            (
                "Cohort structure: the age distribution may vary from year to year due to "
                "sampling effects and differential mortality."
            ): (
                "Repeated cross-sectional sample structure: the age distribution may vary from "
                "year to year because of sampling variation and differential mortality."
            ),
            "Detection/care bias": "Health-care contact or selection (hypothesized)",
            (
                "People diagnosed with hypertension may report more depressive symptoms (stress "
                "of disease management) and have blood pressure controlled by treatment — this "
                "explains the inverse association observed in cross-sectional data."
            ): (
                "Among people with a prior diagnosis, treatment, health-care contact and selection "
                "could jointly produce an inverse cross-sectional pattern. The available timing "
                "data cannot identify any of these mechanisms."
            ),
            (
                "Explanatory hypothesis for the study's finding; it should be made explicit in "
                "the Discussion."
            ): (
                "Explanatory hypothesis only; the study does not demonstrate this mechanism."
            ),
        },
    )
    s5_path = SUBMISSION / "S5_Table.xlsx"
    workbook = load_workbook(s5_path)
    cover = workbook["Cover"]
    cover["A25"] = (
        "DAG edge labels represent theory-informed hypotheses. Model 3 is an additional-adjustment "
        "analysis and is not interpreted as estimating mediation or direct causal effects."
    )
    cover["A25"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.row_dimensions[25].height = 32
    workbook.save(s5_path)

    replace_all(
        SUBMISSION / "S6_Table.xlsx",
        {
            "Comparison of the included cohort versus those excluded from the analysis": (
                "Comparison of records included in the analytic sample with records excluded "
                "during sample selection"
            ),
            "Included = cohort passing all masks of the STROBE flow.": (
                "Included = records passing all masks of the STROBE flow."
            ),
        },
    )
    replace_all(
        SUBMISSION / "S8_Table.xlsx",
        {
            "Geographic confounder": "Geographic/contextual adjustment variable",
            "Potential mediators (Model 3, exploratory)": (
                "Additional covariates (Model 3, exploratory)"
            ),
            "Post-exposure mediator; not adjusted for.": (
                "Post-exposure variable; not included in the adjustment models."
            ),
        },
    )
    s8_path = SUBMISSION / "S8_Table.xlsx"
    workbook = load_workbook(s8_path)
    worksheet = workbook["S8 Table"]
    for row in range(17, 23):
        if worksheet.cell(row, 2).value == "Mediator":
            worksheet.cell(row, 2, "Exploratory additional covariate")
    workbook.save(s8_path)

    replace_all(
        SUBMISSION / "S9_Table.xlsx",
        {
            "Model 2 prespecified, without altitude": (
                "Model 2 defined in the analysis framework, without altitude"
            ),
            "Model 3 (exploratory, with mediators)": (
                "Model 3 (exploratory additional adjustment)"
            ),
        },
    )
    replace_all(
        SUBMISSION / "S10_Table.xlsx",
        {
            "20 MICE imputations": "20 parametric chained-equations imputations",
            "Model 3 — Exploratory (with mediators)": (
                "Model 3 — Exploratory (additional adjustment)"
            ),
            "the MAR covariables were imputed": (
                "Nine targets were declared; five covariates with missing values were imputed"
            ),
            "(potential mediators).": (
                "(additional exploratory covariates; not a mediation model)."
            ),
            (
                "Altitude incorporated into the structural adjustment as a geographic confounder."
            ): (
                "Altitude was included as a geographic and contextual adjustment variable; "
                "the resulting attenuation is not interpreted causally."
            ),
        },
    )
    replace_all(
        SUBMISSION / "S11_Table.xlsx",
        {
            "20 MICE imputations": "20 parametric chained-equations imputations",
            "The MAR covariables were imputed": (
                "Nine targets were declared; five covariates with missing values were imputed"
            ),
        },
    )
    print("Actualizados: S1, S5, S6, S8, S9, S10 y S11 Tables.")


if __name__ == "__main__":
    main()
