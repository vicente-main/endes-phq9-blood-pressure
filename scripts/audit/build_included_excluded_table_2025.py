"""Tabla suplementaria S6 (cohorte 2019-2025): comparacion de la cohorte incluida
final vs los excluidos del analisis (STROBE).

Variante 2019-2025 de build_included_excluded_table.py. Reutiliza las funciones de
reconstruccion/formato del script original y solo cambia: configuracion (config
2025, divisor 7, anios 2019-2025), directorios de QC/salida y el rotulo del anio.

Comparacion descriptiva NO ponderada (convencion STROBE para no respuesta), igual
que la S6 original.

Salida: FIXES_RONDA_2026-07-09/S6_Table.xlsx (+ .csv)
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts" / "audit"))

from endes_pipeline import pipeline as P  # noqa: E402
import build_included_excluded_table as S6  # noqa: E402  (reutiliza helpers)

QC_DIR = ROOT / "data" / "output_2025" / "qc"
OUT_DIR = ROOT / "FIXES_RONDA_2026-07-09"
YEAR_LABEL = "2019-2025"


def _write_xlsx_2025(table: pd.DataFrame, n_inc: int, n_exc: int, path: Path) -> None:
    """Igual que S6._write_xlsx pero con el rotulo de anio 2019-2025."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "S6 Table"

    header_fill = PatternFill("solid", fgColor=S6.FILL_HEADER)
    section_fill = PatternFill("solid", fgColor=S6.FILL_SECTION)
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    title = (f"S6 Table. Comparacion de la cohorte incluida frente a los excluidos "
             f"del analisis (ENDES {YEAR_LABEL})")
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    headers = ["Caracteristica", "Nivel",
               f"Incluidos\n(n = {n_inc:,})".replace(",", " "),
               f"Excluidos del analisis\n(n = {n_exc:,})".replace(",", " "),
               "p"]
    for col, htext in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=htext)
        c.fill = header_fill
        c.font = bold
        c.alignment = center
        c.border = border

    r = 3
    for _, row in table.iterrows():
        is_section = bool(row["_is_section"])
        values = [row["Caracteristica"], row["Nivel"], row["Incluidos"],
                  row["Excluidos del analisis"], row["p"]]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            c.alignment = left if col <= 2 else center
            if is_section:
                c.fill = section_fill
                if col in (1, 5):
                    c.font = bold
        r += 1

    footer = (
        f"Fuente: ENDES {YEAR_LABEL} (INEI), modulo CSALUD01. Comparacion descriptiva NO ponderada "
        "(convencion STROBE para analisis de no respuesta). Continuas: media (DE); categoricas: n (% sobre no faltantes). "
        "p: t de Welch (continuas) o chi-cuadrado de Pearson (categoricas). "
        "Incluidos = cohorte que pasa todas las mascaras del flujo STROBE. "
        "Excluidos del analisis = alcanzan la base estructural pero se excluyen por PHQ-9 no disponible o presion arterial no valida. "
        "Los conteos reproducen filter_flow_by_year.csv y final_counts_by_year.csv (run 2019-2025)."
    )
    ws.cell(row=r + 1, column=1, value=footer).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    ws.row_dimensions[r + 1].height = 90

    widths = [26, 22, 20, 24, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A3"
    wb.save(path)


def main() -> None:
    cfg = json.loads((ROOT / "config" / "pipeline_config_2025.json").read_text(encoding="utf-8-sig"))
    years = cfg["years"]
    divisor = cfg["analysis"]["combined_year_divisor"]
    archives = P._find_outer_archives(ROOT / cfg["paths"]["raw_archives_dir"], years)

    frames = []
    for year in years:
        print(f"  reconstruyendo {year} ...")
        frames.append(S6._build_harmonized_with_flags(year, archives[year], divisor))
    full = pd.concat(frames, ignore_index=True)

    inc = full[full["_incluido"]]
    exc = full[full["_estructural"] & ~full["_incluido"]]

    n_inc = len(inc)
    n_exc = len(exc)
    flow = pd.read_csv(QC_DIR / "filter_flow_by_year.csv")
    expected_exc = int(flow["drop_phq_available"].sum() + flow["drop_bp_available"].sum())
    final_counts = pd.read_csv(QC_DIR / "final_counts_by_year.csv")
    expected_inc = int(final_counts["n_final"].sum())
    print(f"  incluidos={n_inc} (esperado {expected_inc}) | excluidos_analisis={n_exc} (esperado {expected_exc})")
    assert n_inc == expected_inc, f"Incluidos {n_inc} != STROBE {expected_inc}"
    assert n_exc == expected_exc, f"Excluidos {n_exc} != STROBE {expected_exc}"

    rows = []
    rows.append(S6._cont_row("EDAD", "Edad (anios)", inc, exc))
    rows += S6._cat_rows("QSSEXO", "Sexo", S6.SEX_MAP, inc, exc)
    rows += S6._cat_rows("HV025", "Area de residencia", S6.AREA_MAP, inc, exc)
    rows += S6._cat_rows("HV270", "Quintil de riqueza", S6.RIQUEZA_MAP, inc, exc)
    rows += S6._cat_rows("ANIO", "Anio de encuesta", {y: str(y) for y in years}, inc, exc)
    rows += S6._cat_rows("ALTITUD_CAT3", "Altitud (m s.n.m.)",
                         {"<1500": "< 1500", "1500-2499": "1500-2499", ">=2500": ">= 2500"}, inc, exc)
    table = pd.DataFrame(rows)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _write_xlsx_2025(table, n_inc, n_exc, OUT_DIR / "S6_Table.xlsx")
    table.drop(columns=["_is_section"]).to_csv(
        OUT_DIR / "S6_Table.csv", index=False, encoding="utf-8-sig"
    )
    print(f"  escrito: {OUT_DIR / 'S6_Table.xlsx'}")
    # eco del contenido para verificacion
    with pd.option_context("display.max_rows", None, "display.width", 200):
        print(table.drop(columns=["_is_section"]).to_string(index=False))


if __name__ == "__main__":
    main()
